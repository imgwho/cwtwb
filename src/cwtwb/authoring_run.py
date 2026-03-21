"""Run-based guided authoring helpers for the MCP dashboard workflow."""

from __future__ import annotations

import hashlib
import json
import xml.etree.ElementTree as ET
from copy import deepcopy
from datetime import date, datetime
from pathlib import Path
from typing import Any
from zipfile import ZipFile

import xlrd

from .authoring_contract import review_authoring_contract_payload, suggest_profile_matches
from .config import DEFAULT_AUTHORING_RUNS_DIR
from .connections import infer_tableau_semantic_role, inspect_hyper_schema

try:
    import yaml
except ImportError:  # pragma: no cover - optional dependency
    yaml = None

RUN_INDEX_NAME = "index.json"
MANIFEST_NAME = "manifest.json"
APPROVALS_NAME = "approvals.json"

SCHEMA_STAGE = "schema"
ANALYSIS_STAGE = "analysis"
CONTRACT_STAGE = "contract"
WIREFRAME_STAGE = "wireframe"
EXECUTION_STAGE = "execution_plan"
CONFIRMABLE_STAGES = {
    SCHEMA_STAGE,
    ANALYSIS_STAGE,
    CONTRACT_STAGE,
    WIREFRAME_STAGE,
    EXECUTION_STAGE,
}

STATUS_INITIALIZED = "initialized"
STATUS_SCHEMA_INTAKED = "schema_intaked"
STATUS_SCHEMA_CONFIRMED = "schema_confirmed"
STATUS_ANALYSIS_BUILT = "analysis_built"
STATUS_ANALYSIS_FINALIZED = "analysis_finalized"
STATUS_ANALYSIS_CONFIRMED = "analysis_confirmed"
STATUS_CONTRACT_DRAFTED = "contract_drafted"
STATUS_CONTRACT_REVIEWED = "contract_reviewed"
STATUS_CONTRACT_FINALIZED = "contract_finalized"
STATUS_CONTRACT_CONFIRMED = "contract_confirmed"
STATUS_WIREFRAME_BUILT = "wireframe_built"
STATUS_WIREFRAME_FINALIZED = "wireframe_finalized"
STATUS_WIREFRAME_CONFIRMED = "wireframe_confirmed"
STATUS_EXECUTION_PLANNED = "execution_planned"
STATUS_EXECUTION_CONFIRMED = "execution_confirmed"
STATUS_GENERATION_STARTED = "workbook_generation_started"
STATUS_GENERATION_FAILED = "workbook_generation_failed"
STATUS_GENERATED = "workbook_generated"
STATUS_VALIDATED = "validated"
STATUS_ANALYZED = "analyzed"

SUPPORTED_EXCEL_SUFFIXES = {".xls", ".xlsx", ".xlsm"}
SUPPORTED_DATASOURCE_SUFFIXES = SUPPORTED_EXCEL_SUFFIXES | {".hyper"}

ARTIFACT_SCHEMA = "schema_summary"
ARTIFACT_ANALYSIS_BRIEF = "analysis_brief"
ARTIFACT_CONTRACT_DRAFT = "contract_draft"
ARTIFACT_CONTRACT_REVIEW = "contract_review"
ARTIFACT_CONTRACT_FINAL = "contract_final"
ARTIFACT_WIREFRAME = "wireframe"
ARTIFACT_EXECUTION_PLAN = "execution_plan"
ARTIFACT_SEMANTIC_VALIDATION = "semantic_validation"
ARTIFACT_VALIDATION = "validation_report"
ARTIFACT_ANALYSIS = "analysis_report"

STAGE_ARTIFACT_MAP = {
    SCHEMA_STAGE: ARTIFACT_SCHEMA,
    ANALYSIS_STAGE: ARTIFACT_ANALYSIS_BRIEF,
    CONTRACT_STAGE: ARTIFACT_CONTRACT_FINAL,
    WIREFRAME_STAGE: ARTIFACT_WIREFRAME,
    EXECUTION_STAGE: ARTIFACT_EXECUTION_PLAN,
}

ARTIFACT_KEYS = (
    ARTIFACT_SCHEMA,
    ARTIFACT_ANALYSIS_BRIEF,
    ARTIFACT_CONTRACT_DRAFT,
    ARTIFACT_CONTRACT_REVIEW,
    ARTIFACT_CONTRACT_FINAL,
    ARTIFACT_WIREFRAME,
    ARTIFACT_EXECUTION_PLAN,
    ARTIFACT_SEMANTIC_VALIDATION,
    ARTIFACT_VALIDATION,
    ARTIFACT_ANALYSIS,
)

REOPEN_STATUS_ALLOWED = {
    ANALYSIS_STAGE: (
        STATUS_ANALYSIS_CONFIRMED,
        STATUS_CONTRACT_DRAFTED,
        STATUS_CONTRACT_REVIEWED,
        STATUS_CONTRACT_FINALIZED,
        STATUS_CONTRACT_CONFIRMED,
        STATUS_WIREFRAME_BUILT,
        STATUS_WIREFRAME_FINALIZED,
        STATUS_WIREFRAME_CONFIRMED,
        STATUS_EXECUTION_PLANNED,
        STATUS_EXECUTION_CONFIRMED,
        STATUS_GENERATION_FAILED,
        STATUS_GENERATED,
        STATUS_VALIDATED,
        STATUS_ANALYZED,
    ),
    CONTRACT_STAGE: (
        STATUS_CONTRACT_CONFIRMED,
        STATUS_WIREFRAME_BUILT,
        STATUS_WIREFRAME_FINALIZED,
        STATUS_WIREFRAME_CONFIRMED,
        STATUS_EXECUTION_PLANNED,
        STATUS_EXECUTION_CONFIRMED,
        STATUS_GENERATION_FAILED,
        STATUS_GENERATED,
        STATUS_VALIDATED,
        STATUS_ANALYZED,
    ),
    WIREFRAME_STAGE: (
        STATUS_WIREFRAME_CONFIRMED,
        STATUS_EXECUTION_PLANNED,
        STATUS_EXECUTION_CONFIRMED,
        STATUS_GENERATION_FAILED,
        STATUS_GENERATED,
        STATUS_VALIDATED,
        STATUS_ANALYZED,
    ),
    EXECUTION_STAGE: (
        STATUS_EXECUTION_CONFIRMED,
        STATUS_GENERATION_FAILED,
        STATUS_GENERATED,
        STATUS_VALIDATED,
        STATUS_ANALYZED,
    ),
}

DOWNSTREAM_ARTIFACTS = {
    ANALYSIS_STAGE: (
        ARTIFACT_CONTRACT_DRAFT,
        ARTIFACT_CONTRACT_REVIEW,
        ARTIFACT_CONTRACT_FINAL,
        ARTIFACT_WIREFRAME,
        ARTIFACT_EXECUTION_PLAN,
        ARTIFACT_SEMANTIC_VALIDATION,
        ARTIFACT_VALIDATION,
        ARTIFACT_ANALYSIS,
    ),
    CONTRACT_STAGE: (
        ARTIFACT_WIREFRAME,
        ARTIFACT_EXECUTION_PLAN,
        ARTIFACT_SEMANTIC_VALIDATION,
        ARTIFACT_VALIDATION,
        ARTIFACT_ANALYSIS,
    ),
    WIREFRAME_STAGE: (
        ARTIFACT_EXECUTION_PLAN,
        ARTIFACT_SEMANTIC_VALIDATION,
        ARTIFACT_VALIDATION,
        ARTIFACT_ANALYSIS,
    ),
    EXECUTION_STAGE: (
        ARTIFACT_SEMANTIC_VALIDATION,
        ARTIFACT_VALIDATION,
        ARTIFACT_ANALYSIS,
    ),
}

SCOPE_CHANGE_NOTE_PREFIXES = (
    "add ",
    "add a ",
    "add an ",
    "add another ",
    "additional ",
    "include ",
    "include a ",
    "include another ",
    "new ",
)
SCOPE_CHANGE_NOTE_OBJECTS = (
    "worksheet",
    "sheet",
    "chart",
    "view",
)

EXECUTION_STEP_WHITELIST = (
    "create_workbook",
    "open_workbook",
    "add_calculated_field",
    "add_parameter",
    "add_worksheet",
    "configure_chart",
    "configure_dual_axis",
    "configure_chart_recipe",
    "configure_worksheet_style",
    "add_dashboard",
    "add_dashboard_action",
    "set_worksheet_caption",
    "set_excel_connection",
    "set_mysql_connection",
    "set_tableauserver_connection",
    "set_hyper_connection",
)

POST_CHECK_WHITELIST = ("validate_workbook", "analyze_twb")

AUTHORING_MODE_AGENT_FIRST = "agent_first"
AUTHORING_MODE_LEGACY = "legacy"
AUTHORING_MODES = {AUTHORING_MODE_AGENT_FIRST, AUTHORING_MODE_LEGACY}

KNOWN_CALCULATED_FORMULAS = {
    "Profit Ratio": "SUM([Profit])/SUM([Sales])",
}
KNOWN_CALCULATED_FORMULA_LOOKUP = {
    " ".join(name.casefold().replace("_", " ").replace("-", " ").split()): name
    for name in KNOWN_CALCULATED_FORMULAS
}
EXPRESSION_FUNCTION_PREFIXES = (
    "sum(",
    "avg(",
    "count(",
    "countd(",
    "min(",
    "max(",
    "median(",
    "attr(",
    "month(",
    "quarter(",
    "year(",
    "week(",
    "day(",
    "hour(",
    "minute(",
    "second(",
    "date(",
    "datetime(",
    "dateadd(",
    "datediff(",
    "datetrunc(",
    "dateparse(",
)
STRING_DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M:%S",
    "%m/%d/%Y",
    "%m-%d-%Y",
    "%d/%m/%Y",
    "%d-%m-%Y",
)

ENCODING_LIST_KEYS = ("columns", "rows", "measure_values", "tooltip")
ENCODING_SCALAR_KEYS = (
    "color",
    "label",
    "detail",
    "size",
    "wedge_size",
    "geographic_field",
)


class SemanticValidationError(RuntimeError):
    """Raised when the generated workbook does not match the confirmed contract."""

    def __init__(self, payload: dict[str, Any]):
        self.payload = payload
        super().__init__(payload.get("message", "Workbook semantic validation failed."))


def _now() -> datetime:
    return datetime.now()


def _now_token() -> str:
    return _now().strftime("%Y%m%d-%H%M%S-%f")


def _now_iso() -> str:
    return _now().isoformat(timespec="seconds")


def _normalize_path(path: str | Path) -> str:
    return str(Path(path).expanduser().resolve())


def _detect_datasource_type(datasource_path: str | Path) -> str:
    suffix = Path(datasource_path).suffix.lower()
    if suffix in SUPPORTED_EXCEL_SUFFIXES:
        return "excel"
    if suffix == ".hyper":
        return "hyper"
    raise ValueError(
        f"Unsupported datasource type '{suffix}'. Supported: "
        f"{', '.join(sorted(SUPPORTED_DATASOURCE_SUFFIXES))}"
    )


def _ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def _index_path(root_dir: Path) -> Path:
    return root_dir / RUN_INDEX_NAME


def _new_artifact_state() -> dict[str, Any]:
    return {
        "current": "",
        "versions": [],
        "review_current": "",
        "review_versions": [],
    }


def _default_manifest(
    output_root: Path,
    run_dir: Path,
    run_id: str,
    datasource_path: str,
    datasource_type: str,
    authoring_mode: str,
) -> dict[str, Any]:
    now = _now_iso()
    return {
        "run_id": run_id,
        "output_root": str(output_root),
        "run_dir": str(run_dir),
        "datasource_path": datasource_path,
        "datasource_type": datasource_type,
        "authoring_mode": authoring_mode,
        "status": STATUS_INITIALIZED,
        "created_at": now,
        "updated_at": now,
        "selected_primary_object": "",
        "artifacts": {key: _new_artifact_state() for key in ARTIFACT_KEYS},
        "pending_confirmation": {},
        "final_workbook": "",
        "approvals_file": APPROVALS_NAME,
        "last_error": {},
        "resolution_warnings": [],
    }


def _empty_approvals() -> dict[str, Any]:
    return {"events": []}


def _read_json(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if not isinstance(payload, dict):
        raise ValueError(f"Expected JSON object in {path}")
    return payload


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)


def _write_text(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def _authoring_mode(manifest: dict[str, Any]) -> str:
    mode = str(manifest.get("authoring_mode", AUTHORING_MODE_AGENT_FIRST)).strip()
    return mode if mode in AUTHORING_MODES else AUTHORING_MODE_AGENT_FIRST


def _allow_legacy_inference(manifest: dict[str, Any]) -> bool:
    return _authoring_mode(manifest) == AUTHORING_MODE_LEGACY


def _load_index(root_dir: Path) -> dict[str, Any]:
    index_path = _index_path(root_dir)
    if not index_path.exists():
        return {"runs": {}}
    if index_path.stat().st_size == 0:
        return {"runs": {}}
    return _read_json(index_path)


def _save_index(root_dir: Path, index_payload: dict[str, Any]) -> None:
    _write_json(_index_path(root_dir), index_payload)


def _update_index_entry(manifest: dict[str, Any]) -> None:
    entry = {
        "run_dir": manifest["run_dir"],
        "datasource_path": manifest["datasource_path"],
        "status": manifest["status"],
        "created_at": manifest["created_at"],
        "updated_at": manifest["updated_at"],
    }
    for root_dir in {Path(manifest["output_root"]), DEFAULT_AUTHORING_RUNS_DIR}:
        index_payload = _load_index(root_dir)
        runs = index_payload.setdefault("runs", {})
        runs[manifest["run_id"]] = entry
        _save_index(root_dir, index_payload)


def _artifact_entry(manifest: dict[str, Any], artifact_key: str) -> dict[str, Any]:
    entry = manifest.setdefault("artifacts", {}).setdefault(
        artifact_key,
        _new_artifact_state(),
    )
    entry.setdefault("current", "")
    entry.setdefault("versions", [])
    entry.setdefault("review_current", "")
    entry.setdefault("review_versions", [])
    return entry


def _write_versioned_artifact(
    manifest: dict[str, Any],
    artifact_key: str,
    payload: dict[str, Any],
    *,
    markdown_content: str | None = None,
) -> Path:
    run_dir = Path(manifest["run_dir"])
    token = _now_token()
    filename = f"{artifact_key}.{token}.json"
    path = run_dir / filename
    _write_json(path, payload)
    entry = _artifact_entry(manifest, artifact_key)
    entry["current"] = filename
    versions = entry.setdefault("versions", [])
    versions.append(filename)
    if markdown_content is not None:
        review_filename = f"{artifact_key}.{token}.md"
        _write_text(run_dir / review_filename, markdown_content)
        entry["review_current"] = review_filename
        review_versions = entry.setdefault("review_versions", [])
        review_versions.append(review_filename)
    return path


def _current_artifact_path(manifest: dict[str, Any], artifact_key: str) -> Path:
    current = _artifact_entry(manifest, artifact_key).get("current", "")
    if not current:
        raise ValueError(
            f"No current artifact for '{artifact_key}' in run '{manifest['run_id']}'."
        )
    path = Path(manifest["run_dir"]) / current
    if not path.exists():
        raise FileNotFoundError(f"Expected artifact missing: {path}")
    return path


def _load_current_artifact(manifest: dict[str, Any], artifact_key: str) -> dict[str, Any]:
    return _read_json(_current_artifact_path(manifest, artifact_key))


def _current_review_artifact_path(manifest: dict[str, Any], artifact_key: str) -> Path:
    current = _artifact_entry(manifest, artifact_key).get("review_current", "")
    if not current:
        raise ValueError(
            f"No current review artifact for '{artifact_key}' in run '{manifest['run_id']}'."
        )
    path = Path(manifest["run_dir"]) / current
    if not path.exists():
        raise FileNotFoundError(f"Expected review artifact missing: {path}")
    return path


def _clear_current_artifact(manifest: dict[str, Any], artifact_key: str) -> None:
    entry = _artifact_entry(manifest, artifact_key)
    entry["current"] = ""
    entry["review_current"] = ""


def _invalidate_downstream_artifacts(manifest: dict[str, Any], stage: str) -> list[str]:
    cleared: list[str] = []
    for artifact_key in DOWNSTREAM_ARTIFACTS.get(stage, ()):
        entry = _artifact_entry(manifest, artifact_key)
        if entry.get("current") or entry.get("review_current"):
            cleared.append(artifact_key)
        _clear_current_artifact(manifest, artifact_key)
    manifest["final_workbook"] = ""
    manifest["resolution_warnings"] = []
    return cleared


def _load_approvals(manifest: dict[str, Any]) -> dict[str, Any]:
    approvals_path = Path(manifest["run_dir"]) / APPROVALS_NAME
    if not approvals_path.exists():
        approvals = _empty_approvals()
        _write_json(approvals_path, approvals)
        return approvals
    return _read_json(approvals_path)


def _save_approvals(manifest: dict[str, Any], payload: dict[str, Any]) -> None:
    approvals_path = Path(manifest["run_dir"]) / APPROVALS_NAME
    _write_json(approvals_path, payload)


def _update_manifest(
    manifest: dict[str, Any],
    *,
    status: str | None = None,
    last_error: dict[str, Any] | None = None,
    pending_confirmation: dict[str, Any] | None = None,
    resolution_warnings: list[dict[str, Any]] | None = None,
) -> None:
    if status is not None:
        manifest["status"] = status
    if last_error is not None:
        manifest["last_error"] = last_error
    if pending_confirmation is not None:
        manifest["pending_confirmation"] = pending_confirmation
    if resolution_warnings is not None:
        manifest["resolution_warnings"] = resolution_warnings
    manifest["updated_at"] = _now_iso()
    manifest_path = Path(manifest["run_dir"]) / MANIFEST_NAME
    _write_json(manifest_path, manifest)
    _update_index_entry(manifest)


def _load_manifest_by_id(run_id: str) -> dict[str, Any]:
    default_path = DEFAULT_AUTHORING_RUNS_DIR / run_id / MANIFEST_NAME
    if default_path.exists():
        return _read_json(default_path)

    root_dir = DEFAULT_AUTHORING_RUNS_DIR
    index_payload = _load_index(root_dir)
    candidate = index_payload.get("runs", {}).get(run_id, {}).get("run_dir", "")
    if candidate:
        manifest_path = Path(candidate) / MANIFEST_NAME
        if manifest_path.exists():
            return _read_json(manifest_path)

    if root_dir.exists():
        for manifest_path in root_dir.rglob(MANIFEST_NAME):
            manifest = _read_json(manifest_path)
            if manifest.get("run_id") == run_id:
                return manifest

    raise FileNotFoundError(f"Authoring run '{run_id}' not found.")


def _require_status(manifest: dict[str, Any], allowed: tuple[str, ...], action: str) -> None:
    if manifest["status"] not in allowed:
        raise RuntimeError(
            f"Cannot {action} while run '{manifest['run_id']}' is in status "
            f"'{manifest['status']}'. Allowed: {', '.join(allowed)}"
        )


def _current_stage_artifact_name(manifest: dict[str, Any], stage: str) -> str:
    artifact_key = STAGE_ARTIFACT_MAP[stage]
    return _artifact_entry(manifest, artifact_key).get("current", "")


def _current_stage_review_artifact_name(manifest: dict[str, Any], stage: str) -> str:
    artifact_key = STAGE_ARTIFACT_MAP[stage]
    return _artifact_entry(manifest, artifact_key).get("review_current", "")


def request_stage_confirmation(
    run_id: str,
    stage: str,
    *,
    confirmation_mode: str,
    requested_via: str = "interactive_stage_confirmation",
) -> dict[str, Any]:
    """Persist a pending human confirmation request for one guided stage."""

    if stage not in CONFIRMABLE_STAGES:
        raise ValueError(
            f"Unsupported stage '{stage}'. Expected one of: {', '.join(sorted(CONFIRMABLE_STAGES))}"
        )

    manifest = _load_manifest_by_id(run_id)
    artifact = _current_stage_artifact_name(manifest, stage)
    if not artifact:
        raise RuntimeError(f"Cannot request confirmation for '{stage}' before its artifact exists.")

    pending = {
        "stage": stage,
        "artifact": artifact,
        "review_artifact": _current_stage_review_artifact_name(manifest, stage),
        "requested_at": _now_iso(),
        "current_status": manifest.get("status", ""),
        "confirmation_mode": confirmation_mode,
        "requested_via": requested_via,
    }
    _update_manifest(manifest, pending_confirmation=pending)
    return pending


def _require_pending_confirmation(manifest: dict[str, Any], stage: str) -> dict[str, Any]:
    pending = manifest.get("pending_confirmation", {}) or {}
    if pending.get("stage") != stage:
        raise RuntimeError(
            "confirm_authoring_stage requires a fresh interactive_stage_confirmation request "
            f"for stage '{stage}' before approval can be recorded."
        )

    expected_artifact = _current_stage_artifact_name(manifest, stage)
    if pending.get("artifact") != expected_artifact:
        raise RuntimeError(
            f"The pending confirmation for stage '{stage}' is stale. "
            "Re-run interactive_stage_confirmation after the latest stage artifact is finalized."
        )
    return pending


def _sanitize_header(value: Any, index: int, seen: dict[str, int], notes: list[str]) -> str:
    base = str(value).strip()
    if not base:
        base = f"Column_{index + 1}"
        notes.append(f"Column {index + 1} had no header; renamed to '{base}'.")
    count = seen.get(base, 0) + 1
    seen[base] = count
    if count > 1:
        renamed = f"{base}__dup{count}"
        notes.append(f"Duplicate header '{base}' detected; renamed to '{renamed}'.")
        return renamed
    return base


def _looks_like_string_date(value: Any) -> bool:
    text = str(value).strip()
    if not text or len(text) < 6:
        return False
    for fmt in STRING_DATE_FORMATS:
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue
    return False


def _is_probable_string_date_column(values: list[Any], *, threshold: float = 0.6, sample_limit: int = 24) -> bool:
    samples = [str(value).strip() for value in values if isinstance(value, str) and str(value).strip()]
    if len(samples) < 3:
        return False
    sample = samples[:sample_limit]
    matches = sum(1 for value in sample if _looks_like_string_date(value))
    return matches / len(sample) >= threshold


def _sample_rows_from_xls(path: Path) -> list[dict[str, Any]]:
    workbook = xlrd.open_workbook(str(path))
    sheets: list[dict[str, Any]] = []
    for sheet in workbook.sheets():
        rows = [
            sheet.row_values(row_index)
            for row_index in range(min(sheet.nrows, 151))
        ]
        sheets.append({"name": sheet.name, "rows": rows})
    return sheets


def _sample_rows_from_openpyxl(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "Reading .xlsx/.xlsm files requires openpyxl to be installed."
        ) from exc

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    sheets: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        rows = []
        for row in worksheet.iter_rows(min_row=1, max_row=151, values_only=True):
            rows.append(list(row))
        sheets.append({"name": worksheet.title, "rows": rows})
    workbook.close()
    return sheets


def _infer_column_type(header: str, values: list[Any]) -> str:
    lower = header.casefold()
    non_blank = [value for value in values if value not in ("", None)]
    if any(token in lower for token in ("date", "month", "year", "day", "time")):
        return "date"
    if any(token in lower for token in ("latitude", "longitude", "lat", "lon")):
        return "real"
    if not non_blank:
        return "string"
    if any(isinstance(value, (datetime, date)) for value in non_blank):
        return "date"
    if all(isinstance(value, bool) for value in non_blank):
        return "boolean"
    numeric_values: list[float] = []
    all_numeric = True
    for value in non_blank:
        if isinstance(value, bool):
            all_numeric = False
            break
        if isinstance(value, (int, float)):
            numeric_values.append(float(value))
            continue
        try:
            numeric_values.append(float(str(value).strip()))
        except Exception:
            all_numeric = False
            break
    if all_numeric and numeric_values:
        if all(float(v).is_integer() for v in numeric_values):
            return "integer"
        return "real"
    if _is_probable_string_date_column(non_blank):
        return "date"
    return "string"


def _infer_role(field_name: str, datatype: str) -> str:
    lower = field_name.casefold()
    if datatype in {"integer", "real"} and not any(
        token in lower for token in ("id", "code", "zip", "postal")
    ):
        return "measure"
    return "dimension"


def _infer_field_type(role: str, datatype: str) -> str:
    if datatype == "date":
        return "ordinal"
    if role == "measure":
        return "quantitative"
    return "nominal"


def _is_geo_field(field_name: str) -> bool:
    return bool(infer_tableau_semantic_role(field_name))


def _build_field_payloads(source_object: str, rows: list[list[Any]], notes: list[str]) -> list[dict[str, Any]]:
    if not rows:
        notes.append(f"'{source_object}' had no rows to inspect.")
        return []
    header_row = rows[0]
    seen: dict[str, int] = {}
    headers = [
        _sanitize_header(value, index, seen, notes)
        for index, value in enumerate(header_row)
    ]
    value_rows = rows[1:]
    fields: list[dict[str, Any]] = []
    for index, header in enumerate(headers):
        values = [row[index] if index < len(row) else None for row in value_rows]
        datatype = _infer_column_type(header, values)
        role = _infer_role(header, datatype)
        field_type = _infer_field_type(role, datatype)
        fields.append(
            {
                "name": header,
                "source_object": source_object,
                "inferred_type": datatype,
                "role": role,
                "field_type": field_type,
                "semantic_role": infer_tableau_semantic_role(header),
            }
        )
    return fields


def _collect_field_candidates(fields: list[dict[str, Any]]) -> dict[str, list[str]]:
    dimensions: list[str] = []
    measures: list[str] = []
    date_fields: list[str] = []
    geo_fields: list[str] = []
    for field in fields:
        name = field["name"]
        datatype = field["inferred_type"]
        role = field["role"]
        if role == "measure":
            measures.append(name)
        else:
            dimensions.append(name)
        if datatype == "date":
            date_fields.append(name)
        if field.get("semantic_role"):
            geo_fields.append(name)
    return {
        "dimensions": dimensions,
        "measures": measures,
        "date_fields": date_fields,
        "geo_fields": geo_fields,
    }


def _build_excel_schema_summary(path: Path, preferred_sheet: str = "") -> dict[str, Any]:
    notes: list[str] = []
    if path.suffix.lower() == ".xls":
        sheet_payloads = _sample_rows_from_xls(path)
    else:
        sheet_payloads = _sample_rows_from_openpyxl(path)

    if not sheet_payloads:
        raise ValueError(f"No worksheets found in Excel file: {path}")

    sheets: list[dict[str, Any]] = []
    selected_sheet = ""
    if preferred_sheet:
        if any(payload["name"] == preferred_sheet for payload in sheet_payloads):
            selected_sheet = preferred_sheet
        else:
            notes.append(
                f"Preferred sheet '{preferred_sheet}' was not found. Review available sheets before confirming schema."
            )
    if not selected_sheet and len(sheet_payloads) == 1:
        selected_sheet = sheet_payloads[0]["name"]

    selected_fields: list[dict[str, Any]] = []
    for payload in sheet_payloads:
        sheet_notes: list[str] = []
        fields = _build_field_payloads(payload["name"], payload["rows"], sheet_notes)
        sheets.append(
            {
                "name": payload["name"],
                "row_count": max(len(payload["rows"]) - 1, 0),
                "fields": fields,
                "notes": sheet_notes,
            }
        )
        notes.extend(sheet_notes)
        if payload["name"] == selected_sheet:
            selected_fields = fields

    if not selected_sheet:
        selected_fields = sheets[0]["fields"]
        if len(sheets) > 1:
            notes.append(
                "Multiple Excel sheets were discovered. Re-run intake_datasource_schema with preferred_sheet before confirming schema."
            )
        else:
            selected_sheet = sheets[0]["name"]

    field_candidates = _collect_field_candidates(selected_fields)
    profile_matches = suggest_profile_matches(
        available_fields=[field["name"] for field in selected_fields]
    )
    return {
        "datasource": {
            "path": _normalize_path(path),
            "type": "excel",
        },
        "sheets": sheets,
        "selected_primary_object": selected_sheet,
        "fields": selected_fields,
        "field_candidates": field_candidates,
        "recommended_profile_matches": profile_matches,
        "notes": notes,
    }


def _build_hyper_schema_summary(path: Path) -> dict[str, Any]:
    notes: list[str] = []
    schema = inspect_hyper_schema(str(path))
    tables = schema.get("tables", [])
    if not tables:
        raise ValueError(f"No tables found in Hyper file: {path}")

    selected_table = tables[0]["name"]
    if len(tables) > 1:
        notes.append(
            "Multiple Hyper tables were discovered. V1 treats the first table as primary and does not plan joins automatically."
        )

    table_payloads: list[dict[str, Any]] = []
    selected_fields: list[dict[str, Any]] = []
    for table in tables:
        fields: list[dict[str, Any]] = []
        for column in table.get("columns", []):
            datatype = str(column.get("type", "string")).strip().casefold()
            if "date" in datatype or "timestamp" in datatype:
                inferred = "date"
            elif any(token in datatype for token in ("double", "numeric", "real", "float", "decimal")):
                inferred = "real"
            elif any(token in datatype for token in ("int", "bigint", "smallint")):
                inferred = "integer"
            elif "bool" in datatype:
                inferred = "boolean"
            else:
                inferred = "string"
            name = str(column.get("name", "")).strip()
            role = _infer_role(name, inferred)
            fields.append(
                {
                    "name": name,
                    "source_object": table["name"],
                    "inferred_type": inferred,
                    "role": role,
                    "field_type": _infer_field_type(role, inferred),
                    "semantic_role": infer_tableau_semantic_role(name),
                }
            )
        table_payloads.append(
            {
                "schema": table.get("schema", ""),
                "name": table["name"],
                "fields": fields,
                "notes": [],
            }
        )
        if table["name"] == selected_table:
            selected_fields = fields

    field_candidates = _collect_field_candidates(selected_fields)
    profile_matches = suggest_profile_matches(
        available_fields=[field["name"] for field in selected_fields]
    )
    return {
        "datasource": {
            "path": _normalize_path(path),
            "type": "hyper",
        },
        "tables": table_payloads,
        "selected_primary_object": selected_table,
        "fields": selected_fields,
        "field_candidates": field_candidates,
        "recommended_profile_matches": profile_matches,
        "notes": notes,
    }


def _extract_audience(brief: str) -> str:
    lines = [line.strip(" -\t") for line in brief.splitlines() if line.strip()]
    for line in lines:
        lower = line.casefold()
        if lower.startswith("audience:") or lower.startswith("audience -"):
            return line.split(":", 1)[1].strip() if ":" in line else line.split("-", 1)[1].strip()
        if lower.startswith("受众：") or lower.startswith("受众:"):
            return line.split(":", 1)[1].strip() if ":" in line else line.split("：", 1)[1].strip()
    for line in lines:
        lower = line.casefold()
        if lower.startswith("for "):
            return line[4:].strip()
        if "管理层" in line or "leaders" in lower or "executive" in lower or "manager" in lower:
            return line
    return ""


def _extract_primary_question(brief: str) -> str:
    lines = [line.strip(" -\t") for line in brief.splitlines() if line.strip()]
    for line in lines:
        lower = line.casefold()
        if lower.startswith("primary question:") or lower.startswith("primary question -"):
            return line.split(":", 1)[1].strip() if ":" in line else line.split("-", 1)[1].strip()
        if lower.startswith("question:") or lower.startswith("question -"):
            return line.split(":", 1)[1].strip() if ":" in line else line.split("-", 1)[1].strip()
        if line.startswith("核心问题：") or line.startswith("核心问题:"):
            return line.split(":", 1)[1].strip() if ":" in line else line.split("：", 1)[1].strip()
    for line in lines:
        if line.endswith("?") or "？" in line:
            return line
    if lines:
        return lines[0]
    return ""


def _infer_interaction_requirement(brief: str) -> bool | None:
    lower = brief.casefold()
    if any(token in lower for token in ("no interaction", "static only", "不要交互", "无需交互")):
        return False
    if any(
        token in lower
        for token in (
            "filter",
            "drill",
            "interaction",
            "click",
            "link",
            "url",
            "go-to-sheet",
            "jump",
            "跳转",
            "联动",
            "筛选",
            "交互",
        )
    ):
        return True
    return None


def _choose_measure(fields: dict[str, list[str]], fallback: str = "Sales") -> str:
    return fields["measures"][0] if fields["measures"] else fallback


def _choose_dimension(fields: dict[str, list[str]], fallback: str = "Category") -> str:
    return fields["dimensions"][0] if fields["dimensions"] else fallback


def _choose_geo(fields: dict[str, list[str]], fallback: str = "State/Province") -> str:
    geo_fields = fields.get("geo_fields", [])
    if not geo_fields:
        return fallback

    preferred_order = [
        "region",
        "state/province",
        "state",
        "province",
        "country/region",
        "country",
        "city",
        "postal code",
        "zip code",
        "zipcode",
        "latitude",
        "longitude",
        "lat",
        "lon",
    ]
    ranked = {
        name: index
        for index, name in enumerate(preferred_order)
    }

    def _rank(field_name: str) -> tuple[int, str]:
        normalized = " ".join(
            field_name.casefold().replace("_", " ").replace("-", " ").split()
        )
        return (ranked.get(normalized, len(preferred_order)), field_name)

    return sorted(geo_fields, key=_rank)[0]


def _choose_date(fields: dict[str, list[str]], fallback: str = "Order Date") -> str:
    return fields["date_fields"][0] if fields["date_fields"] else fallback


def _resolve_mark_type(question: str, priority: str, field_candidates: dict[str, list[str]]) -> str:
    lower = " ".join((question or "", priority or "")).casefold()
    if "map" in lower or "region" in lower or "state" in lower or "geograph" in lower:
        if field_candidates["geo_fields"]:
            return "Map"
    if any(token in lower for token in ("trend", "month", "time", "over time", "timeline")):
        if field_candidates["date_fields"]:
            return "Line"
    if priority == "summary":
        return "Text"
    return "Bar"


def _default_worksheet_specs(field_candidates: dict[str, list[str]]) -> list[dict[str, Any]]:
    primary_measure = _choose_measure(field_candidates)
    primary_dimension = _choose_dimension(field_candidates)
    geo_dimension = _choose_geo(field_candidates)
    date_field = _choose_date(field_candidates)
    return [
        {
            "name": "Summary View",
            "question": f"What high-level {primary_measure} summary should this dashboard show first?",
            "mark_type": "Text",
            "priority": "summary",
        },
        {
            "name": "Primary View",
            "question": f"Which {geo_dimension or primary_dimension} is driving {primary_measure}?",
            "mark_type": "Map" if field_candidates["geo_fields"] else "Bar",
            "priority": "primary",
        },
        {
            "name": "Detail View",
            "question": f"How is {primary_measure} changing over {date_field}?",
            "mark_type": "Line" if field_candidates["date_fields"] else "Bar",
            "priority": "detail",
        },
    ]


def _deep_merge(base: dict[str, Any], updates: dict[str, Any]) -> dict[str, Any]:
    merged = deepcopy(base)
    for key, value in updates.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = _deep_merge(merged[key], value)
        else:
            merged[key] = value
    return merged


def _unique_strings(values: list[str]) -> list[str]:
    seen: set[str] = set()
    result: list[str] = []
    for value in values:
        if value in ("", None):
            continue
        cleaned = str(value).strip()
        if cleaned and cleaned not in seen:
            seen.add(cleaned)
            result.append(cleaned)
    return result


def _json_code_fence(payload: dict[str, Any]) -> str:
    return "```yaml\n" + json.dumps(payload, ensure_ascii=False, indent=2) + "\n```"


def _format_bullets(items: list[str]) -> str:
    if not items:
        return "- (none)"
    return "\n".join(f"- {item}" for item in items)


def _extract_override_block(markdown_text: str) -> str:
    in_block = False
    block_lines: list[str] = []
    for line in markdown_text.splitlines():
        stripped = line.strip()
        if not in_block and stripped.startswith("```"):
            label = stripped[3:].strip().casefold()
            if label in {"yaml", "yml"}:
                in_block = True
            continue
        if in_block and stripped.startswith("```"):
            break
        if in_block:
            block_lines.append(line)
    block = "\n".join(block_lines).strip()
    if not block:
        raise ValueError("No fenced YAML override block was found in the Markdown artifact.")
    return block


def _parse_override_block(block: str) -> dict[str, Any]:
    try:
        if yaml is not None:
            parsed = yaml.safe_load(block)
        else:
            parsed = json.loads(block)
    except Exception as exc:  # pragma: no cover - exact parser varies by env
        raise ValueError(
            "Failed to parse the fenced YAML override block. "
            "Use JSON-compatible YAML in the code fence for V1.1."
        ) from exc
    if parsed is None:
        return {}
    if not isinstance(parsed, dict):
        raise ValueError("The fenced YAML override block must parse to an object.")
    return parsed


def _load_markdown_overrides(markdown_path: str | Path) -> dict[str, Any]:
    normalized_path = Path(markdown_path).expanduser().resolve()
    if not normalized_path.exists():
        raise FileNotFoundError(f"Markdown review artifact not found: {normalized_path}")
    markdown_text = normalized_path.read_text(encoding="utf-8")
    return _parse_override_block(_extract_override_block(markdown_text))


def _combined_overrides(
    *,
    markdown_path: str = "",
    user_answers_json: str = "",
) -> dict[str, Any]:
    overrides: dict[str, Any] = {}
    if markdown_path.strip():
        overrides = _load_markdown_overrides(markdown_path)
    if user_answers_json.strip():
        parsed = json.loads(user_answers_json)
        if not isinstance(parsed, dict):
            raise ValueError("user_answers_json must be a JSON object.")
        overrides = _deep_merge(overrides, parsed)
    return overrides


def _choose_named_field(fields: list[str], preferred_names: list[str]) -> str:
    normalized_map = {
        " ".join(field.casefold().replace("_", " ").replace("-", " ").split()): field
        for field in fields
    }
    for preferred in preferred_names:
        normalized = " ".join(preferred.casefold().replace("_", " ").replace("-", " ").split())
        if normalized in normalized_map:
            return normalized_map[normalized]
    return ""


def _normalize_field_key(value: str) -> str:
    return " ".join(str(value).casefold().replace("_", " ").replace("-", " ").split())


def _known_calculated_field_name(field_name: str) -> str:
    normalized = _normalize_field_key(field_name)
    return KNOWN_CALCULATED_FORMULA_LOOKUP.get(normalized, "")


def _available_field_summary(
    available_fields: list[str],
    calculated_field_lookup: dict[str, dict[str, Any]] | None = None,
    *,
    limit: int = 12,
) -> str:
    summary = list(available_fields)
    if calculated_field_lookup:
        summary.extend(
            spec["name"]
            for spec in calculated_field_lookup.values()
            if str(spec.get("name", "")).strip()
        )
    unique = _dedupe_strings(summary)
    if not unique:
        return "(none)"
    if len(unique) <= limit:
        return ", ".join(unique)
    return ", ".join(unique[:limit]) + f", ... (+{len(unique) - limit} more)"


def _dedupe_resolution_warnings(warnings: list[dict[str, Any]]) -> list[dict[str, Any]]:
    deduped: list[dict[str, Any]] = []
    seen: set[str] = set()
    for warning in warnings:
        if not isinstance(warning, dict):
            continue
        token = json.dumps(warning, ensure_ascii=False, sort_keys=True)
        if token in seen:
            continue
        seen.add(token)
        deduped.append(warning)
    return deduped


def _record_resolution_warning(
    warnings: list[dict[str, Any]] | None,
    *,
    field_name: str,
    context: str,
    datasource_path: str,
    available_fields: list[str],
    calculated_field_lookup: dict[str, dict[str, Any]] | None = None,
) -> None:
    if warnings is None:
        return
    warnings.append(
        {
            "type": "field_resolution_fallback",
            "field_name": field_name,
            "context": context,
            "datasource_path": datasource_path,
            "fallback": "omitted",
            "available_fields_excerpt": _available_field_summary(
                available_fields,
                calculated_field_lookup,
            ),
        }
    )


def _is_expression(value: str) -> bool:
    text = str(value).strip()
    if not text:
        return False
    return text.startswith("[") or text.casefold().startswith(EXPRESSION_FUNCTION_PREFIXES)


def _available_field_lookup(available_fields: list[str]) -> dict[str, str]:
    return {_normalize_field_key(field): field for field in available_fields if str(field).strip()}


def _resolve_field_name(
    field_name: str,
    available_fields: list[str],
    *,
    strict: bool = False,
    resolution_warnings: list[dict[str, Any]] | None = None,
    context: str = "",
    datasource_path: str = "",
    calculated_field_lookup: dict[str, dict[str, Any]] | None = None,
) -> str:
    text = str(field_name).strip()
    if not text:
        return ""
    known_calculated_name = _known_calculated_field_name(text)
    if _is_expression(text):
        return text
    if known_calculated_name:
        return known_calculated_name

    available_lookup = _available_field_lookup(available_fields)
    normalized = _normalize_field_key(text)
    if normalized in available_lookup:
        return available_lookup[normalized]

    calculated_field_lookup = calculated_field_lookup or {}
    calculated_field = calculated_field_lookup.get(normalized)
    if calculated_field:
        return str(calculated_field.get("name", "")).strip()

    if strict:
        raise RuntimeError(
            f"Could not resolve field '{text}'"
            + (f" for {context}" if context else "")
            + (
                f" against datasource '{datasource_path}'. "
                if datasource_path
                else ". "
            )
            + "Available fields: "
            + _available_field_summary(available_fields, calculated_field_lookup)
        )

    _record_resolution_warning(
        resolution_warnings,
        field_name=text,
        context=context,
        datasource_path=datasource_path,
        available_fields=available_fields,
        calculated_field_lookup=calculated_field_lookup,
    )
    return ""


def _default_measure_expression(
    field_name: str,
    *,
    calculated_field_names: set[str] | None = None,
) -> str:
    text = str(field_name).strip()
    if not text:
        return ""
    if _is_expression(text):
        return text
    if _known_calculated_field_name(text):
        return _known_calculated_field_name(text)
    calculated_field_names = calculated_field_names or set()
    if text in calculated_field_names:
        return text
    if _normalize_field_key(text) == "discount":
        return f"AVG({text})"
    return f"SUM({text})"


def _default_date_expression(field_name: str) -> str:
    text = str(field_name).strip()
    if not text:
        return ""
    if _is_expression(text):
        return text
    return f"MONTH({text})"


def _worksheet_text(worksheet: dict[str, Any]) -> str:
    return " ".join(
        [
            str(worksheet.get("name", "")),
            str(worksheet.get("question", "")),
            str(worksheet.get("priority", "")),
            str(worksheet.get("mark_type", "")),
        ]
    ).casefold()


def _dedupe_strings(values: list[str]) -> list[str]:
    return _unique_strings([str(value).strip() for value in values if str(value).strip()])


def _normalize_string_list(value: Any) -> list[str]:
    if not isinstance(value, list):
        return []
    return _dedupe_strings([str(item).strip() for item in value if str(item).strip()])


def _normalize_encoding_spec(value: Any) -> dict[str, Any]:
    normalized = {key: [] for key in ENCODING_LIST_KEYS}
    normalized.update({key: "" for key in ENCODING_SCALAR_KEYS})
    if not isinstance(value, dict):
        return normalized
    for key in ENCODING_LIST_KEYS:
        normalized[key] = _dedupe_strings(value.get(key, [])) if isinstance(value.get(key, []), list) else []
    for key in ENCODING_SCALAR_KEYS:
        normalized[key] = str(value.get(key, "")).strip()
    return normalized


def _worksheet_priority_name(contract: dict[str, Any], priority: str) -> str:
    worksheet = _worksheet_by_priority(contract, priority)
    return str(worksheet.get("name", "")).strip() if worksheet else ""


def _choose_requested_dimension(text: str, field_candidates: dict[str, list[str]]) -> str:
    dimensions = field_candidates.get("dimensions", [])
    geo_fields = field_candidates.get("geo_fields", [])
    date_fields = field_candidates.get("date_fields", [])

    ordered_matches = [
        (("sub-category", "subcategory", "sub category"), dimensions, ["Sub-Category"]),
        (("category",), dimensions, ["Category"]),
        (("segment",), dimensions, ["Segment"]),
        (("ship mode",), dimensions, ["Ship Mode"]),
        (("customer",), dimensions, ["Customer Name"]),
        (("product",), dimensions, ["Product Name"]),
        (("region",), geo_fields + dimensions, ["Region"]),
        (("state/province", "state", "province"), geo_fields + dimensions, ["State/Province", "State", "Province"]),
        (("country/region", "country"), geo_fields + dimensions, ["Country/Region", "Country"]),
        (("city",), geo_fields + dimensions, ["City"]),
        (("postal", "zip"), geo_fields + dimensions, ["Postal Code"]),
        (("ship date",), date_fields, ["Ship Date"]),
        (("order date", "time", "trend", "timeline", "date"), date_fields, ["Order Date"]),
    ]

    for tokens, pool, preferred in ordered_matches:
        if any(token in text for token in tokens):
            chosen = _choose_named_field(pool, preferred)
            if chosen:
                return chosen

    return _choose_dimension(field_candidates, "")


def _choose_requested_geo(text: str, field_candidates: dict[str, list[str]]) -> str:
    geo_fields = field_candidates.get("geo_fields", [])
    for tokens, preferred in (
        (("state/province", "state", "province"), ["State/Province", "State", "Province"]),
        (("region",), ["Region"]),
        (("country/region", "country"), ["Country/Region", "Country"]),
        (("city",), ["City"]),
    ):
        if any(token in text for token in tokens):
            chosen = _choose_named_field(geo_fields, preferred)
            if chosen:
                return chosen
    return _choose_geo(field_candidates, "")


def _choose_requested_date(text: str, field_candidates: dict[str, list[str]]) -> str:
    date_fields = field_candidates.get("date_fields", [])
    for tokens, preferred in (
        (("ship date",), ["Ship Date"]),
        (("order date", "time", "trend", "timeline", "month", "date"), ["Order Date"]),
    ):
        if any(token in text for token in tokens):
            chosen = _choose_named_field(date_fields, preferred)
            if chosen:
                return chosen
    return _choose_date(field_candidates, "")


def _choose_requested_measure(text: str, field_candidates: dict[str, list[str]]) -> str:
    measures = field_candidates.get("measures", [])
    if "profit ratio" in text:
        return "Profit Ratio"
    for tokens, preferred in (
        (("sales", "revenue"), ["Sales"]),
        (("profit", "margin"), ["Profit"]),
        (("quantity", "volume", "units"), ["Quantity"]),
        (("discount",), ["Discount"]),
    ):
        if any(token in text for token in tokens):
            chosen = _choose_named_field(measures, preferred)
            if chosen:
                return chosen
    return _choose_measure(field_candidates, "")


def _normalized_worksheet_filters(
    worksheet: dict[str, Any],
    contract: dict[str, Any],
) -> list[str]:
    worksheet_filters = worksheet.get("filters", [])
    if isinstance(worksheet_filters, list) and worksheet_filters:
        return _dedupe_strings(worksheet_filters)
    return _dedupe_strings(contract.get("constraints", {}).get("filters", []))


def _default_contract_actions(
    contract: dict[str, Any],
    field_candidates: dict[str, list[str]],
) -> list[dict[str, Any]]:
    worksheet_names = _contract_worksheet_names(contract)
    if not contract.get("require_interaction") or len(worksheet_names) < 2:
        return []

    source_sheet = _worksheet_priority_name(contract, "primary") or worksheet_names[0]
    targets = [name for name in worksheet_names if name != source_sheet]
    action_field = (
        _choose_geo(field_candidates, "")
        if field_candidates.get("geo_fields")
        else _choose_dimension(field_candidates, "")
    )
    if not source_sheet or not targets:
        return []
    return [
        {
            "type": "filter",
            "source": source_sheet,
            "targets": targets,
            "fields": [action_field] if action_field else [],
            "caption": f"Filter related views from {source_sheet}",
            "url": "",
        }
    ]


def _normalized_contract_actions(
    contract: dict[str, Any],
    field_candidates: dict[str, list[str]],
    *,
    allow_defaults: bool,
) -> list[dict[str, Any]]:
    raw_actions = contract.get("actions") or (
        _default_contract_actions(contract, field_candidates) if allow_defaults else []
    )
    normalized: list[dict[str, Any]] = []
    worksheet_names = [
        str(worksheet.get("name", "")).strip()
        for worksheet in contract.get("worksheets", [])
        if isinstance(worksheet, dict) and str(worksheet.get("name", "")).strip()
    ]
    legacy_sheet_aliases = {
        "summary view": _worksheet_priority_name(contract, "summary"),
        "primary view": _worksheet_priority_name(contract, "primary"),
        "detail view": _worksheet_priority_name(contract, "detail"),
    }

    def _resolve_sheet_alias(name: Any, *, fallback: str = "") -> str:
        text = str(name).strip()
        if not text:
            return fallback
        mapped = legacy_sheet_aliases.get(_normalize_field_key(text), "")
        if mapped:
            return mapped
        if text in worksheet_names:
            return text
        return text

    def _action_fields(action: dict[str, Any]) -> list[str]:
        raw_fields = action.get("fields", [])
        if isinstance(raw_fields, list):
            return _dedupe_strings(raw_fields)
        if str(raw_fields).strip():
            return [str(raw_fields).strip()]
        return []

    for action in raw_actions:
        if not isinstance(action, dict):
            continue
        source = _resolve_sheet_alias(
            action.get("source", ""),
            fallback=_worksheet_priority_name(contract, "primary"),
        )
        targets = [_resolve_sheet_alias(target) for target in _dedupe_strings(action.get("targets", []))]
        target = _resolve_sheet_alias(action.get("target", ""))
        if not targets and target:
            targets = [target]
        targets = [target_name for target_name in targets if target_name and target_name != source]
        normalized.append(
            {
                "type": str(action.get("type", "filter")).strip() or "filter",
                "source": source,
                "target": targets[0] if targets else target,
                "targets": targets,
                "fields": _action_fields(action),
                "url": str(action.get("url", "")).strip(),
                "caption": str(action.get("caption", "")).strip(),
            }
        )
    return normalized


def _normalized_contract_calculated_fields(contract: dict[str, Any]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for item in contract.get("calculated_fields", []):
        if not isinstance(item, dict):
            continue
        name = str(item.get("name", "")).strip()
        if not name:
            continue
        key = _normalize_field_key(name)
        if key in seen:
            continue
        seen.add(key)
        normalized.append(
            {
                "name": name,
                "formula": str(item.get("formula", "")).strip(),
                "datatype": str(item.get("datatype", "real")).strip() or "real",
            }
        )
    return normalized


def _contract_calculated_field_lookup(contract: dict[str, Any]) -> dict[str, dict[str, Any]]:
    return {
        _normalize_field_key(spec["name"]): spec
        for spec in _normalized_contract_calculated_fields(contract)
        if str(spec.get("name", "")).strip()
    }


def _populate_worksheet_execution_spec(
    worksheet: dict[str, Any],
    contract: dict[str, Any],
    schema_summary: dict[str, Any],
    *,
    fail_on_unresolved: bool,
    allow_inference: bool,
    resolution_warnings: list[dict[str, Any]] | None = None,
    calculated_field_lookup: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any]:
    field_candidates = schema_summary.get("field_candidates", {})
    available_fields = [field["name"] for field in schema_summary.get("fields", [])]
    calculated_field_lookup = calculated_field_lookup or {}
    calculated_field_names = {
        str(spec.get("name", "")).strip()
        for spec in calculated_field_lookup.values()
        if str(spec.get("name", "")).strip()
    }
    date_fields = set(field_candidates.get("date_fields", []))
    geo_fields = set(field_candidates.get("geo_fields", []))
    text = _worksheet_text(worksheet)
    worksheet_name = str(worksheet.get("name", "")).strip() or "Worksheet"
    datasource_path = str(schema_summary.get("datasource", {}).get("path", "")).strip()
    mark_type = str(worksheet.get("mark_type", "")).strip()
    if not mark_type and allow_inference:
        mark_type = _resolve_mark_type(
            worksheet.get("question", ""),
            worksheet.get("priority", ""),
            field_candidates,
        )

    def resolve_field(value: Any, context: str) -> str:
        return _resolve_field_name(
            str(value).strip(),
            available_fields,
            strict=fail_on_unresolved,
            resolution_warnings=resolution_warnings,
            context=f"worksheet '{worksheet_name}' {context}",
            datasource_path=datasource_path,
            calculated_field_lookup=calculated_field_lookup,
        )

    def resolve_many(values: list[Any], context: str) -> list[str]:
        return _dedupe_strings([resolve_field(value, context) for value in values])

    dimensions = resolve_many(worksheet.get("dimensions", []), "dimensions")
    measures = resolve_many(worksheet.get("measures", []), "measures")
    kpi_fields = resolve_many(worksheet.get("kpi_fields", []), "kpi_fields")
    encodings = _normalize_encoding_spec(worksheet.get("encodings", {}))
    for key in ENCODING_LIST_KEYS:
        encodings[key] = resolve_many(encodings.get(key, []), f"encodings.{key}")
    for key in ENCODING_SCALAR_KEYS:
        encodings[key] = resolve_field(encodings.get(key, ""), f"encodings.{key}")
    filters = resolve_many(_normalize_string_list(worksheet.get("filters")), "filters")
    if not filters:
        filters = resolve_many(
            _normalize_string_list(contract.get("constraints", {}).get("filters", [])),
            "filters",
        )
    sort_descending = str(worksheet.get("sort_descending", "")).strip()

    def _explicit_geo_dimension() -> str:
        for field in dimensions:
            if field in geo_fields or _is_geo_field(field):
                return field
        return ""

    def _explicit_date_dimension() -> str:
        for field in dimensions:
            if field in date_fields:
                return field
        return ""

    if mark_type == "Text":
        if not kpi_fields:
            explicit_kpis = resolve_many(
                _normalize_string_list(contract.get("constraints", {}).get("kpis", [])),
                "constraints.kpis",
            )
            if explicit_kpis:
                kpi_fields = explicit_kpis
            elif allow_inference:
                kpi_fields = _dedupe_strings(_suggest_kpis(field_candidates))
        if not encodings["measure_values"]:
            encodings["measure_values"] = [
                _default_measure_expression(
                    resolve_field(field, "kpi_fields"),
                    calculated_field_names=calculated_field_names,
                )
                for field in kpi_fields
                if resolve_field(field, "kpi_fields")
            ]
    elif mark_type == "Pie":
        dimension = encodings["color"] or (dimensions[0] if dimensions else "")
        if not dimension and allow_inference:
            dimension = _choose_requested_dimension(text, field_candidates)
        dimension = resolve_field(dimension, "encodings.color")
        measure = encodings["wedge_size"] or (measures[0] if measures else "")
        if not measure and allow_inference:
            measure = _choose_requested_measure(text, field_candidates)
        measure = resolve_field(measure, "encodings.wedge_size")
        if dimension:
            dimensions = _dedupe_strings([dimension] + dimensions)
            encodings["color"] = encodings["color"] or dimension
            encodings["label"] = encodings["label"] or dimension
        if measure:
            if (
                not _known_calculated_field_name(measure)
                and measure not in calculated_field_names
                and not _is_expression(measure)
            ):
                measures = _dedupe_strings([measure] + measures)
            encodings["wedge_size"] = encodings["wedge_size"] or _default_measure_expression(
                measure,
                calculated_field_names=calculated_field_names,
            )
        if not encodings["tooltip"] and dimension and encodings["wedge_size"]:
            encodings["tooltip"] = [dimension, encodings["wedge_size"]]
    elif mark_type == "Map":
        geo_field = encodings["geographic_field"] or _explicit_geo_dimension()
        if not geo_field and allow_inference:
            geo_field = _choose_requested_geo(text, field_candidates)
        geo_field = resolve_field(geo_field, "encodings.geographic_field")
        measure = encodings["color"] or (measures[0] if measures else "")
        if not measure and allow_inference:
            measure = _choose_requested_measure(text, field_candidates)
        measure = resolve_field(measure, "encodings.color")
        if geo_field:
            dimensions = _dedupe_strings([geo_field] + dimensions)
            encodings["geographic_field"] = geo_field
        if measure:
            if (
                not _known_calculated_field_name(measure)
                and measure not in calculated_field_names
                and not _is_expression(measure)
            ):
                measures = _dedupe_strings([measure] + measures)
            encodings["color"] = encodings["color"] or _default_measure_expression(
                measure,
                calculated_field_names=calculated_field_names,
            )
        if not encodings["tooltip"] and geo_field:
            tooltip_values = [geo_field]
            if encodings["color"]:
                tooltip_values.append(encodings["color"])
            encodings["tooltip"] = tooltip_values
    elif mark_type == "Line":
        date_field = _explicit_date_dimension()
        if not date_field and allow_inference:
            date_field = _choose_requested_date(text, field_candidates)
        date_field = resolve_field(date_field, "encodings.columns")
        measure = measures[0] if measures else ""
        if not measure and allow_inference:
            measure = _choose_requested_measure(text, field_candidates)
        measure = resolve_field(measure, "encodings.rows")
        if date_field:
            dimensions = _dedupe_strings([date_field] + dimensions)
            encodings["columns"] = encodings["columns"] or [_default_date_expression(date_field)]
        if measure:
            if (
                not _known_calculated_field_name(measure)
                and measure not in calculated_field_names
                and not _is_expression(measure)
            ):
                measures = _dedupe_strings([measure] + measures)
            encodings["rows"] = encodings["rows"] or [
                _default_measure_expression(
                    measure,
                    calculated_field_names=calculated_field_names,
                )
            ]
        if not encodings["tooltip"] and date_field:
            tooltip_values = [date_field]
            if encodings["rows"]:
                tooltip_values.extend(encodings["rows"])
            encodings["tooltip"] = tooltip_values
    else:
        dimension = encodings["rows"][0] if encodings["rows"] else (dimensions[0] if dimensions else "")
        if not dimension and allow_inference:
            dimension = _choose_requested_dimension(text, field_candidates)
        dimension = resolve_field(dimension, "encodings.rows")
        measure = encodings["columns"][0] if encodings["columns"] else (measures[0] if measures else "")
        if not measure and allow_inference:
            measure = _choose_requested_measure(text, field_candidates)
        measure = resolve_field(measure, "encodings.columns")
        if dimension:
            dimensions = _dedupe_strings([dimension] + dimensions)
            encodings["rows"] = encodings["rows"] or [dimension]
            encodings["label"] = encodings["label"] or dimension
        if measure:
            if (
                not _known_calculated_field_name(measure)
                and measure not in calculated_field_names
                and not _is_expression(measure)
            ):
                measures = _dedupe_strings([measure] + measures)
            encodings["columns"] = encodings["columns"] or [
                _default_measure_expression(
                    measure,
                    calculated_field_names=calculated_field_names,
                )
            ]
            sort_descending = sort_descending or encodings["columns"][0]
        if not encodings["tooltip"] and dimension:
            tooltip_values = [dimension]
            if encodings["columns"]:
                tooltip_values.extend(encodings["columns"])
            encodings["tooltip"] = tooltip_values

    worksheet["mark_type"] = mark_type
    worksheet["dimensions"] = dimensions
    worksheet["measures"] = measures
    worksheet["kpi_fields"] = kpi_fields
    worksheet["encodings"] = encodings
    worksheet["sort_descending"] = sort_descending
    worksheet["filters"] = filters

    errors: list[str] = []
    if not mark_type:
        errors.append("Worksheet mark_type must be explicit in agent-first mode.")
    if mark_type == "Text" and not encodings["measure_values"]:
        errors.append("Text worksheets require kpi_fields or encodings.measure_values.")
    if mark_type == "Pie":
        if not encodings["color"]:
            errors.append("Pie worksheets require a categorical encodings.color field.")
        if not encodings["wedge_size"]:
            errors.append("Pie worksheets require encodings.wedge_size.")
    if mark_type == "Map" and not encodings["geographic_field"]:
        errors.append("Map worksheets require encodings.geographic_field.")
    if mark_type == "Line":
        if not encodings["columns"]:
            errors.append("Line worksheets require encodings.columns.")
        if not encodings["rows"]:
            errors.append("Line worksheets require encodings.rows.")
    if mark_type not in {"Text", "Pie", "Map", "Line"}:
        if not encodings["rows"] or not encodings["columns"]:
            errors.append(f"{mark_type or 'Worksheet'} requires both row and column encodings.")

    if fail_on_unresolved and errors:
        raise RuntimeError(
            f"Worksheet '{worksheet.get('name', 'Worksheet')}' is not executable: "
            + " ".join(errors)
        )

    return worksheet


def _ensure_contract_execution_spec(
    contract: dict[str, Any],
    schema_summary: dict[str, Any],
    *,
    fail_on_unresolved: bool,
    allow_inference: bool,
) -> dict[str, Any]:
    normalized = deepcopy(contract)
    normalized["calculated_fields"] = _normalized_contract_calculated_fields(normalized)
    calculated_field_lookup = _contract_calculated_field_lookup(normalized)
    available_fields = [field["name"] for field in schema_summary.get("fields", [])]
    datasource_path = str(schema_summary.get("datasource", {}).get("path", "")).strip()
    resolution_warnings: list[dict[str, Any]] = []

    def resolve_contract_field(value: Any, context: str) -> str:
        return _resolve_field_name(
            str(value).strip(),
            available_fields,
            strict=fail_on_unresolved,
            resolution_warnings=resolution_warnings,
            context=context,
            datasource_path=datasource_path,
            calculated_field_lookup=calculated_field_lookup,
        )

    constraints = normalized.setdefault("constraints", {})
    if isinstance(constraints, dict):
        constraints["filters"] = _dedupe_strings(
            [
                resolve_contract_field(field, "constraints.filters")
                for field in _normalize_string_list(constraints.get("filters", []))
            ]
        )
        constraints["kpis"] = _dedupe_strings(
            [
                resolve_contract_field(field, "constraints.kpis")
                for field in _normalize_string_list(constraints.get("kpis", []))
            ]
        )
    normalized["actions"] = _normalized_contract_actions(
        normalized,
        schema_summary.get("field_candidates", {}),
        allow_defaults=allow_inference,
    )
    for action in normalized["actions"]:
        if not isinstance(action, dict):
            continue
        action["fields"] = _dedupe_strings(
            [
                resolve_contract_field(
                    field,
                    f"action '{action.get('source', '') or '(unspecified)'}' fields",
                )
                for field in _normalize_string_list(action.get("fields", []))
            ]
        )
    normalized["worksheets"] = [
        _populate_worksheet_execution_spec(
            deepcopy(worksheet),
            normalized,
            schema_summary,
            fail_on_unresolved=fail_on_unresolved,
            allow_inference=allow_inference,
            resolution_warnings=resolution_warnings,
            calculated_field_lookup=calculated_field_lookup,
        )
        for worksheet in normalized.get("worksheets", [])
        if isinstance(worksheet, dict)
    ]
    normalized["resolution_warnings"] = _dedupe_resolution_warnings(resolution_warnings)
    return normalized


def _suggest_kpis(field_candidates: dict[str, list[str]]) -> list[str]:
    measures = field_candidates.get("measures", [])
    kpis: list[str] = []
    for preferred in ("Sales", "Profit", "Quantity", "Discount"):
        chosen = _choose_named_field(measures, [preferred])
        if chosen:
            kpis.append(chosen)
    if _choose_named_field(measures, ["Sales"]) and _choose_named_field(measures, ["Profit"]):
        kpis.append("Profit Ratio")
    if not kpis and measures:
        kpis.extend(measures[:3])
    return _unique_strings(kpis)


def _suggest_filters(field_candidates: dict[str, list[str]]) -> list[str]:
    dimensions = field_candidates.get("dimensions", [])
    filters: list[str] = []
    date_field = _choose_date(field_candidates, "")
    geo_field = _choose_geo(field_candidates, "")
    if date_field:
        filters.append(date_field)
    if geo_field:
        filters.append(geo_field)
    for preferred in ("Category", "Sub-Category", "Segment"):
        chosen = _choose_named_field(dimensions, [preferred])
        if chosen:
            filters.append(chosen)
    if not filters:
        filters.extend(dimensions[:3])
    return _unique_strings(filters)


def _worksheet_plan(
    *,
    summary_question: str,
    primary_name: str,
    primary_question: str,
    primary_mark: str,
    detail_question: str,
    detail_mark: str,
) -> list[dict[str, Any]]:
    return [
        {
            "name": "Summary View",
            "question": summary_question,
            "mark_type": "Text",
            "priority": "summary",
        },
        {
            "name": primary_name,
            "question": primary_question,
            "mark_type": primary_mark,
            "priority": "primary",
        },
        {
            "name": "Detail View",
            "question": detail_question,
            "mark_type": detail_mark,
            "priority": "detail",
        },
    ]


def _analysis_direction(
    *,
    direction_id: str,
    title: str,
    business_question: str,
    why_it_matters: str,
    recommended_kpis: list[str],
    primary_view: dict[str, Any],
    detail_view: dict[str, Any],
    recommended_filters: list[str],
    interaction_pattern: str,
    caveats: list[str],
    layout_pattern: str = "executive overview",
) -> dict[str, Any]:
    worksheet_plan = _worksheet_plan(
        summary_question=f"What top-level summary should {title} show first?",
        primary_name=primary_view["name"],
        primary_question=primary_view["question"],
        primary_mark=primary_view["mark_type"],
        detail_question=detail_view["question"],
        detail_mark=detail_view["mark_type"],
    )
    return {
        "id": direction_id,
        "title": title,
        "business_question": business_question,
        "why_it_matters": why_it_matters,
        "recommended_kpis": recommended_kpis,
        "primary_view": primary_view,
        "detail_view": detail_view,
        "recommended_filters": recommended_filters,
        "interaction_pattern": interaction_pattern,
        "caveats": caveats,
        "contract_seed": {
            "dashboard_name": title,
            "layout_pattern": layout_pattern,
            "kpis": recommended_kpis,
            "filters": recommended_filters,
            "interaction_pattern": interaction_pattern,
            "worksheets": worksheet_plan,
        },
    }


def _build_analysis_directions(schema_summary: dict[str, Any]) -> list[dict[str, Any]]:
    field_candidates = schema_summary.get("field_candidates", {})
    measures = field_candidates.get("measures", [])
    dimensions = field_candidates.get("dimensions", [])
    primary_measure = _choose_named_field(measures, ["Sales", "Profit"]) or _choose_measure(field_candidates)
    profit_measure = _choose_named_field(measures, ["Profit"]) or primary_measure
    quantity_measure = _choose_named_field(measures, ["Quantity"])
    geo_field = _choose_geo(field_candidates, "")
    date_field = _choose_date(field_candidates, "")
    category_field = _choose_named_field(dimensions, ["Category"]) or _choose_dimension(field_candidates)
    subcategory_field = _choose_named_field(dimensions, ["Sub-Category"])
    segment_field = _choose_named_field(dimensions, ["Segment"])

    common_kpis = _suggest_kpis(field_candidates)
    common_filters = _suggest_filters(field_candidates)
    directions: list[dict[str, Any]] = []

    primary_axis = geo_field or category_field or _choose_dimension(field_candidates)
    directions.append(
        _analysis_direction(
            direction_id="executive_overview",
            title="Executive Overview",
            business_question=(
                f"Which {primary_axis} and categories are driving {primary_measure}"
                + (f" and {profit_measure}" if profit_measure != primary_measure else "")
                + "?"
            ),
            why_it_matters="Gives leaders a fast cross-cut of performance, mix, and drill paths.",
            recommended_kpis=common_kpis,
            primary_view={
                "name": "Primary View",
                "question": f"Which {primary_axis} is driving {primary_measure}?",
                "mark_type": "Map" if geo_field else "Bar",
            },
            detail_view={
                "name": "Detail View",
                "question": f"How is {primary_measure} changing over {date_field or 'time'}?",
                "mark_type": "Line" if date_field else "Bar",
            },
            recommended_filters=common_filters,
            interaction_pattern="Click Primary View to filter Detail View.",
            caveats=[
                "Keep the dashboard to one executive page.",
                "URL-style interactions should be confirmed before plan generation.",
            ],
        )
    )

    if geo_field:
        directions.append(
            _analysis_direction(
                direction_id="geographic_performance",
                title="Regional Performance",
                business_question=f"Which {geo_field} are outperforming or underperforming on {primary_measure} and {profit_measure}?",
                why_it_matters="Helps regional leaders spot where to drill deeper and compare markets quickly.",
                recommended_kpis=_unique_strings([primary_measure, profit_measure, quantity_measure, "Profit Ratio"]),
                primary_view={
                    "name": "Primary View",
                    "question": f"Which {geo_field} is driving the biggest performance gap?",
                    "mark_type": "Map",
                },
                detail_view={
                    "name": "Detail View",
                    "question": f"How does each {geo_field} trend over {date_field or 'time'}?",
                    "mark_type": "Line" if date_field else "Bar",
                },
                recommended_filters=_unique_strings([date_field, geo_field, category_field]),
                interaction_pattern="Click a region to filter the detail trend.",
                caveats=["Direct dashboard-title interactions may require a worksheet-zone fallback."],
            )
        )

    if category_field:
        directions.append(
            _analysis_direction(
                direction_id="product_mix",
                title="Product Mix",
                business_question=(
                    f"Which {subcategory_field or category_field} are driving {primary_measure}"
                    + (f" and {profit_measure}" if profit_measure != primary_measure else "")
                    + "?"
                ),
                why_it_matters="Useful when leaders want to understand category contribution and margin tradeoffs.",
                recommended_kpis=_unique_strings([primary_measure, profit_measure, quantity_measure]),
                primary_view={
                    "name": "Primary View",
                    "question": f"Which {category_field} is driving {primary_measure}?",
                    "mark_type": "Bar",
                },
                detail_view={
                    "name": "Detail View",
                    "question": f"How do {subcategory_field or category_field} compare inside the selected slice?",
                    "mark_type": "Bar",
                },
                recommended_filters=_unique_strings([date_field, category_field, segment_field]),
                interaction_pattern="Click Primary View to filter Detail View.",
                caveats=["Works best when category fields are clean and stable."],
            )
        )

    if segment_field:
        directions.append(
            _analysis_direction(
                direction_id="segment_health",
                title="Segment Health",
                business_question=f"Which {segment_field} segments are strongest on {primary_measure} and {profit_measure}?",
                why_it_matters="Good for leadership teams comparing consumer, corporate, or enterprise performance.",
                recommended_kpis=_unique_strings([primary_measure, profit_measure, quantity_measure]),
                primary_view={
                    "name": "Primary View",
                    "question": f"Which {segment_field} is driving performance?",
                    "mark_type": "Bar",
                },
                detail_view={
                    "name": "Detail View",
                    "question": f"How does the selected {segment_field} trend over {date_field or 'time'}?",
                    "mark_type": "Line" if date_field else "Bar",
                },
                recommended_filters=_unique_strings([date_field, segment_field, geo_field]),
                interaction_pattern="Click a segment to filter the detail view.",
                caveats=["Segment dashboards are best when the audience cares about customer mix."],
            )
        )

    return directions[:4]


def _build_analysis_brief_payload(run_id: str, schema_summary: dict[str, Any]) -> dict[str, Any]:
    directions = _build_analysis_directions(schema_summary)
    if not directions:
        raise RuntimeError("Could not derive any analysis directions from the current schema.")
    return {
        "run_id": run_id,
        "source_schema": schema_summary.get("selected_primary_object", ""),
        "field_candidates": schema_summary.get("field_candidates", {}),
        "directions": directions,
        "selected_direction_id": directions[0]["id"],
        "selected_direction_title": directions[0]["title"],
        "notes": [
            "V1.1 uses schema-driven heuristics instead of a full exploratory analysis engine.",
            "Choose the direction that best matches the business conversation before drafting the contract.",
        ],
    }


def _build_agent_first_analysis_brief_payload(run_id: str, schema_summary: dict[str, Any]) -> dict[str, Any]:
    return {
        "run_id": run_id,
        "source_schema": schema_summary.get("selected_primary_object", ""),
        "field_candidates": schema_summary.get("field_candidates", {}),
        "required_next_step": (
            "Author 2-4 explicit dashboard directions, set selected_direction_id, "
            "then call finalize_analysis_brief before requesting confirmation."
        ),
        "required_fields": [
            "directions[].id",
            "directions[].title",
            "directions[].business_question",
            "directions[].recommended_kpis",
            "directions[].primary_view",
            "directions[].detail_view",
            "selected_direction_id",
        ],
        "strict_mode_note": (
            "Agent-first mode is fail-closed: the server will not infer directions or "
            "fill missing execution details on your behalf."
        ),
        "directions": [],
        "selected_direction_id": "",
        "selected_direction_title": "",
        "direction_template": {
            "id": "executive_overview",
            "title": "Executive Overview",
            "business_question": "",
            "why_it_matters": "",
            "recommended_kpis": [],
            "primary_view": {
                "name": "Primary View",
                "question": "",
                "mark_type": "",
            },
            "detail_view": {
                "name": "Detail View",
                "question": "",
                "mark_type": "",
            },
            "recommended_filters": [],
            "interaction_pattern": "",
            "caveats": [],
            "contract_seed": {
                "dashboard": {"name": "", "layout_pattern": ""},
                "constraints": {
                    "kpis": [],
                    "filters": [],
                    "interaction_pattern": "",
                    "layout_pattern": "",
                },
                "calculated_fields": [],
                "worksheets": [],
                "actions": [],
            },
        },
        "notes": [
            "Agent-first mode: the MCP client agent must propose the dashboard directions; the server will not infer them from the schema.",
            "Use finalize_analysis_brief to write 2-4 explicit directions plus the selected_direction_id before requesting analysis confirmation.",
        ],
    }


def _selected_analysis_direction(analysis_brief: dict[str, Any]) -> dict[str, Any]:
    selected_id = str(analysis_brief.get("selected_direction_id", "")).strip()
    for direction in analysis_brief.get("directions", []):
        if isinstance(direction, dict) and direction.get("id") == selected_id:
            return direction
    raise RuntimeError("The analysis brief does not have a valid selected_direction_id.")


def _render_schema_summary_markdown(schema_summary: dict[str, Any]) -> str:
    field_candidates = schema_summary.get("field_candidates", {})
    selected_primary_object = schema_summary.get("selected_primary_object", "")
    datasource = schema_summary.get("datasource", {})
    return (
        "# Schema Review\n\n"
        f"- Datasource: `{datasource.get('path', '')}`\n"
        f"- Type: `{datasource.get('type', '')}`\n"
        f"- Selected primary object: `{selected_primary_object or '(not selected)'}`\n"
        f"- Measures: {', '.join(field_candidates.get('measures', []) or ['(none)'])}\n"
        f"- Dimensions: {', '.join(field_candidates.get('dimensions', []) or ['(none)'])}\n"
        f"- Date fields: {', '.join(field_candidates.get('date_fields', []) or ['(none)'])}\n"
        f"- Geo fields: {', '.join(field_candidates.get('geo_fields', []) or ['(none)'])}\n\n"
        "## Notes\n"
        f"{_format_bullets(schema_summary.get('notes', []))}\n"
    )


def _render_analysis_brief_markdown(analysis_brief: dict[str, Any]) -> str:
    direction_lines: list[str] = []
    for direction in analysis_brief.get("directions", []):
        if not isinstance(direction, dict):
            continue
        direction_lines.append(
            f"### {direction.get('title', '')}\n"
            f"- id: `{direction.get('id', '')}`\n"
            f"- Business question: {direction.get('business_question', '')}\n"
            f"- Why it matters: {direction.get('why_it_matters', '')}\n"
            f"- KPIs: {', '.join(direction.get('recommended_kpis', []) or ['(none)'])}\n"
            f"- Primary view: {direction.get('primary_view', {}).get('question', '')}\n"
            f"- Detail view: {direction.get('detail_view', {}).get('question', '')}\n"
            f"- Filters: {', '.join(direction.get('recommended_filters', []) or ['(none)'])}\n"
            f"- Interaction: {direction.get('interaction_pattern', '')}\n"
            f"- Caveats: {'; '.join(direction.get('caveats', []) or ['(none)'])}\n"
        )
    if not direction_lines:
        direction_lines.append(
            "## ACTION REQUIRED\n"
            "No candidate directions have been authored yet.\n"
            "In agent-first mode, the MCP client agent must add 2-4 explicit directions before this stage can be confirmed.\n"
            + (
                f"Next step: {analysis_brief.get('required_next_step', '')}\n"
                if analysis_brief.get("required_next_step")
                else ""
            )
        )
    overrides = {
        "selected_direction_id": analysis_brief.get("selected_direction_id", ""),
    }
    if analysis_brief.get("direction_template"):
        overrides["directions"] = [analysis_brief["direction_template"]]
    return (
        "# Analysis Brief Review\n\n"
        "Choose or author candidate dashboard directions before drafting the contract.\n\n"
        + (
            f"Strict mode note: {analysis_brief.get('strict_mode_note', '')}\n\n"
            if analysis_brief.get("strict_mode_note")
            else ""
        )
        + "\n".join(direction_lines)
        + "\n## Editable Overrides\n"
        "Use this JSON-compatible YAML block to add directions and change the selected direction.\n\n"
        + _json_code_fence(overrides)
        + "\n"
    )


def _render_worksheet_execution_summary(worksheet: dict[str, Any]) -> str:
    encodings = _normalize_encoding_spec(worksheet.get("encodings", {}))
    lines = [
        f"### {worksheet.get('name', '')}",
        f"- Mark type: `{worksheet.get('mark_type', '')}`",
        f"- Priority: `{worksheet.get('priority', '')}`",
        f"- Question: {worksheet.get('question', '')}",
        f"- Dimensions: {', '.join(worksheet.get('dimensions', []) or ['(none)'])}",
        f"- Measures: {', '.join(worksheet.get('measures', []) or ['(none)'])}",
        f"- KPI fields: {', '.join(worksheet.get('kpi_fields', []) or ['(none)'])}",
    ]
    encoding_lines = []
    for key in ("columns", "rows", "measure_values", "tooltip"):
        values = encodings.get(key, [])
        if values:
            encoding_lines.append(f"  - `{key}`: {', '.join(values)}")
    for key in ("color", "label", "detail", "size", "wedge_size", "geographic_field"):
        value = encodings.get(key, "")
        if value:
            encoding_lines.append(f"  - `{key}`: {value}")
    lines.append("- Encodings:")
    lines.append("\n".join(encoding_lines) if encoding_lines else "  - (none)")
    lines.append(f"- Worksheet filters: {', '.join(worksheet.get('filters', []) or ['(none)'])}")
    if worksheet.get("sort_descending"):
        lines.append(f"- Sort descending: `{worksheet.get('sort_descending', '')}`")
    return "\n".join(lines)


def _render_action_execution_summary(action: dict[str, Any]) -> str:
    targets = action.get("targets", []) or ([action.get("target")] if action.get("target") else [])
    return (
        f"- `{action.get('type', '')}` from `{action.get('source', '') or '(unspecified)'}` "
        f"to `{', '.join(targets) if targets else action.get('url', '') or '(none)'}` "
        f"fields: {', '.join(action.get('fields', []) or ['(none)'])} "
        f"caption: {action.get('caption', '') or '(default)'}"
    )


def _render_contract_markdown(contract: dict[str, Any]) -> str:
    constraints = contract.get("constraints", {})
    summary_lines = [
        f"- Goal: {contract.get('goal', '')}",
        f"- Audience: {contract.get('audience', '')}",
        f"- Primary question: {contract.get('primary_question', '')}",
        f"- Dashboard name: {contract.get('dashboard', {}).get('name', '')}",
        f"- KPIs: {', '.join(constraints.get('kpis', []) or ['(none)'])}",
        f"- Filters: {', '.join(constraints.get('filters', []) or ['(none)'])}",
        f"- Interaction required: {contract.get('require_interaction')}",
    ]
    worksheet_sections = [
        _render_worksheet_execution_summary(worksheet)
        for worksheet in contract.get("worksheets", [])
        if isinstance(worksheet, dict)
    ]
    action_lines = [
        _render_action_execution_summary(action)
        for action in contract.get("actions", [])
        if isinstance(action, dict)
    ]
    calculated_field_lines = [
        f"- `{field.get('name', '')}` = {field.get('formula', '') or '(missing formula)'}"
        for field in contract.get("calculated_fields", [])
        if isinstance(field, dict)
    ]
    warning_lines = [
        f"- {warning.get('field_name', '')}: {warning.get('context', '')}"
        for warning in contract.get("resolution_warnings", [])
        if isinstance(warning, dict)
    ]
    return (
        "# Contract Review\n\n"
        + "\n".join(summary_lines)
        + "\n\n## Calculated Fields\n"
        + ("\n".join(calculated_field_lines) if calculated_field_lines else "- (none)")
        + "\n\n## Worksheet Execution Specs\n"
        + ("\n\n".join(worksheet_sections) if worksheet_sections else "- (none)")
        + "\n\n## Actions\n"
        + ("\n".join(action_lines) if action_lines else "- (none)")
        + "\n\n## Resolution Warnings\n"
        + ("\n".join(warning_lines) if warning_lines else "- (none)")
        + "\n\n## Editable Overrides\n"
        "Edit the JSON-compatible YAML block if you want to change the contract before approval.\n\n"
        + _json_code_fence(contract)
        + "\n"
    )


def _worksheet_by_priority(contract: dict[str, Any], priority: str) -> dict[str, Any] | None:
    for worksheet in contract.get("worksheets", []):
        if (
            isinstance(worksheet, dict)
            and str(worksheet.get("priority", "")).strip().casefold() == priority.casefold()
        ):
            return worksheet
    return None


def _normalize_wireframe_actions(
    contract: dict[str, Any],
    field_candidates: dict[str, list[str]],
    *,
    allow_inference: bool,
) -> list[dict[str, Any]]:
    worksheets = [
        worksheet for worksheet in contract.get("worksheets", []) if isinstance(worksheet, dict)
    ]
    worksheet_names = [
        str(worksheet.get("name", "")).strip()
        for worksheet in worksheets
        if str(worksheet.get("name", "")).strip()
    ]
    primary_sheet, detail_sheet = _choose_default_action_sheets(worksheets, worksheet_names)
    summary_sheet = _worksheet_by_priority(contract, "summary")
    summary_name = str(summary_sheet.get("name", "")).strip() if summary_sheet else ""
    action_field = (
        (
            _choose_geo(field_candidates, "")
            if field_candidates.get("geo_fields")
            else _choose_dimension(field_candidates, "")
        )
        if allow_inference
        else ""
    )
    normalized_actions: list[dict[str, Any]] = []

    raw_actions = contract.get("actions") or []
    if allow_inference and not raw_actions and contract.get("require_interaction") and primary_sheet:
        fallback_targets = [name for name in worksheet_names if name != primary_sheet]
        if fallback_targets:
            raw_actions = [
                {
                    "type": "filter",
                    "source": primary_sheet,
                    "targets": fallback_targets,
                    "fields": [action_field] if action_field else [],
                    "caption": f"Filter related views from {primary_sheet}",
                }
            ]

    for action in raw_actions:
        if not isinstance(action, dict):
            continue
        action_type = str(action.get("type", "filter")).strip().casefold() or "filter"
        requested_source = str(action.get("source", "")).strip()
        requested_targets = _dedupe_strings(action.get("targets", []))
        requested_target = str(action.get("target", "")).strip()
        if not requested_targets and requested_target:
            requested_targets = [requested_target]
        resolved_source = requested_source
        resolved_targets = requested_targets[:]
        support_level = "supported"
        note = ""

        if action_type == "filter":
            if not resolved_source:
                if allow_inference:
                    resolved_source = primary_sheet
                    note = "Defaulted filter source to the primary worksheet."
                else:
                    support_level = "unsupported"
                    note = "Filter action is missing an explicit source worksheet."
            if not resolved_targets:
                if allow_inference and resolved_source:
                    resolved_targets = [name for name in worksheet_names if name != resolved_source]
                    note = (note + " " if note else "") + "Defaulted filter targets to the other dashboard worksheets."
                else:
                    support_level = "unsupported"
                    note = (note + " " if note else "") + "Filter action is missing explicit target worksheets."
            resolved_targets = [target for target in resolved_targets if target and target != resolved_source]
            if not resolved_source or not resolved_targets:
                support_level = "unsupported"
                note = "Filter action could not be resolved to at least one target worksheet."
        elif action_type == "url":
            if requested_source.casefold() in {
                "dashboard",
                "dashboard_title",
                "title",
                "title_zone",
                "executive overview",
            }:
                resolved_source = summary_name or primary_sheet or (worksheet_names[0] if worksheet_names else "")
                support_level = "workaround" if resolved_source else "unsupported"
                note = (
                    "Dashboard-title URL actions are modeled as a top worksheet-zone click in V1.1."
                    if resolved_source
                    else "No worksheet zone was available for a URL-action workaround."
                )
            elif not resolved_source:
                resolved_source = summary_name or primary_sheet or (worksheet_names[0] if worksheet_names else "")
                support_level = "workaround" if resolved_source else "unsupported"
                note = (
                    "Defaulted URL action to the top summary zone."
                    if resolved_source
                    else "No worksheet zone was available for a URL action."
                )

        normalized_actions.append(
            {
                "type": action_type,
                "requested_source": requested_source,
                "requested_target": requested_target,
                "requested_targets": requested_targets,
                "source": resolved_source,
                "target": resolved_targets[0] if resolved_targets else "",
                "targets": resolved_targets,
                "fields": (
                    _dedupe_strings(action.get("fields", []))
                    if isinstance(action.get("fields", []), list)
                    else ([str(action.get("fields", "")).strip()] if str(action.get("fields", "")).strip() else [])
                ),
                "caption": action.get("caption", ""),
                "url": action.get("url", ""),
                "support_level": support_level,
                "note": note,
            }
        )

    return normalized_actions


def _ascii_box(lines: list[str], width: int = 70) -> str:
    horizontal = "+" + "-" * width + "+"
    rendered = [horizontal]
    for line in lines:
        content = line[:width].ljust(width)
        rendered.append(f"|{content}|")
    rendered.append(horizontal)
    return "\n".join(rendered)


def _layout_description(
    *,
    layout_pattern: str,
    summary_name: str,
    primary_name: str,
    detail_name: str,
    auxiliary_worksheets: list[str],
) -> str:
    secondary_text = ", ".join(auxiliary_worksheets) if auxiliary_worksheets else "no secondary views"
    return (
        f"Layout pattern '{layout_pattern}': title and KPI strip at the top, "
        f"summary zone '{summary_name}', primary zone '{primary_name}', "
        f"detail zone '{detail_name}', and {secondary_text} beneath them."
    )


def _render_wireframe_ascii(
    *,
    dashboard_name: str,
    layout_pattern: str,
    kpis: list[str],
    summary_name: str,
    primary_name: str,
    primary_question: str,
    detail_name: str,
    detail_question: str,
    auxiliary_worksheets: list[str],
    filters: list[str],
    actions: list[dict[str, Any]],
    support_notes: list[str],
) -> str:
    ascii_lines = [
        f"{dashboard_name} [{layout_pattern}]",
        "",
        "KPI Zone: " + " | ".join(kpis or ["(define KPIs)"]),
        "",
        f"Summary Zone: {summary_name}",
        f"Primary Zone: {primary_name}",
        primary_question or "(question not provided)",
        f"Detail Zone: {detail_name}",
        detail_question or "(question not provided)",
    ]
    if auxiliary_worksheets:
        ascii_lines.extend(["", "Secondary Zones: " + " | ".join(auxiliary_worksheets)])
    ascii_lines.extend(
        [
            "",
            "Filters: " + " | ".join(filters or ["(none)"]),
            "Actions: "
            + " ; ".join(
                [
                    f"{action['type']} {action.get('source', '')}->{','.join(action.get('targets', [])) or action.get('url', '')} [{action.get('support_level', '')}]"
                    for action in actions
                ]
                or ["(none)"]
            ),
            "Notes: " + " ; ".join(support_notes),
        ]
    )
    return _ascii_box(ascii_lines)


def _build_wireframe_payload(
    *,
    run_id: str,
    contract: dict[str, Any],
    schema_summary: dict[str, Any],
    allow_inference: bool,
) -> dict[str, Any]:
    field_candidates = schema_summary.get("field_candidates", {})
    constraints = contract.get("constraints", {})
    layout_pattern = str(
        contract.get("dashboard", {}).get("layout_pattern")
        or constraints.get("layout_pattern", "executive overview")
    ).strip() or "executive overview"
    contract_worksheets = [
        worksheet
        for worksheet in contract.get("worksheets", [])
        if isinstance(worksheet, dict) and str(worksheet.get("name", "")).strip()
    ]
    worksheet_names = _contract_worksheet_names(contract)
    used_zone_names: set[str] = set()

    def _select_zone(priority: str, fallback: dict[str, Any] | None = None) -> dict[str, Any]:
        explicit = _worksheet_by_priority(contract, priority)
        if explicit and str(explicit.get("name", "")).strip():
            used_zone_names.add(str(explicit.get("name", "")).strip())
            return explicit
        for worksheet in contract_worksheets:
            name = str(worksheet.get("name", "")).strip()
            if name and name not in used_zone_names:
                used_zone_names.add(name)
                return worksheet
        return fallback or (contract_worksheets[0] if contract_worksheets else {"name": f"{priority.title()} View", "question": ""})

    summary_view = _select_zone("summary")
    primary_view = _select_zone("primary", fallback=summary_view)
    detail_view = _select_zone("detail", fallback=primary_view)
    summary_name = summary_view.get("name", "Summary View")
    primary_name = primary_view.get("name", "Primary View")
    detail_name = detail_view.get("name", "Detail View")
    auxiliary_worksheets = [
        name
        for name in worksheet_names
        if name not in {str(summary_name), str(primary_name), str(detail_name)}
    ]
    filters = constraints.get("filters", []) or (_suggest_filters(field_candidates) if allow_inference else [])
    actions = _normalize_wireframe_actions(
        contract,
        field_candidates,
        allow_inference=allow_inference,
    )
    support_notes = [action["note"] for action in actions if action.get("note")]
    if not support_notes:
        support_notes = ["No extra workaround notes were needed for the current interaction design."]

    return {
        "run_id": run_id,
        "dashboard_name": contract.get("dashboard", {}).get("name", "Analytical Dashboard"),
        "worksheet_names": worksheet_names,
        "layout_description": _layout_description(
            layout_pattern=layout_pattern,
            summary_name=str(summary_name),
            primary_name=str(primary_name),
            detail_name=str(detail_name),
            auxiliary_worksheets=auxiliary_worksheets,
        ),
        "zones": {
            "title_zone": contract.get("dashboard", {}).get("name", "Analytical Dashboard"),
            "summary_zone": summary_name,
            "primary_zone": primary_name,
            "detail_zone": detail_name,
            "secondary_zones": auxiliary_worksheets,
            "filter_zone": filters,
        },
        "actions": actions,
        "support_notes": support_notes,
        "ascii_wireframe": _render_wireframe_ascii(
            dashboard_name=contract.get("dashboard", {}).get("name", "Analytical Dashboard"),
            layout_pattern=layout_pattern,
            kpis=constraints.get("kpis", []) or [],
            summary_name=str(summary_name),
            primary_name=str(primary_name),
            primary_question=str(primary_view.get("question", "")),
            detail_name=str(detail_name),
            detail_question=str(detail_view.get("question", "")),
            auxiliary_worksheets=auxiliary_worksheets,
            filters=filters,
            actions=actions,
            support_notes=support_notes,
        ),
    }


def _render_wireframe_markdown(wireframe: dict[str, Any]) -> str:
    overrides = {
        "layout_description": wireframe.get("layout_description", ""),
        "support_notes": wireframe.get("support_notes", []),
        "actions": wireframe.get("actions", []),
    }
    zone_lines = [
        f"- Layout: {wireframe.get('layout_description', '')}",
        f"- Summary zone: `{wireframe.get('zones', {}).get('summary_zone', '')}`",
        f"- Primary zone: `{wireframe.get('zones', {}).get('primary_zone', '')}`",
        f"- Detail zone: `{wireframe.get('zones', {}).get('detail_zone', '')}`",
        f"- Secondary zones: {', '.join(wireframe.get('zones', {}).get('secondary_zones', []) or ['(none)'])}",
    ]
    action_lines = [
        f"- `{action.get('type', '')}` `{action.get('source', '')}` -> "
        f"`{', '.join(action.get('targets', []) or ([action.get('target')] if action.get('target') else [])) or action.get('url', '') or '(none)'}` "
        f"[{action.get('support_level', '')}]"
        for action in wireframe.get("actions", [])
        if isinstance(action, dict)
    ]
    return (
        "# Wireframe Review\n\n"
        "Use this ASCII wireframe to confirm the information hierarchy and action bindings.\n\n"
        "```text\n"
        + wireframe.get("ascii_wireframe", "")
        + "\n```\n\n"
        "## Zones\n"
        + "\n".join(zone_lines)
        + "\n\n## Actions\n"
        + ("\n".join(action_lines) if action_lines else "- (none)")
        + "\n\n"
        "## Editable Overrides\n"
        "Edit the JSON-compatible YAML block if you need to adjust notes or normalized actions before approval.\n\n"
        + _json_code_fence(overrides)
        + "\n"
    )


def _render_execution_plan_markdown(plan: dict[str, Any]) -> str:
    def _step_summary(step: dict[str, Any], index: int) -> str:
        tool = str(step.get("tool", "")).strip()
        args = step.get("args", {}) or {}
        if tool == "configure_chart":
            summary_parts = [
                f"worksheet=`{args.get('worksheet_name', '')}`",
                f"mark=`{args.get('mark_type', '')}`",
            ]
            for key in ("measure_values", "columns", "rows", "color", "wedge_size", "geographic_field"):
                value = args.get(key)
                if value:
                    summary_parts.append(f"{key}={value}")
            return f"{index}. `{tool}` " + " | ".join(summary_parts)
        if tool == "add_dashboard_action":
            target = args.get("target_sheet", "") or args.get("url", "")
            return (
                f"{index}. `{tool}` source=`{args.get('source_sheet', '')}` "
                f"target=`{target}` type=`{args.get('action_type', '')}` fields={args.get('fields', [])}"
            )
        if tool == "add_dashboard":
            return (
                f"{index}. `{tool}` dashboard=`{args.get('dashboard_name', '')}` "
                f"worksheets={args.get('worksheet_names', [])}"
            )
        if tool in {"add_worksheet", "set_worksheet_caption"}:
            return f"{index}. `{tool}` worksheet=`{args.get('worksheet_name', '')}`"
        return f"{index}. `{tool}`"

    step_lines = [_step_summary(step, index) for index, step in enumerate(plan.get("steps", []), start=1)]
    return (
        "# Execution Plan Review\n\n"
        "This plan is read-only in V1.1. If you want changes, reopen `contract` or `wireframe`, rebuild, and confirm again.\n\n"
        "## Planned Steps\n"
        + "\n".join(step_lines)
        + "\n\n## Post Checks\n"
        + _format_bullets([check.get("tool", "") for check in plan.get("post_checks", [])])
        + "\n"
    )


def _json_response(**payload: Any) -> str:
    return json.dumps(payload, ensure_ascii=False, indent=2)


def _artifact_names(manifest: dict[str, Any]) -> list[str]:
    names: list[str] = []
    for entry in manifest.get("artifacts", {}).values():
        current = entry.get("current", "")
        if current:
            names.append(current)
        review_current = entry.get("review_current", "")
        if review_current:
            names.append(review_current)
    approvals = Path(manifest["run_dir"]) / APPROVALS_NAME
    if approvals.exists():
        names.append(APPROVALS_NAME)
    if manifest.get("final_workbook"):
        names.append(Path(manifest["final_workbook"]).name)
    return sorted(set(names))


def _safe_current_artifact_payload(manifest: dict[str, Any], artifact_key: str) -> dict[str, Any]:
    entry = _artifact_entry(manifest, artifact_key)
    if not entry.get("current"):
        return {}
    try:
        return _load_current_artifact(manifest, artifact_key)
    except (FileNotFoundError, ValueError):
        return {}


def start_authoring_run(
    datasource_path: str,
    output_dir: str | Path = DEFAULT_AUTHORING_RUNS_DIR,
    resume_if_exists: bool = False,
    authoring_mode: str = AUTHORING_MODE_AGENT_FIRST,
    force_new: bool = False,
) -> str:
    """Create a new guided authoring run rooted in tmp/agentic_run/{run_id}."""

    normalized_path = _normalize_path(datasource_path)
    if not Path(normalized_path).exists():
        raise FileNotFoundError(f"Datasource file not found: {normalized_path}")
    normalized_mode = str(authoring_mode).strip() or AUTHORING_MODE_AGENT_FIRST
    if normalized_mode not in AUTHORING_MODES:
        raise ValueError(
            f"Unsupported authoring_mode '{authoring_mode}'. Expected one of: "
            f"{', '.join(sorted(AUTHORING_MODES))}"
        )

    root_dir = _ensure_dir(Path(output_dir))
    if resume_if_exists and not force_new:
        index_payload = _load_index(root_dir)
        for run_id, info in index_payload.get("runs", {}).items():
            if info.get("datasource_path") == normalized_path:
                manifest_path = Path(info["run_dir"]) / MANIFEST_NAME
                if manifest_path.exists():
                    manifest = _read_json(manifest_path)
                    resumed_mode = _authoring_mode(manifest)
                    mode_mismatch = resumed_mode != normalized_mode
                    return _json_response(
                        run_id=run_id,
                        run_dir=manifest["run_dir"],
                        datasource_path=manifest["datasource_path"],
                        datasource_type=manifest["datasource_type"],
                        status=manifest["status"],
                        authoring_mode=resumed_mode,
                        requested_authoring_mode=normalized_mode,
                        resumed_authoring_mode=resumed_mode,
                        mode_mismatch=mode_mismatch,
                        mode_mismatch_note=(
                            f"Requested authoring_mode '{normalized_mode}' but resumed existing run in "
                            f"'{resumed_mode}'. Use force_new=True to skip resume reuse."
                            if mode_mismatch
                            else ""
                        ),
                        resumed=True,
                        force_new=False,
                    )

    datasource_type = _detect_datasource_type(normalized_path)
    hash_seed = f"{normalized_path}|{_now().timestamp()}".encode("utf-8")
    run_id = f"{_now_token()}-{hashlib.sha1(hash_seed).hexdigest()[:8]}"
    run_dir = _ensure_dir(root_dir / run_id)
    manifest = _default_manifest(
        root_dir,
        run_dir,
        run_id,
        normalized_path,
        datasource_type,
        normalized_mode,
    )
    _write_json(run_dir / MANIFEST_NAME, manifest)
    _write_json(run_dir / APPROVALS_NAME, _empty_approvals())
    _update_index_entry(manifest)
    return _json_response(
        run_id=run_id,
        run_dir=str(run_dir),
        datasource_path=normalized_path,
        datasource_type=datasource_type,
        authoring_mode=normalized_mode,
        status=manifest["status"],
        resumed=False,
        requested_authoring_mode=normalized_mode,
        resumed_authoring_mode=normalized_mode,
        mode_mismatch=False,
        mode_mismatch_note="",
        force_new=force_new,
    )


def list_authoring_runs(output_dir: str | Path = DEFAULT_AUTHORING_RUNS_DIR) -> str:
    """List run ids, timestamps, status, and datasource paths for all runs."""

    root_dir = Path(output_dir)
    index_payload = _load_index(root_dir)
    runs: list[dict[str, Any]] = []
    for run_id, info in sorted(
        index_payload.get("runs", {}).items(),
        key=lambda item: item[1].get("created_at", ""),
        reverse=True,
    ):
        manifest_path = Path(info.get("run_dir", "")) / MANIFEST_NAME
        manifest = _read_json(manifest_path) if manifest_path.exists() else info
        runs.append(
            {
                "run_id": run_id,
                "created_at": manifest.get("created_at", ""),
                "updated_at": manifest.get("updated_at", ""),
                "status": manifest.get("status", ""),
                "authoring_mode": _authoring_mode(manifest),
                "datasource_path": manifest.get("datasource_path", ""),
                "artifacts_present": _artifact_names(manifest),
            }
        )
    return _json_response(runs=runs)


def get_run_status(run_id: str) -> str:
    """Return the current manifest-backed state for one authoring run."""

    manifest = _load_manifest_by_id(run_id)
    return _json_response(
        run_id=manifest["run_id"],
        status=manifest["status"],
        authoring_mode=_authoring_mode(manifest),
        datasource_path=manifest["datasource_path"],
        datasource_type=manifest["datasource_type"],
        selected_primary_object=manifest.get("selected_primary_object", ""),
        created_at=manifest.get("created_at", ""),
        updated_at=manifest.get("updated_at", ""),
        artifacts=manifest.get("artifacts", {}),
        artifacts_present=_artifact_names(manifest),
        pending_confirmation=manifest.get("pending_confirmation", {}),
        final_workbook=manifest.get("final_workbook", ""),
        last_error=manifest.get("last_error", {}),
        resolution_warnings=manifest.get("resolution_warnings", []),
        semantic_validation=_safe_current_artifact_payload(manifest, ARTIFACT_SEMANTIC_VALIDATION),
    )


def resume_authoring_run(run_id: str) -> str:
    """Resume an interrupted authoring run by returning its current status."""

    manifest = _load_manifest_by_id(run_id)
    return _json_response(
        run_id=manifest["run_id"],
        status=manifest["status"],
        authoring_mode=_authoring_mode(manifest),
        run_dir=manifest["run_dir"],
        needs_attention=(manifest["status"] == STATUS_GENERATION_FAILED),
        last_error=manifest.get("last_error", {}),
        resolution_warnings=manifest.get("resolution_warnings", []),
        semantic_validation=_safe_current_artifact_payload(manifest, ARTIFACT_SEMANTIC_VALIDATION),
        artifacts_present=_artifact_names(manifest),
    )


def intake_datasource_schema(run_id: str, preferred_sheet: str = "") -> str:
    """Inspect the manifest datasource and write schema_summary.json + .md."""

    manifest = _load_manifest_by_id(run_id)
    _require_status(manifest, (STATUS_INITIALIZED, STATUS_SCHEMA_INTAKED), "intake datasource schema")

    datasource_path = Path(manifest["datasource_path"])
    if manifest["datasource_type"] == "excel":
        summary = _build_excel_schema_summary(datasource_path, preferred_sheet=preferred_sheet)
    elif manifest["datasource_type"] == "hyper":
        summary = _build_hyper_schema_summary(datasource_path)
    else:
        raise ValueError(f"Unsupported datasource type: {manifest['datasource_type']}")

    manifest["selected_primary_object"] = summary.get("selected_primary_object", "")
    artifact_path = _write_versioned_artifact(
        manifest,
        ARTIFACT_SCHEMA,
        summary,
        markdown_content=_render_schema_summary_markdown(summary),
    )
    _update_manifest(manifest, status=STATUS_SCHEMA_INTAKED, last_error={})
    return _json_response(
        run_id=run_id,
        status=manifest["status"],
        selected_primary_object=manifest.get("selected_primary_object", ""),
        artifact=str(artifact_path),
        review_artifact=str(_current_review_artifact_path(manifest, ARTIFACT_SCHEMA)),
    )


def build_analysis_brief(run_id: str) -> str:
    """Generate a light analyst-copilot brief with 2-4 candidate dashboard directions."""

    manifest = _load_manifest_by_id(run_id)
    _require_status(
        manifest,
        (STATUS_SCHEMA_CONFIRMED, STATUS_ANALYSIS_BUILT, STATUS_ANALYSIS_FINALIZED),
        "build analysis brief",
    )
    schema_summary = _load_current_artifact(manifest, ARTIFACT_SCHEMA)
    if _allow_legacy_inference(manifest):
        payload = _build_analysis_brief_payload(run_id, schema_summary)
    else:
        payload = _build_agent_first_analysis_brief_payload(run_id, schema_summary)
    artifact_path = _write_versioned_artifact(
        manifest,
        ARTIFACT_ANALYSIS_BRIEF,
        payload,
        markdown_content=_render_analysis_brief_markdown(payload),
    )
    _update_manifest(manifest, status=STATUS_ANALYSIS_BUILT, last_error={})
    return _json_response(
        run_id=run_id,
        status=manifest["status"],
        artifact=str(artifact_path),
        review_artifact=str(_current_review_artifact_path(manifest, ARTIFACT_ANALYSIS_BRIEF)),
        direction_count=len(payload.get("directions", [])),
    )


def finalize_analysis_brief(
    run_id: str,
    user_answers_json: str = "",
    markdown_path: str = "",
) -> str:
    """Finalize the selected analysis direction from chat overrides or a Markdown review file."""

    manifest = _load_manifest_by_id(run_id)
    _require_status(
        manifest,
        (STATUS_ANALYSIS_BUILT, STATUS_ANALYSIS_FINALIZED),
        "finalize analysis brief",
    )
    analysis_brief = _load_current_artifact(manifest, ARTIFACT_ANALYSIS_BRIEF)
    overrides = _combined_overrides(
        markdown_path=markdown_path,
        user_answers_json=user_answers_json,
    )
    merged = _deep_merge(analysis_brief, overrides)
    direction = _selected_analysis_direction(merged)
    merged["selected_direction_title"] = direction.get("title", "")
    artifact_path = _write_versioned_artifact(
        manifest,
        ARTIFACT_ANALYSIS_BRIEF,
        merged,
        markdown_content=_render_analysis_brief_markdown(merged),
    )
    _update_manifest(manifest, status=STATUS_ANALYSIS_FINALIZED, last_error={})
    return _json_response(
        run_id=run_id,
        status=manifest["status"],
        selected_direction_id=merged.get("selected_direction_id", ""),
        selected_direction_title=merged.get("selected_direction_title", ""),
        artifact=str(artifact_path),
        review_artifact=str(_current_review_artifact_path(manifest, ARTIFACT_ANALYSIS_BRIEF)),
    )


def draft_authoring_contract(run_id: str, human_brief: str, rewrite: bool = False) -> str:
    """Draft a contract from the human brief plus the current schema summary."""

    manifest = _load_manifest_by_id(run_id)
    allowed = (STATUS_ANALYSIS_CONFIRMED,)
    if rewrite:
        allowed = (STATUS_CONTRACT_REVIEWED, STATUS_CONTRACT_FINALIZED, STATUS_CONTRACT_DRAFTED)
    _require_status(manifest, allowed, "draft authoring contract")

    schema_summary = _load_current_artifact(manifest, ARTIFACT_SCHEMA)
    analysis_brief = _load_current_artifact(manifest, ARTIFACT_ANALYSIS_BRIEF)
    selected_direction = _selected_analysis_direction(analysis_brief)
    seed = selected_direction.get("contract_seed", {}) if isinstance(selected_direction, dict) else {}
    from .config import CONTRACTS_DIR

    template = _read_json(CONTRACTS_DIR / "dashboard_authoring_v1.json")
    field_candidates = schema_summary.get("field_candidates", {})
    brief = human_brief.strip()

    contract = deepcopy(template)
    contract["goal"] = brief
    contract["dataset"] = Path(manifest["datasource_path"]).stem
    contract["workbook_template"] = ""
    contract["available_fields"] = [field["name"] for field in schema_summary.get("fields", [])]
    contract["worksheets"] = []
    contract["actions"] = []
    contract["analysis_direction_id"] = selected_direction.get("id", "")
    contract["analysis_direction_title"] = selected_direction.get("title", "")
    if _allow_legacy_inference(manifest):
        recommended_profiles = schema_summary.get("recommended_profile_matches", [])
        contract["dataset_profile"] = recommended_profiles[0]["id"] if recommended_profiles else ""
        contract["audience"] = _extract_audience(brief)
        contract["primary_question"] = _extract_primary_question(brief) or selected_direction.get(
            "business_question", ""
        )
        interaction_requirement = _infer_interaction_requirement(brief)
        contract["require_interaction"] = (
            interaction_requirement
            if interaction_requirement is not None
            else bool(seed.get("interaction_pattern"))
        )
        contract["worksheets"] = deepcopy(seed.get("worksheets", [])) or _default_worksheet_specs(field_candidates)
        contract["constraints"]["kpis"] = deepcopy(seed.get("kpis", [])) or contract["constraints"].get("kpis", [])
        contract["constraints"]["filters"] = deepcopy(seed.get("filters", [])) or contract["constraints"].get("filters", [])
        contract["constraints"]["interaction_pattern"] = seed.get("interaction_pattern", "")
        contract["constraints"]["layout_pattern"] = seed.get("layout_pattern", "executive overview")
        contract["dashboard"]["name"] = seed.get("dashboard_name", contract["dashboard"].get("name", ""))
        contract["dashboard"]["layout_pattern"] = seed.get(
            "layout_pattern",
            contract["dashboard"].get("layout_pattern", ""),
        )
        if contract["worksheets"]:
            for worksheet in contract["worksheets"]:
                worksheet["mark_type"] = worksheet["mark_type"] or _resolve_mark_type(
                    worksheet.get("question", ""),
                    worksheet.get("priority", ""),
                    field_candidates,
                )
    else:
        if isinstance(seed, dict):
            contract = _deep_merge(contract, seed)
        contract["goal"] = brief
        contract["dataset"] = Path(manifest["datasource_path"]).stem
        contract["workbook_template"] = ""
        contract["available_fields"] = [field["name"] for field in schema_summary.get("fields", [])]
        contract["analysis_direction_id"] = selected_direction.get("id", "")
        contract["analysis_direction_title"] = selected_direction.get("title", "")
    contract = _ensure_contract_execution_spec(
        contract,
        schema_summary,
        fail_on_unresolved=False,
        allow_inference=_allow_legacy_inference(manifest),
    )

    artifact_path = _write_versioned_artifact(manifest, ARTIFACT_CONTRACT_DRAFT, contract)
    _update_manifest(
        manifest,
        status=STATUS_CONTRACT_DRAFTED,
        last_error={},
        resolution_warnings=contract.get("resolution_warnings", []),
    )
    return _json_response(
        run_id=run_id,
        status=manifest["status"],
        artifact=str(artifact_path),
    )


def review_authoring_contract_for_run(run_id: str) -> str:
    """Review the current contract draft and persist the review result."""

    manifest = _load_manifest_by_id(run_id)
    _require_status(manifest, (STATUS_CONTRACT_DRAFTED,), "review contract")
    contract = _load_current_artifact(manifest, ARTIFACT_CONTRACT_DRAFT)
    review = json.loads(
        review_authoring_contract_payload(
            json.dumps(contract),
            allow_profile_defaults=_allow_legacy_inference(manifest),
            strict_execution=not _allow_legacy_inference(manifest),
        ).to_json()
    )
    artifact_path = _write_versioned_artifact(manifest, ARTIFACT_CONTRACT_REVIEW, review)
    _update_manifest(manifest, status=STATUS_CONTRACT_REVIEWED, last_error={})
    return _json_response(
        run_id=run_id,
        status=manifest["status"],
        valid=review["valid"],
        clarification_questions=review["clarification_questions"],
        artifact=str(artifact_path),
    )


def finalize_authoring_contract(
    run_id: str,
    user_answers_json: str = "",
    markdown_path: str = "",
) -> str:
    """Merge review defaults with human overrides and write contract_final.json + .md."""

    manifest = _load_manifest_by_id(run_id)
    _require_status(
        manifest,
        (STATUS_CONTRACT_REVIEWED, STATUS_CONTRACT_FINALIZED),
        "finalize contract",
    )
    review_payload = _load_current_artifact(manifest, ARTIFACT_CONTRACT_REVIEW)
    normalized_contract = deepcopy(review_payload["normalized_contract"])
    overrides = _combined_overrides(
        markdown_path=markdown_path,
        user_answers_json=user_answers_json,
    )
    merged = _deep_merge(normalized_contract, overrides)
    refreshed = json.loads(
        review_authoring_contract_payload(
            json.dumps(merged),
            allow_profile_defaults=_allow_legacy_inference(manifest),
            strict_execution=not _allow_legacy_inference(manifest),
        ).to_json()
    )
    if not refreshed["valid"] and not _allow_legacy_inference(manifest):
        raise RuntimeError(
            "Contract is not executable yet. "
            + " ".join(refreshed.get("clarification_questions", []) or ["Add the missing worksheet/action spec and try again."])
        )
    schema_summary = _load_current_artifact(manifest, ARTIFACT_SCHEMA)
    final_contract = _ensure_contract_execution_spec(
        refreshed["normalized_contract"],
        schema_summary,
        fail_on_unresolved=not _allow_legacy_inference(manifest),
        allow_inference=_allow_legacy_inference(manifest),
    )
    artifact_path = _write_versioned_artifact(
        manifest,
        ARTIFACT_CONTRACT_FINAL,
        final_contract,
        markdown_content=_render_contract_markdown(final_contract),
    )
    _update_manifest(
        manifest,
        status=STATUS_CONTRACT_FINALIZED,
        last_error={},
        resolution_warnings=final_contract.get("resolution_warnings", []),
    )
    return _json_response(
        run_id=run_id,
        status=manifest["status"],
        valid=refreshed["valid"],
        missing_required=refreshed["missing_required"],
        artifact=str(artifact_path),
        review_artifact=str(_current_review_artifact_path(manifest, ARTIFACT_CONTRACT_FINAL)),
    )


def build_wireframe(run_id: str) -> str:
    """Build a human-reviewable ASCII wireframe from the confirmed contract."""

    manifest = _load_manifest_by_id(run_id)
    _require_status(
        manifest,
        (STATUS_CONTRACT_CONFIRMED, STATUS_WIREFRAME_BUILT, STATUS_WIREFRAME_FINALIZED),
        "build wireframe",
    )
    contract = _load_current_artifact(manifest, ARTIFACT_CONTRACT_FINAL)
    schema_summary = _load_current_artifact(manifest, ARTIFACT_SCHEMA)
    payload = _build_wireframe_payload(
        run_id=run_id,
        contract=contract,
        schema_summary=schema_summary,
        allow_inference=_allow_legacy_inference(manifest),
    )
    artifact_path = _write_versioned_artifact(
        manifest,
        ARTIFACT_WIREFRAME,
        payload,
        markdown_content=_render_wireframe_markdown(payload),
    )
    _update_manifest(manifest, status=STATUS_WIREFRAME_BUILT, last_error={})
    return _json_response(
        run_id=run_id,
        status=manifest["status"],
        artifact=str(artifact_path),
        review_artifact=str(_current_review_artifact_path(manifest, ARTIFACT_WIREFRAME)),
    )


def finalize_wireframe(
    run_id: str,
    user_answers_json: str = "",
    markdown_path: str = "",
) -> str:
    """Finalize the wireframe review artifact before execution planning."""

    manifest = _load_manifest_by_id(run_id)
    _require_status(
        manifest,
        (STATUS_WIREFRAME_BUILT, STATUS_WIREFRAME_FINALIZED),
        "finalize wireframe",
    )
    wireframe = _load_current_artifact(manifest, ARTIFACT_WIREFRAME)
    overrides = _combined_overrides(
        markdown_path=markdown_path,
        user_answers_json=user_answers_json,
    )
    merged = _deep_merge(wireframe, overrides)
    artifact_path = _write_versioned_artifact(
        manifest,
        ARTIFACT_WIREFRAME,
        merged,
        markdown_content=_render_wireframe_markdown(merged),
    )
    _update_manifest(manifest, status=STATUS_WIREFRAME_FINALIZED, last_error={})
    return _json_response(
        run_id=run_id,
        status=manifest["status"],
        artifact=str(artifact_path),
        review_artifact=str(_current_review_artifact_path(manifest, ARTIFACT_WIREFRAME)),
    )


def confirm_authoring_stage(run_id: str, stage: str, approved: bool, notes: str = "") -> str:
    """Record a human confirmation decision and advance or roll back state."""

    if stage not in CONFIRMABLE_STAGES:
        raise ValueError(
            f"Unsupported stage '{stage}'. Expected one of: {', '.join(sorted(CONFIRMABLE_STAGES))}"
        )

    manifest = _load_manifest_by_id(run_id)
    _require_pending_confirmation(manifest, stage)
    stage_current: dict[str, Any] = {}
    if stage == SCHEMA_STAGE:
        _require_status(manifest, (STATUS_SCHEMA_INTAKED,), "confirm schema")
        schema_payload = _load_current_artifact(manifest, ARTIFACT_SCHEMA)
        if approved and not schema_payload.get("selected_primary_object"):
            raise RuntimeError(
                "Schema cannot be confirmed until a primary sheet/table is selected."
            )
        next_status = STATUS_SCHEMA_CONFIRMED if approved else STATUS_SCHEMA_INTAKED
        stage_current = {
            "artifact": _artifact_entry(manifest, ARTIFACT_SCHEMA).get("current", ""),
            "selected_primary_object": schema_payload.get("selected_primary_object", ""),
        }
    elif stage == ANALYSIS_STAGE:
        _require_status(manifest, (STATUS_ANALYSIS_FINALIZED,), "confirm analysis brief")
        analysis_payload = _load_current_artifact(manifest, ARTIFACT_ANALYSIS_BRIEF)
        if approved:
            _selected_analysis_direction(analysis_payload)
            next_status = STATUS_ANALYSIS_CONFIRMED
        else:
            next_status = STATUS_ANALYSIS_BUILT
        stage_current = {
            "artifact": _artifact_entry(manifest, ARTIFACT_ANALYSIS_BRIEF).get("current", ""),
            "selected_direction_id": analysis_payload.get("selected_direction_id", ""),
            "selected_direction_title": analysis_payload.get("selected_direction_title", ""),
        }
    elif stage == CONTRACT_STAGE:
        _require_status(
            manifest,
            (STATUS_CONTRACT_FINALIZED, STATUS_CONTRACT_REVIEWED),
            "confirm contract",
        )
        if approved:
            final_contract = _load_current_artifact(manifest, ARTIFACT_CONTRACT_FINAL)
            refreshed = json.loads(
                review_authoring_contract_payload(
                    json.dumps(final_contract),
                    allow_profile_defaults=_allow_legacy_inference(manifest),
                    strict_execution=not _allow_legacy_inference(manifest),
                ).to_json()
            )
            if not refreshed["valid"]:
                raise RuntimeError(
                    "Contract cannot be confirmed until all required intent is captured."
                )
            schema_summary = _load_current_artifact(manifest, ARTIFACT_SCHEMA)
            _ensure_contract_execution_spec(
                final_contract,
                schema_summary,
                fail_on_unresolved=not _allow_legacy_inference(manifest),
                allow_inference=_allow_legacy_inference(manifest),
            )
            next_status = STATUS_CONTRACT_CONFIRMED
            stage_current = {
                "artifact": _artifact_entry(manifest, ARTIFACT_CONTRACT_FINAL).get("current", ""),
                "missing_required": refreshed["missing_required"],
            }
        else:
            next_status = STATUS_CONTRACT_REVIEWED
            stage_current = {
                "artifact": _artifact_entry(manifest, ARTIFACT_CONTRACT_REVIEW).get("current", ""),
                "rewrite_hint": "rewrite" in notes.casefold() or "重写" in notes,
            }
    elif stage == WIREFRAME_STAGE:
        _require_status(manifest, (STATUS_WIREFRAME_FINALIZED,), "confirm wireframe")
        wireframe_payload = _load_current_artifact(manifest, ARTIFACT_WIREFRAME)
        next_status = STATUS_WIREFRAME_CONFIRMED if approved else STATUS_WIREFRAME_BUILT
        stage_current = {
            "artifact": _artifact_entry(manifest, ARTIFACT_WIREFRAME).get("current", ""),
            "support_notes": wireframe_payload.get("support_notes", []),
        }
    else:
        _require_status(manifest, (STATUS_EXECUTION_PLANNED,), "confirm execution plan")
        next_status = STATUS_EXECUTION_CONFIRMED if approved else STATUS_EXECUTION_PLANNED
        stage_current = {
            "artifact": _artifact_entry(manifest, ARTIFACT_EXECUTION_PLAN).get("current", ""),
        }

    approvals = _load_approvals(manifest)
    approvals.setdefault("events", []).append(
        {
            "stage": stage,
            "approved": approved,
            "notes": notes,
            "timestamp": _now_iso(),
            "artifact": stage_current.get("artifact", ""),
        }
    )
    _save_approvals(manifest, approvals)
    resolution_warnings = manifest.get("resolution_warnings", [])
    _update_manifest(
        manifest,
        status=next_status,
        last_error={},
        pending_confirmation={},
        resolution_warnings=resolution_warnings,
    )
    return _json_response(
        run_id=run_id,
        status=manifest["status"],
        stage=stage,
        approved=approved,
        notes=notes,
        current=stage_current,
    )


def reopen_authoring_stage(run_id: str, stage: str, notes: str = "") -> str:
    """Reopen one editable guided-run stage after confirmation or a generation failure."""

    if stage not in {ANALYSIS_STAGE, CONTRACT_STAGE, WIREFRAME_STAGE, EXECUTION_STAGE}:
        raise ValueError(
            "reopen_authoring_stage only supports 'analysis', 'contract', 'wireframe', or 'execution_plan'."
        )

    manifest = _load_manifest_by_id(run_id)
    _require_status(manifest, REOPEN_STATUS_ALLOWED[stage], "reopen authoring stage")
    status_map = {
        ANALYSIS_STAGE: STATUS_ANALYSIS_FINALIZED,
        CONTRACT_STAGE: STATUS_CONTRACT_FINALIZED,
        WIREFRAME_STAGE: STATUS_WIREFRAME_FINALIZED,
        EXECUTION_STAGE: STATUS_EXECUTION_PLANNED,
    }
    artifact_name = _artifact_entry(manifest, STAGE_ARTIFACT_MAP[stage]).get("current", "")
    if not artifact_name:
        raise RuntimeError(f"Cannot reopen '{stage}' because no current artifact exists for that stage.")

    previous_status = manifest.get("status", "")
    cleared_artifacts = _invalidate_downstream_artifacts(manifest, stage)
    approvals = _load_approvals(manifest)
    approvals.setdefault("events", []).append(
        {
            "stage": stage,
            "approved": None,
            "notes": notes,
            "timestamp": _now_iso(),
            "artifact": artifact_name,
            "event_type": "reopen",
            "previous_status": previous_status,
            "cleared_artifacts": cleared_artifacts,
        }
    )
    _save_approvals(manifest, approvals)
    _update_manifest(
        manifest,
        status=status_map[stage],
        last_error={},
        pending_confirmation={},
        resolution_warnings=[],
    )
    return _json_response(
        run_id=run_id,
        status=manifest["status"],
        stage=stage,
        artifact=artifact_name,
        notes=notes,
        previous_status=previous_status,
        cleared_artifacts=cleared_artifacts,
        next_steps=_reopen_next_steps(stage),
    )


def _reopen_next_steps(stage: str) -> list[str]:
    step_map = {
        ANALYSIS_STAGE: [
            "Edit the analysis brief artifact or review markdown.",
            "Call finalize_analysis_brief with the updated directions and selected_direction_id.",
            "Request analysis confirmation again before continuing.",
        ],
        CONTRACT_STAGE: [
            "Update the contract artifact or review markdown.",
            "Call finalize_authoring_contract to rebuild the executable contract.",
            "Request contract confirmation again, then rebuild downstream artifacts.",
        ],
        WIREFRAME_STAGE: [
            "Rebuild or edit the wireframe artifact.",
            "Call finalize_wireframe after the layout and action bindings are updated.",
            "Request wireframe confirmation again before rebuilding the execution plan.",
        ],
        EXECUTION_STAGE: [
            "Rebuild the execution plan from the confirmed contract and wireframe.",
            "Review the read-only execution plan artifact.",
            "Request execution_plan confirmation again before generation.",
        ],
    }
    return step_map.get(stage, [])


def _plan_calculated_fields(
    contract: dict[str, Any],
    available_fields: list[str],
    *,
    fail_on_unresolved: bool,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    requested_fields = _dedupe_strings(contract.get("constraints", {}).get("kpis", []))
    for worksheet in contract.get("worksheets", []):
        if not isinstance(worksheet, dict):
            continue
        requested_fields.extend(_dedupe_strings(worksheet.get("kpi_fields", [])))
        requested_fields.extend(_dedupe_strings(worksheet.get("measures", [])))
    available_lookup = _available_field_lookup(available_fields)
    calculated_field_lookup = _contract_calculated_field_lookup(contract)
    steps: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    for kpi in _dedupe_strings(requested_fields):
        normalized_kpi = _normalize_field_key(kpi)
        if normalized_kpi in available_lookup:
            continue
        contract_calculated = calculated_field_lookup.get(normalized_kpi)
        if contract_calculated is not None:
            formula = str(contract_calculated.get("formula", "")).strip()
            if not formula:
                if fail_on_unresolved:
                    raise RuntimeError(
                        f"Calculated field '{contract_calculated.get('name', kpi)}' is missing a formula. "
                        "Add contract.calculated_fields[].formula before building the execution plan."
                    )
                warnings.append(
                    {
                        "type": "calculated_field_omitted",
                        "field_name": str(contract_calculated.get("name", kpi)).strip(),
                        "context": "contract.calculated_fields",
                        "reason": "missing_formula",
                    }
                )
                continue
            steps.append(
                {
                    "tool": "add_calculated_field",
                    "args": {
                        "field_name": str(contract_calculated.get("name", kpi)).strip(),
                        "formula": formula,
                        "datatype": str(contract_calculated.get("datatype", "real")).strip() or "real",
                    },
                }
            )
            continue
        known_name = _known_calculated_field_name(kpi)
        formula = KNOWN_CALCULATED_FORMULAS.get(known_name)
        if formula:
            steps.append(
                {
                    "tool": "add_calculated_field",
                    "args": {
                        "field_name": known_name,
                        "formula": formula,
                        "datatype": "real",
                    },
                }
            )
            continue
        if fail_on_unresolved:
            raise RuntimeError(
                f"Could not materialize KPI/measure '{kpi}'. Add it to contract.calculated_fields "
                "with an explicit formula before building the execution plan."
            )
        warnings.append(
            {
                "type": "calculated_field_omitted",
                "field_name": kpi,
                "context": "execution_plan",
                "reason": "unknown_formula",
            }
        )
    return steps, warnings


def _build_chart_step(
    worksheet: dict[str, Any],
    field_candidates: dict[str, list[str]],
) -> dict[str, Any]:
    name = worksheet.get("name", "Worksheet")
    mark_type = str(worksheet.get("mark_type", "")).strip()
    encodings = _normalize_encoding_spec(worksheet.get("encodings", {}))

    args: dict[str, Any] = {
        "worksheet_name": name,
        "mark_type": mark_type,
    }
    if not mark_type:
        raise RuntimeError(
            f"Worksheet '{name}' is missing mark_type in the confirmed contract."
        )
    if mark_type == "Map":
        geographic_field = encodings.get("geographic_field", "")
        if not geographic_field:
            raise RuntimeError(
                f"Worksheet '{name}' is missing encodings.geographic_field for a Map chart."
            )
        args["geographic_field"] = geographic_field
        if encodings.get("color"):
            args["color"] = encodings["color"]
        if encodings.get("size"):
            args["size"] = encodings["size"]
        if encodings.get("label"):
            args["label"] = encodings["label"]
        if encodings.get("detail"):
            args["detail"] = encodings["detail"]
        if encodings.get("tooltip"):
            args["tooltip"] = encodings["tooltip"]
    elif mark_type == "Line":
        if not encodings.get("columns") or not encodings.get("rows"):
            raise RuntimeError(
                f"Worksheet '{name}' is missing encodings.columns or encodings.rows for a Line chart."
            )
        args["columns"] = encodings["columns"]
        args["rows"] = encodings["rows"]
        if encodings.get("color"):
            args["color"] = encodings["color"]
        if encodings.get("label"):
            args["label"] = encodings["label"]
        if encodings.get("detail"):
            args["detail"] = encodings["detail"]
        if encodings.get("tooltip"):
            args["tooltip"] = encodings["tooltip"]
    elif mark_type == "Text":
        if not encodings.get("measure_values"):
            raise RuntimeError(
                f"Worksheet '{name}' is missing encodings.measure_values for a Text/KPI chart."
            )
        args["measure_values"] = encodings["measure_values"]
        if encodings.get("color"):
            args["color"] = encodings["color"]
        if encodings.get("label"):
            args["label"] = encodings["label"]
        if encodings.get("detail"):
            args["detail"] = encodings["detail"]
        if encodings.get("tooltip"):
            args["tooltip"] = encodings["tooltip"]
    elif mark_type == "Pie":
        if not encodings.get("color") or not encodings.get("wedge_size"):
            raise RuntimeError(
                f"Worksheet '{name}' is missing encodings.color or encodings.wedge_size for a Pie chart."
            )
        args["color"] = encodings["color"]
        args["wedge_size"] = encodings["wedge_size"]
        if encodings.get("label"):
            args["label"] = encodings["label"]
        if encodings.get("detail"):
            args["detail"] = encodings["detail"]
        if encodings.get("tooltip"):
            args["tooltip"] = encodings["tooltip"]
    else:
        if not encodings.get("rows") or not encodings.get("columns"):
            raise RuntimeError(
                f"Worksheet '{name}' is missing encodings.rows or encodings.columns for a {mark_type} chart."
            )
        args["rows"] = encodings["rows"]
        args["columns"] = encodings["columns"]
        if encodings.get("color"):
            args["color"] = encodings["color"]
        if encodings.get("size"):
            args["size"] = encodings["size"]
        if encodings.get("label"):
            args["label"] = encodings["label"]
        if encodings.get("detail"):
            args["detail"] = encodings["detail"]
        if encodings.get("tooltip"):
            args["tooltip"] = encodings["tooltip"]
        if worksheet.get("sort_descending"):
            args["sort_descending"] = worksheet["sort_descending"]

    return {"tool": "configure_chart", "args": args}


def _choose_default_action_sheets(
    worksheets: list[dict[str, Any]],
    worksheet_names: list[str],
) -> tuple[str, str]:
    primary_name = ""
    detail_name = ""

    for worksheet in worksheets:
        if not isinstance(worksheet, dict):
            continue
        name = str(worksheet.get("name", "")).strip()
        if not name or name not in worksheet_names:
            continue
        priority = str(worksheet.get("priority", "")).strip().casefold()
        if priority == "primary" and not primary_name:
            primary_name = name
        elif priority == "detail" and not detail_name:
            detail_name = name

    non_summary_names = [
        str(worksheet.get("name", "")).strip()
        for worksheet in worksheets
        if isinstance(worksheet, dict)
        and str(worksheet.get("name", "")).strip() in worksheet_names
        and str(worksheet.get("priority", "")).strip().casefold() != "summary"
    ]

    if not primary_name and non_summary_names:
        primary_name = non_summary_names[0]
    if not detail_name and len(non_summary_names) >= 2:
        detail_name = non_summary_names[1]

    if not primary_name and worksheet_names:
        primary_name = worksheet_names[0]
    if not detail_name and len(worksheet_names) >= 2:
        detail_name = worksheet_names[1]

    return primary_name, detail_name


def _contract_worksheet_names(contract: dict[str, Any]) -> list[str]:
    names: list[str] = []
    seen: set[str] = set()
    for worksheet in contract.get("worksheets", []):
        if not isinstance(worksheet, dict):
            continue
        name = str(worksheet.get("name", "")).strip()
        if not name:
            continue
        if name in seen:
            raise RuntimeError(
                f"Contract contains duplicate worksheet name '{name}'. "
                "Reopen the contract stage and make worksheet names unique before planning."
            )
        seen.add(name)
        names.append(name)
    if not names:
        raise RuntimeError(
            "The confirmed contract does not contain any worksheets to plan. "
            "Reopen the contract stage and add at least one worksheet."
        )
    return names


def _wireframe_mentions_new_scope(wireframe: dict[str, Any]) -> list[str]:
    flagged: list[str] = []
    for note in wireframe.get("support_notes", []):
        text = str(note).strip()
        lower = text.casefold()
        if not lower:
            continue
        if any(prefix in lower for prefix in SCOPE_CHANGE_NOTE_PREFIXES) and any(
            obj in lower for obj in SCOPE_CHANGE_NOTE_OBJECTS
        ):
            flagged.append(text)
    return flagged


def _validate_execution_scope(contract: dict[str, Any], wireframe: dict[str, Any]) -> list[str]:
    contract_names = _contract_worksheet_names(contract)
    contract_name_set = set(contract_names)
    errors: list[str] = []

    wireframe_names = _unique_strings(
        [
            str(name).strip()
            for name in wireframe.get("worksheet_names", [])
            if isinstance(name, str) and str(name).strip()
        ]
    )
    if wireframe_names:
        extras = [name for name in wireframe_names if name not in contract_name_set]
        missing = [name for name in contract_names if name not in set(wireframe_names)]
        if extras or missing:
            details: list[str] = []
            if extras:
                details.append("extra worksheets in wireframe: " + ", ".join(extras))
            if missing:
                details.append("missing worksheets from wireframe: " + ", ".join(missing))
            errors.append(
                "Wireframe scope no longer matches the confirmed contract (" + "; ".join(details) + ")."
            )

    zones = wireframe.get("zones", {}) if isinstance(wireframe.get("zones", {}), dict) else {}
    for key in ("summary_zone", "primary_zone", "detail_zone"):
        name = str(zones.get(key, "")).strip()
        if name and name not in contract_name_set:
            errors.append(
                f"Wireframe zone '{key}' points to worksheet '{name}', which is not in the confirmed contract."
            )
    secondary = zones.get("secondary_zones", [])
    secondary_names = secondary if isinstance(secondary, list) else [secondary]
    for name in _unique_strings([str(item).strip() for item in secondary_names if str(item).strip()]):
        if name not in contract_name_set:
            errors.append(
                f"Wireframe secondary zone '{name}' is not in the confirmed contract."
            )

    for action in wireframe.get("actions", []):
        if not isinstance(action, dict) or action.get("support_level") == "unsupported":
            continue
        action_type = str(action.get("type", "")).strip().casefold()
        source = str(action.get("source", "")).strip()
        targets = _dedupe_strings(action.get("targets", []))
        if not targets and action.get("target"):
            targets = [str(action.get("target", "")).strip()]
        if source and source not in contract_name_set:
            errors.append(
                f"Wireframe action source '{source}' is not in the confirmed contract."
            )
        if action_type == "filter":
            for target in targets:
                if target and target not in contract_name_set:
                    errors.append(
                        f"Wireframe action target '{target}' is not in the confirmed contract."
                    )

    contract_action_set: set[tuple[str, str, str]] = set()
    for action in contract.get("actions", []):
        if not isinstance(action, dict):
            continue
        action_type = str(action.get("type", "")).strip().casefold()
        source = str(action.get("source", "")).strip()
        if action_type == "url":
            contract_action_set.add((action_type, source, str(action.get("url", "")).strip()))
            continue
        targets = _dedupe_strings(action.get("targets", []))
        if not targets and action.get("target"):
            targets = [str(action.get("target", "")).strip()]
        for target in targets:
            contract_action_set.add((action_type, source, target))

    wireframe_action_set: set[tuple[str, str, str]] = set()
    for action in wireframe.get("actions", []):
        if not isinstance(action, dict) or action.get("support_level") == "unsupported":
            continue
        action_type = str(action.get("type", "")).strip().casefold()
        source = str(action.get("source", "")).strip()
        if action_type == "url":
            wireframe_action_set.add((action_type, source, str(action.get("url", "")).strip()))
            continue
        targets = _dedupe_strings(action.get("targets", []))
        if not targets and action.get("target"):
            targets = [str(action.get("target", "")).strip()]
        for target in targets:
            wireframe_action_set.add((action_type, source, target))

    missing_contract_actions = sorted(contract_action_set - wireframe_action_set)
    if missing_contract_actions:
        errors.append(
            "Wireframe no longer includes all confirmed contract actions: "
            + "; ".join(f"{item[0]} {item[1]} -> {item[2]}" for item in missing_contract_actions)
        )

    scope_drift_notes = _wireframe_mentions_new_scope(wireframe)
    if scope_drift_notes:
        errors.append(
            "Wireframe notes are requesting new worksheets or views after contract confirmation: "
            + "; ".join(scope_drift_notes)
        )

    if errors:
        raise RuntimeError(
            "Execution plan scope drift detected. "
            + " ".join(errors)
            + " Reopen the contract stage and update the confirmed contract before rebuilding the wireframe or execution plan."
        )

    return contract_names


def build_execution_plan(run_id: str) -> str:
    """Create a mechanical MCP tool sequence from the current final contract."""

    manifest = _load_manifest_by_id(run_id)
    _require_status(manifest, (STATUS_WIREFRAME_CONFIRMED,), "build execution plan")
    schema_summary = _load_current_artifact(manifest, ARTIFACT_SCHEMA)
    contract = _ensure_contract_execution_spec(
        _load_current_artifact(manifest, ARTIFACT_CONTRACT_FINAL),
        schema_summary,
        fail_on_unresolved=not _allow_legacy_inference(manifest),
        allow_inference=_allow_legacy_inference(manifest),
    )
    wireframe = _load_current_artifact(manifest, ARTIFACT_WIREFRAME)
    worksheet_names = _validate_execution_scope(contract, wireframe)
    field_candidates = schema_summary.get("field_candidates", {})
    available_fields = [field["name"] for field in schema_summary.get("fields", [])]
    workbook_template = contract.get("workbook_template", "") or ""

    steps: list[dict[str, Any]] = [
        {
            "tool": "create_workbook",
            "args": {
                "template_path": workbook_template,
                "workbook_name": contract.get("dashboard", {}).get("name", ""),
            },
        }
    ]

    if manifest["datasource_type"] == "excel":
        steps.append(
            {
                "tool": "set_excel_connection",
                "args": {
                    "filepath": manifest["datasource_path"],
                    "sheet_name": schema_summary.get("selected_primary_object", ""),
                    "fields": schema_summary.get("fields", []),
                },
            }
        )
    elif manifest["datasource_type"] == "hyper":
        steps.append(
            {
                "tool": "set_hyper_connection",
                "args": {
                    "filepath": manifest["datasource_path"],
                    "table_name": schema_summary.get("selected_primary_object", "Extract"),
                    "tables": [
                        {
                            "name": table["name"],
                            "columns": [field["name"] for field in table.get("fields", [])],
                        }
                        for table in schema_summary.get("tables", [])
                    ],
                },
            }
        )

    calculated_field_steps, calculated_field_warnings = _plan_calculated_fields(
        contract,
        available_fields,
        fail_on_unresolved=not _allow_legacy_inference(manifest),
    )
    steps.extend(calculated_field_steps)

    for worksheet in contract.get("worksheets", []):
        if not isinstance(worksheet, dict):
            continue
        name = str(worksheet.get("name", "")).strip()
        if not name:
            continue
        steps.append({"tool": "add_worksheet", "args": {"worksheet_name": name}})
        steps.append(_build_chart_step(worksheet, field_candidates))
        caption = str(worksheet.get("question", "")).strip()
        if caption:
            steps.append(
                {
                    "tool": "set_worksheet_caption",
                    "args": {"worksheet_name": name, "caption": caption},
                }
            )

    dashboard_name = contract.get("dashboard", {}).get("name") or "Analytical Dashboard"
    steps.append(
        {
            "tool": "add_dashboard",
            "args": {
                "dashboard_name": dashboard_name,
                "worksheet_names": worksheet_names,
                "layout": contract.get("dashboard", {}).get("layout_pattern")
                or contract.get("constraints", {}).get("layout_pattern", "vertical"),
            },
        }
    )

    actions = wireframe.get("actions") or contract.get("actions") or []
    if actions:
        for action in actions:
            if not isinstance(action, dict) or action.get("support_level") == "unsupported":
                continue
            action_type = str(action.get("type", "filter")).strip() or "filter"
            source_sheet = action.get("source", worksheet_names[0] if worksheet_names else "")
            targets = _dedupe_strings(action.get("targets", []))
            if not targets and action.get("target"):
                targets = [str(action.get("target", "")).strip()]
            if action_type == "url":
                steps.append(
                    {
                        "tool": "add_dashboard_action",
                        "args": {
                            "dashboard_name": dashboard_name,
                            "action_type": action_type,
                            "source_sheet": source_sheet,
                            "target_sheet": "",
                            "fields": action.get("fields", []),
                            "caption": action.get("caption", ""),
                            "url": action.get("url", ""),
                        },
                    }
                )
                continue
            if not targets:
                raise RuntimeError(
                    f"Action '{action_type}' from '{source_sheet}' is missing targets in the confirmed wireframe."
                )
            for target_sheet in targets:
                steps.append(
                    {
                        "tool": "add_dashboard_action",
                        "args": {
                            "dashboard_name": dashboard_name,
                            "action_type": action_type,
                            "source_sheet": source_sheet,
                            "target_sheet": target_sheet,
                            "fields": action.get("fields", []),
                            "caption": action.get("caption", ""),
                            "url": action.get("url", ""),
                        },
                    }
                )
    elif contract.get("require_interaction") and len(worksheet_names) >= 2:
        if not _allow_legacy_inference(manifest):
            raise RuntimeError(
                "The confirmed contract requires interaction, but no explicit actions were provided. "
                "Reopen the contract stage and add explicit action specs before building the execution plan."
            )
        source_sheet, target_sheet = _choose_default_action_sheets(
            [
                worksheet
                for worksheet in contract.get("worksheets", [])
                if isinstance(worksheet, dict)
            ],
            worksheet_names,
        )
        action_field = (
            _choose_geo(field_candidates)
            if field_candidates.get("geo_fields")
            else _choose_dimension(field_candidates)
        )
        if source_sheet and target_sheet and source_sheet != target_sheet:
            fallback_targets = [name for name in worksheet_names if name != source_sheet]
            for target_sheet in fallback_targets:
                steps.append(
                    {
                        "tool": "add_dashboard_action",
                        "args": {
                            "dashboard_name": dashboard_name,
                            "action_type": "filter",
                            "source_sheet": source_sheet,
                            "target_sheet": target_sheet,
                            "fields": [action_field] if action_field else [],
                            "caption": f"Filter {target_sheet} from {source_sheet}",
                        },
                    }
                )

    plan = {
        "run_id": run_id,
        "source_contract": _artifact_entry(manifest, ARTIFACT_CONTRACT_FINAL).get("current", ""),
        "source_wireframe": _artifact_entry(manifest, ARTIFACT_WIREFRAME).get("current", ""),
        "workbook_template": workbook_template,
        "resolution_warnings": _dedupe_resolution_warnings(
            contract.get("resolution_warnings", []) + calculated_field_warnings
        ),
        "steps": steps,
        "post_checks": [
            {"tool": "validate_workbook", "args": {}},
            {"tool": "analyze_twb", "args": {}},
        ],
    }
    artifact_path = _write_versioned_artifact(
        manifest,
        ARTIFACT_EXECUTION_PLAN,
        plan,
        markdown_content=_render_execution_plan_markdown(plan),
    )
    _update_manifest(
        manifest,
        status=STATUS_EXECUTION_PLANNED,
        last_error={},
        resolution_warnings=plan.get("resolution_warnings", []),
    )
    return _json_response(
        run_id=run_id,
        status=manifest["status"],
        artifact=str(artifact_path),
        review_artifact=str(_current_review_artifact_path(manifest, ARTIFACT_EXECUTION_PLAN)),
        step_count=len(steps),
    )


def _workbook_root(path: str | Path) -> ET.Element:
    workbook_path = Path(path)
    if workbook_path.suffix.casefold() == ".twbx":
        with ZipFile(workbook_path, "r") as archive:
            workbook_entries = [name for name in archive.namelist() if name.casefold().endswith(".twb")]
            if not workbook_entries:
                raise RuntimeError(f"Packaged workbook '{workbook_path}' does not contain a .twb payload.")
            return ET.fromstring(archive.read(workbook_entries[0]))
    return ET.parse(workbook_path).getroot()


def _worksheet_map(root: ET.Element) -> dict[str, ET.Element]:
    return {
        worksheet.get("name", ""): worksheet
        for worksheet in root.findall(".//worksheet")
        if worksheet.get("name")
    }


def _worksheet_mark_class(worksheet_el: ET.Element) -> str:
    mark = worksheet_el.find(".//pane/mark")
    return mark.get("class", "") if mark is not None else ""


def _worksheet_encoding_values(worksheet_el: ET.Element, encoding_name: str) -> list[str]:
    values: list[str] = []
    for encoding in worksheet_el.findall(f".//pane/encodings/{encoding_name}"):
        column = encoding.get("column", "")
        if column:
            values.append(column)
    return values


def _worksheet_shelf_text(worksheet_el: ET.Element, shelf: str) -> str:
    return (worksheet_el.findtext(f"./table/{shelf}") or "").strip()


def _worksheet_measure_value_members(worksheet_el: ET.Element) -> list[str]:
    members: list[str] = []
    for filter_el in worksheet_el.findall(".//view/filter"):
        if ":Measure Names" not in filter_el.get("column", ""):
            continue
        for member in filter_el.findall(".//groupfilter[@function='member']"):
            value = member.get("member", "")
            if value:
                members.append(value)
    return members


def _instance_name_from_reference(reference: str) -> str:
    text = str(reference).strip().strip('"')
    marker = text.rfind(".[")
    if marker == -1:
        return ""
    return text[marker + 1 :]


def _worksheet_measure_value_labels(worksheet_el: ET.Element) -> list[str]:
    labels = _worksheet_measure_value_members(worksheet_el)
    columns_by_name: dict[str, str] = {}
    for column in worksheet_el.findall(".//view/datasource-dependencies/column"):
        name = column.get("name", "")
        caption = column.get("caption", "").strip() or name.strip("[]")
        if name:
            columns_by_name[name] = caption

    captions: list[str] = []
    for column_instance in worksheet_el.findall(".//view/datasource-dependencies/column-instance"):
        instance_name = column_instance.get("name", "")
        column_name = column_instance.get("column", "")
        if not instance_name or not column_name:
            continue
        caption = columns_by_name.get(column_name, column_name.strip("[]"))
        for member in labels:
            if _instance_name_from_reference(member) == instance_name and caption not in captions:
                captions.append(caption)

    return labels + captions


def _expression_field_token(value: str) -> str:
    text = str(value).strip()
    if not text:
        return ""
    if text in KNOWN_CALCULATED_FORMULAS:
        return _normalize_field_key(text)
    if "(" in text and text.endswith(")"):
        inner = text[text.find("(") + 1 : -1].strip()
        return _normalize_field_key(inner.strip("[]"))
    return _normalize_field_key(text.strip("[]"))


def _encoding_matches(expected: str, actual_values: list[str]) -> bool:
    token = _expression_field_token(expected)
    if not token:
        return False
    return any(token in _normalize_field_key(value) for value in actual_values)


def _dashboard_sheet_names(root: ET.Element, dashboard_name: str) -> list[str]:
    dashboard = root.find(f".//dashboards/dashboard[@name='{dashboard_name}']")
    if dashboard is None:
        return []
    sheets: list[str] = []
    for zone in dashboard.findall(".//zone"):
        name = zone.get("name", "")
        if name and name not in sheets:
            sheets.append(name)
    return sheets


def _collect_dashboard_actions(root: ET.Element, dashboard_name: str) -> list[dict[str, Any]]:
    dashboard_sheets = _dashboard_sheet_names(root, dashboard_name)
    actions: list[dict[str, Any]] = []
    for action_el in root.findall("./actions/action"):
        source = action_el.find("source")
        source_sheet = source.get("worksheet", "") if source is not None else ""
        caption = action_el.get("caption", "")
        command = action_el.find("command")
        link = action_el.find("link")
        if command is not None:
            command_name = command.get("command", "")
            if command_name == "tsc:tsl-filter":
                exclude_value = ""
                for param in command.findall("param"):
                    if param.get("name") == "exclude":
                        exclude_value = param.get("value", "")
                        break
                excluded = {item.strip() for item in exclude_value.split(",") if item.strip()}
                targets = [sheet for sheet in dashboard_sheets if sheet not in excluded]
                for target in targets:
                    actions.append(
                        {
                            "type": "filter",
                            "source_sheet": source_sheet,
                            "target_sheet": target,
                            "caption": caption,
                        }
                    )
            elif command_name == "tsc:brush":
                exclude_value = ""
                for param in command.findall("param"):
                    if param.get("name") == "exclude":
                        exclude_value = param.get("value", "")
                        break
                excluded = {item.strip() for item in exclude_value.split(",") if item.strip()}
                targets = [sheet for sheet in dashboard_sheets if sheet not in excluded]
                for target in targets:
                    actions.append(
                        {
                            "type": "highlight",
                            "source_sheet": source_sheet,
                            "target_sheet": target,
                            "caption": caption,
                        }
                    )
            elif command_name == "tabdoc:goto-sheet":
                target_sheet = ""
                for param in command.findall("param"):
                    if param.get("name") == "target":
                        target_sheet = param.get("value", "")
                        break
                actions.append(
                    {
                        "type": "go-to-sheet",
                        "source_sheet": source_sheet,
                        "target_sheet": target_sheet,
                        "caption": caption,
                    }
                )
        elif link is not None:
            actions.append(
                {
                    "type": "url",
                    "source_sheet": source_sheet,
                    "url": link.get("expression", ""),
                    "caption": caption,
                }
            )
    return actions


def validate_generated_workbook_semantics(run_id: str, workbook_path: str) -> dict[str, Any]:
    """Fail closed when the generated workbook does not match the confirmed contract."""

    manifest = _load_manifest_by_id(run_id)
    schema_summary = _load_current_artifact(manifest, ARTIFACT_SCHEMA)
    contract = _ensure_contract_execution_spec(
        _load_current_artifact(manifest, ARTIFACT_CONTRACT_FINAL),
        schema_summary,
        fail_on_unresolved=not _allow_legacy_inference(manifest),
        allow_inference=_allow_legacy_inference(manifest),
    )
    root = _workbook_root(workbook_path)
    worksheets = _worksheet_map(root)
    errors: list[str] = []
    dashboard_name = str(contract.get("dashboard", {}).get("name", "")).strip()
    expected_sheets = _contract_worksheet_names(contract)
    actual_sheets = _dashboard_sheet_names(root, dashboard_name)
    for sheet_name in expected_sheets:
        if sheet_name not in actual_sheets:
            errors.append(f"Dashboard '{dashboard_name}' is missing worksheet '{sheet_name}'.")

    for worksheet in contract.get("worksheets", []):
        if not isinstance(worksheet, dict):
            continue
        args = {"worksheet_name": worksheet.get("name", ""), "mark_type": worksheet.get("mark_type", "")}
        args.update(_normalize_encoding_spec(worksheet.get("encodings", {})))
        worksheet_name = str(worksheet.get("name", "")).strip()
        worksheet_el = worksheets.get(worksheet_name)
        if worksheet_el is None:
            errors.append(f"Configured worksheet '{worksheet_name}' was not found in the generated workbook.")
            continue

        mark_type = str(worksheet.get("mark_type", "")).strip()
        actual_mark = _worksheet_mark_class(worksheet_el)
        expected_mark = "Multipolygon" if mark_type == "Map" else mark_type
        if expected_mark and actual_mark != expected_mark:
            errors.append(
                f"Worksheet '{worksheet_name}' expected mark '{expected_mark}' but found '{actual_mark or '(none)'}'."
            )

        if mark_type == "Text":
            members = _worksheet_measure_value_labels(worksheet_el)
            for expr in worksheet.get("encodings", {}).get("measure_values", []) or []:
                if not _encoding_matches(expr, members):
                    errors.append(
                        f"Worksheet '{worksheet_name}' is missing KPI/measure value '{expr}'."
                    )
        elif mark_type == "Pie":
            if not _encoding_matches(worksheet.get("encodings", {}).get("color", ""), _worksheet_encoding_values(worksheet_el, "color")):
                errors.append(
                    f"Worksheet '{worksheet_name}' is missing pie color encoding '{worksheet.get('encodings', {}).get('color', '')}'."
                )
            if not _encoding_matches(worksheet.get("encodings", {}).get("wedge_size", ""), _worksheet_encoding_values(worksheet_el, "wedge-size")):
                errors.append(
                    f"Worksheet '{worksheet_name}' is missing pie wedge-size encoding '{worksheet.get('encodings', {}).get('wedge_size', '')}'."
                )
        elif mark_type == "Map":
            rows_text = _worksheet_shelf_text(worksheet_el, "rows")
            cols_text = _worksheet_shelf_text(worksheet_el, "cols")
            if "latitude (generated)" not in rows_text.casefold() or "longitude (generated)" not in cols_text.casefold():
                errors.append(
                    f"Worksheet '{worksheet_name}' is missing generated latitude/longitude shelves for a map."
                )
            lod_values = _worksheet_encoding_values(worksheet_el, "lod")
            if not _encoding_matches(worksheet.get("encodings", {}).get("geographic_field", ""), lod_values):
                errors.append(
                    f"Worksheet '{worksheet_name}' is missing map geography binding '{worksheet.get('encodings', {}).get('geographic_field', '')}'."
                )
            if worksheet.get("encodings", {}).get("color") and not _encoding_matches(
                worksheet.get("encodings", {}).get("color", ""),
                _worksheet_encoding_values(worksheet_el, "color"),
            ):
                errors.append(
                    f"Worksheet '{worksheet_name}' is missing map color encoding '{worksheet.get('encodings', {}).get('color', '')}'."
                )
        else:
            rows_text = _worksheet_shelf_text(worksheet_el, "rows")
            cols_text = _worksheet_shelf_text(worksheet_el, "cols")
            for expr in worksheet.get("encodings", {}).get("rows", []) or []:
                if _expression_field_token(expr) not in _normalize_field_key(rows_text):
                    errors.append(
                        f"Worksheet '{worksheet_name}' rows do not include '{expr}'."
                    )
            for expr in worksheet.get("encodings", {}).get("columns", []) or []:
                if _expression_field_token(expr) not in _normalize_field_key(cols_text):
                    errors.append(
                        f"Worksheet '{worksheet_name}' columns do not include '{expr}'."
                    )

    if dashboard_name:
        actual_actions = _collect_dashboard_actions(root, dashboard_name)
        for action in contract.get("actions", []):
            if not isinstance(action, dict):
                continue
            expected_type = str(action.get("type", "")).strip()
            expected_source = str(action.get("source", "")).strip()
            expected_url = str(action.get("url", "")).strip()
            targets = _dedupe_strings(action.get("targets", []))
            if expected_type == "url":
                targets = [""]
            elif not targets and action.get("target"):
                targets = [str(action.get("target", "")).strip()]

            for expected_target in targets:
                matched = any(
                    actual_action.get("type") == expected_type
                    and actual_action.get("source_sheet") == expected_source
                    and (
                        (expected_type == "url" and actual_action.get("url", "") == expected_url)
                        or (expected_type != "url" and actual_action.get("target_sheet", "") == expected_target)
                    )
                    for actual_action in actual_actions
                )
                if matched:
                    continue
                if expected_type == "url":
                    errors.append(
                        f"Dashboard action '{expected_type}' from '{expected_source}' to '{expected_url}' is missing."
                    )
                else:
                    errors.append(
                        f"Dashboard action '{expected_type}' from '{expected_source}' to '{expected_target}' is missing."
                    )

    if errors:
        payload = {
            "status": "semantic_validation_failed",
            "workbook_path": str(workbook_path),
            "dashboard_name": dashboard_name,
            "worksheet_count": len(worksheets),
            "error_count": len(errors),
            "errors": errors,
            "message": "Workbook semantic validation failed: " + " | ".join(errors),
        }
        raise SemanticValidationError(payload)

    return {
        "status": "semantic_validation_passed",
        "workbook_path": str(workbook_path),
        "worksheet_count": len(worksheets),
        "dashboard_name": dashboard_name,
        "error_count": 0,
        "errors": [],
    }


def load_execution_plan_for_run(run_id: str) -> dict[str, Any]:
    """Load the current execution plan and enforce confirmation state."""

    manifest = _load_manifest_by_id(run_id)
    _require_status(
        manifest,
        (STATUS_EXECUTION_CONFIRMED, STATUS_GENERATION_STARTED, STATUS_GENERATION_FAILED),
        "load execution plan",
    )
    return _load_current_artifact(manifest, ARTIFACT_EXECUTION_PLAN)


def mark_generation_started(run_id: str) -> dict[str, Any]:
    """Set the run to workbook_generation_started and return the manifest."""

    manifest = _load_manifest_by_id(run_id)
    _require_status(manifest, (STATUS_EXECUTION_CONFIRMED,), "generate workbook")
    _update_manifest(manifest, status=STATUS_GENERATION_STARTED, last_error={})
    return manifest


def mark_generation_failed(
    run_id: str,
    step_tool: str,
    error_message: str,
    details: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Persist a generation failure payload and return the manifest."""

    manifest = _load_manifest_by_id(run_id)
    payload = {
        "failed_at": _now_iso(),
        "step_tool": step_tool,
        "message": error_message,
    }
    if details:
        payload["details"] = details
        if details.get("status", "").startswith("semantic_validation_"):
            payload["semantic_validation"] = details
    _update_manifest(
        manifest,
        status=STATUS_GENERATION_FAILED,
        last_error=payload,
    )
    return manifest


def mark_generation_success(run_id: str, final_workbook: str) -> dict[str, Any]:
    """Persist a successful workbook generation result and return the manifest."""

    manifest = _load_manifest_by_id(run_id)
    manifest["final_workbook"] = final_workbook
    _update_manifest(
        manifest,
        status=STATUS_GENERATED,
        last_error={},
        resolution_warnings=manifest.get("resolution_warnings", []),
    )
    return manifest


def write_post_check_artifact(run_id: str, artifact_key: str, payload: dict[str, Any], status: str) -> dict[str, Any]:
    """Write a validation or analysis artifact and advance the run state."""

    manifest = _load_manifest_by_id(run_id)
    _write_versioned_artifact(manifest, artifact_key, payload)
    _update_manifest(
        manifest,
        status=status,
        last_error={},
        resolution_warnings=manifest.get("resolution_warnings", []),
    )
    return manifest
