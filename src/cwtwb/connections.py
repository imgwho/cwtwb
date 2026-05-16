"""Connection configuration mixin for TWBEditor.

Handles local tabular files plus MySQL and Tableau Server connection setup.
"""

from __future__ import annotations

import csv
import os
import shutil
import tempfile
from copy import deepcopy
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, List, Optional

from lxml import etree
import xlrd

from .config import TMP_DIR, _generate_uuid

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - optional dependency
    load_workbook = None


_VALID_EXTERNAL_DATATYPES = {"string", "integer", "real", "boolean", "date", "datetime"}
_SEMANTIC_ROLE_BY_FIELD_NAME = {
    "city": "[City].[Name]",
    "country": "[Country].[ISO3166_2]",
    "country region": "[Country].[ISO3166_2]",
    "country/region": "[Country].[ISO3166_2]",
    "postal code": "[ZipCode].[Name]",
    "province": "[State].[Name]",
    "state": "[State].[Name]",
    "state province": "[State].[Name]",
    "state/province": "[State].[Name]",
    "zip": "[ZipCode].[Name]",
    "zip code": "[ZipCode].[Name]",
    "zipcode": "[ZipCode].[Name]",
}
_EXCEL_SAMPLE_LIMIT = 10_000
_STRING_LIKE_FIELD_HINTS = (
    "postal code",
    "zip code",
    "zipcode",
    "postcode",
    "postal",
    "zip",
    "phone number",
    "phone",
    "fax",
    "sku",
    "serial number",
    "tracking number",
    "account number",
    "part number",
)
_REMOTE_TYPE_BY_DATATYPE = {
    "boolean": "11",
    "date": "7",
    "datetime": "7",
    "integer": "20",
    "real": "5",
    "string": "130",
}
_DEBUG_REMOTE_TYPE_BY_DATATYPE = {
    "boolean": "BOOL",
    "date": "DATE",
    "datetime": "DATETIME",
    "integer": "I8",
    "real": "R8",
    "string": "WSTR",
}
_STRING_DATE_FORMATS = (
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M:%S",
    "%m/%d/%Y",
    "%m-%d-%Y",
    "%d/%m/%Y",
    "%d-%m-%Y",
)
_STRING_DATETIME_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y/%m/%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y/%m/%d %H:%M",
    "%m/%d/%Y %H:%M:%S",
    "%m-%d-%Y %H:%M:%S",
    "%m/%d/%Y %H:%M",
    "%m-%d-%Y %H:%M",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%dT%H:%M:%S.%f",
    "%Y-%m-%d %I:%M:%S %p",
    "%m/%d/%Y %I:%M:%S %p",
)


def infer_tableau_semantic_role(field_name: str) -> str:
    """Infer a Tableau semantic-role qualified name from a user-visible field name."""

    normalized = " ".join(
        field_name.casefold().replace("_", " ").replace("-", " ").split()
    )
    return _SEMANTIC_ROLE_BY_FIELD_NAME.get(normalized, "")


def _normalize_external_datatype(datatype: str) -> str:
    normalized = datatype.strip().lower()
    if normalized not in _VALID_EXTERNAL_DATATYPES:
        return "string"
    return normalized


def _infer_external_role(field_name: str, datatype: str) -> str:
    lower = field_name.casefold()
    if datatype in {"integer", "real"} and not any(
        token in lower for token in ("id", "code", "zip", "postal")
    ):
        return "measure"
    return "dimension"


def _infer_external_field_type(role: str, datatype: str) -> str:
    if datatype in {"date", "datetime"}:
        return "ordinal"
    if datatype == "integer" and role != "measure":
        return "ordinal"
    if role == "measure":
        return "quantitative"
    return "nominal"


def _default_local_name(field_name: str, source_object: str, semantic_role: str) -> str:
    if semantic_role in {"[Country].[ISO3166_2]", "[State].[Name]", "[ZipCode].[Name]"}:
        return f"[{field_name}]"
    if source_object:
        return f"[{field_name} ({source_object})]"
    return f"[{field_name}]"


def _excel_local_name(field_name: str, source_object: str, *, is_primary: bool, is_shared: bool) -> str:
    if is_primary or not is_shared:
        return f"[{field_name}]"
    return f"[{field_name} ({source_object})]"


def _extract_remote_name_from_map_value(value: str) -> str:
    """Return the right-most identifier from a Tableau map value like [Orders].[Sales]."""

    if not value:
        return ""
    stripped = value.strip()
    if stripped.startswith("[") and stripped.endswith("]"):
        stripped = stripped[1:-1]
    parts = stripped.split("].[")
    return parts[-1].replace("]]", "]").strip()


def _read_csv_rows(
    filepath: str,
    *,
    delimiter: str | None = None,
    encoding: str = "utf-8-sig",
    max_rows: int = 151,
) -> tuple[str, list[list[Any]]]:
    """Read a small sample of rows from a CSV file for schema inference."""

    path = Path(filepath)
    with path.open("r", encoding=encoding, newline="") as handle:
        sample = handle.read(8192)
        handle.seek(0)

        actual_delimiter = delimiter
        if not actual_delimiter:
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=[",", ";", "\t", "|"])
                actual_delimiter = dialect.delimiter
            except csv.Error:
                actual_delimiter = ","

        reader = csv.reader(handle, delimiter=actual_delimiter)
        rows: list[list[Any]] = []
        for row_index, row in enumerate(reader):
            rows.append(list(row))
            if row_index + 1 >= max_rows:
                break

    return actual_delimiter, rows


def _default_aggregation(datatype: str) -> str:
    if datatype in {"date", "datetime"}:
        return "Year"
    if datatype in {"integer", "real"}:
        return "Sum"
    return "Count"


def _column_values_from_rows(rows: list[list[Any]], index: int) -> list[Any]:
    values: list[Any] = []
    for row in rows:
        values.append(row[index] if index < len(row) else None)
    return values


def _looks_like_string_date(value: Any) -> bool:
    text = str(value).strip()
    if not text or len(text) < 6:
        return False
    for fmt in _STRING_DATE_FORMATS:
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue
    return False


def _looks_like_string_datetime(value: Any) -> bool:
    text = str(value).strip()
    if not text or len(text) < 11:
        return False
    for fmt in _STRING_DATETIME_FORMATS:
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            continue
    return False


def _is_string_like_field_name(header: str) -> bool:
    normalized = " ".join(header.casefold().replace("_", " ").replace("-", " ").split())
    return any(hint in normalized for hint in _STRING_LIKE_FIELD_HINTS)


def _is_probable_string_date_column(values: list[Any], *, threshold: float = 0.6, sample_limit: int = 24) -> bool:
    samples = [str(value).strip() for value in values if isinstance(value, str) and str(value).strip()]
    if len(samples) < 3:
        return False
    sample = samples[:sample_limit]
    matches = sum(1 for value in sample if _looks_like_string_date(value))
    return matches / len(sample) >= threshold


def _is_probable_string_datetime_column(values: list[Any], *, threshold: float = 0.6, sample_limit: int = 24) -> bool:
    samples = [str(value).strip() for value in values if isinstance(value, str) and str(value).strip()]
    if len(samples) < 3:
        return False
    sample = samples[:sample_limit]
    matches = sum(1 for value in sample if _looks_like_string_datetime(value))
    return matches / len(sample) >= threshold


def _datetime_has_time_component(value: Any) -> bool:
    if not isinstance(value, datetime):
        return False
    return value.time() != datetime.min.time()


def _infer_excel_datatype(header: str, values: list[Any]) -> str:
    lower = header.casefold()
    non_blank = [value for value in values if value not in ("", None)]
    if _is_string_like_field_name(header):
        return "string"
    if any(token in lower for token in ("datetime", "date & time", "date and time")):
        return "datetime"
    if any(token in lower for token in ("date", "month", "year", "day")):
        return "date"
    if any(token in lower for token in ("latitude", "longitude", "lat", "lon")):
        return "real"
    if not non_blank:
        return "string"
    if any(_datetime_has_time_component(value) for value in non_blank):
        return "datetime"
    if any(isinstance(value, (datetime, date)) for value in non_blank):
        return "date"
    if all(isinstance(value, bool) for value in non_blank):
        return "boolean"

    numeric_values: list[float] = []
    all_numeric = True
    for value in non_blank:
        if isinstance(value, bool):
            all_numeric = False
            break
        if isinstance(value, (int, float)):
            numeric_values.append(float(value))
            continue
        try:
            numeric_values.append(float(str(value).strip()))
        except Exception:
            all_numeric = False
            break

    if all_numeric and numeric_values:
        if all(float(item).is_integer() for item in numeric_values):
            return "integer"
        return "real"
    if _is_probable_string_datetime_column(non_blank):
        return "datetime"
    if _is_probable_string_date_column(non_blank):
        return "date"
    return "string"


def _read_xls_rows(filepath: str, *, sheet_name: str = "", max_rows: int = _EXCEL_SAMPLE_LIMIT) -> tuple[str, list[list[Any]]]:
    """Read a small sample from an .xls file while preserving cell types."""

    path = Path(filepath)
    workbook = xlrd.open_workbook(str(path))
    if sheet_name:
        try:
            worksheet = workbook.sheet_by_name(sheet_name)
        except xlrd.biffh.XLRDError:
            worksheet = workbook.sheet_by_index(0)
    else:
        worksheet = workbook.sheet_by_index(0)

    actual_sheet_name = worksheet.name
    rows: list[list[Any]] = []
    for row_index in range(min(worksheet.nrows, max_rows)):
        row: list[Any] = []
        for col_index in range(worksheet.ncols):
            cell = worksheet.cell(row_index, col_index)
            value: Any = cell.value
            if cell.ctype == getattr(xlrd, "XL_CELL_DATE", 3):
                value = xlrd.xldate_as_datetime(value, workbook.datemode)
            elif cell.ctype == getattr(xlrd, "XL_CELL_BOOLEAN", 4):
                value = bool(value)
            elif cell.ctype == getattr(xlrd, "XL_CELL_ERROR", 5):
                value = None
            elif cell.ctype in {getattr(xlrd, "XL_CELL_EMPTY", 0), getattr(xlrd, "XL_CELL_BLANK", 6)}:
                value = None
            row.append(value)
        rows.append(row)

    return actual_sheet_name, rows


def _excel_column_letter(index: int) -> str:
    """Return a 1-based Excel column letter."""

    if index < 1:
        return "A"
    letters: list[str] = []
    current = index
    while current > 0:
        current, remainder = divmod(current - 1, 26)
        letters.append(chr(65 + remainder))
    return "".join(reversed(letters))


def _excel_grid_origin(rows: list[list[Any]]) -> str:
    """Build a Tableau-style gridOrigin string for an Excel sheet sample."""

    if not rows:
        return "A1:A1:no:A1:A1:0"
    col_count = max((len(row) for row in rows), default=0)
    if col_count <= 0:
        col_count = 1
    row_count = max(len(rows), 1)
    last_col = _excel_column_letter(col_count)
    last_row = str(row_count)
    return f"A1:{last_col}{last_row}:no:A1:{last_col}{last_row}:0"


def _read_excel_sheet_rows(filepath: str, sheet_name: str = "") -> tuple[str, list[list[Any]]]:
    """Read rows from one Excel sheet, preserving cell types when possible."""

    path = Path(filepath)
    suffix = path.suffix.lower()
    if suffix == ".xls":
        return _read_xls_rows(filepath, sheet_name=sheet_name)
    if suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:  # pragma: no cover - depends on optional dependency
            raise RuntimeError(
                "Reading .xlsx/.xlsm files requires openpyxl to be installed."
            )
        workbook = load_workbook(str(path), read_only=True, data_only=True)
        try:
            if sheet_name and sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
            else:
                worksheet = workbook[workbook.sheetnames[0]]
            rows: list[list[Any]] = []
            for row in worksheet.iter_rows(min_row=1, max_row=_EXCEL_SAMPLE_LIMIT, values_only=True):
                rows.append(list(row))
            return worksheet.title, rows
        finally:
            workbook.close()
    raise ValueError(f"Unsupported Excel file type: {suffix}")


def _list_excel_sheet_names(filepath: str) -> list[str]:
    """Return Excel sheet names in workbook order."""

    path = Path(filepath)
    suffix = path.suffix.lower()
    if suffix == ".xls":
        workbook = xlrd.open_workbook(str(path))
        try:
            return list(workbook.sheet_names())
        finally:
            pass
    if suffix in {".xlsx", ".xlsm"}:
        if load_workbook is None:  # pragma: no cover - depends on optional dependency
            raise RuntimeError(
                "Reading .xlsx/.xlsm files requires openpyxl to be installed."
            )
        workbook = load_workbook(str(path), read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()
    raise ValueError(f"Unsupported Excel file type: {suffix}")


def _sanitize_headers(header_row: list[Any]) -> list[str]:
    seen: dict[str, int] = {}
    headers: list[str] = []
    for index, value in enumerate(header_row):
        header = str(value).strip() if value is not None else ""
        if not header:
            header = f"Field {index + 1}"
        if header in seen:
            seen[header] += 1
            header = f"{header} {seen[header]}"
        else:
            seen[header] = 1
        headers.append(header)
    return headers


def inspect_hyper_schema(filepath: str) -> dict:
    """Read a Hyper file and return its schema.

    Returns a dict of the form::

        {"tables": [
            {"schema": "Extract", "name": "Orders",
             "columns": [{"name": "Sales", "type": "double"}, ...]},
            ...
        ]}

    If the file is locked (e.g. open in Tableau), it is copied to a
    temporary location first.
    """
    from tableauhyperapi import HyperProcess, Connection, Telemetry, SchemaName

    # Copy to a temp file so we don't fail on locked hyper files
    TMP_DIR.mkdir(parents=True, exist_ok=True)
    tmp_dir = tempfile.mkdtemp(prefix="cwtwb_hyper_", dir=str(TMP_DIR))
    tmp_path = os.path.join(tmp_dir, os.path.basename(filepath))
    try:
        database_path = filepath
        try:
            shutil.copy2(filepath, tmp_path)
            database_path = tmp_path
        except PermissionError:
            database_path = filepath

        tables_out: list[dict] = []
        try:
            with HyperProcess(telemetry=Telemetry.DO_NOT_SEND_USAGE_DATA_TO_TABLEAU) as hyper:
                with Connection(
                    endpoint=hyper.endpoint,
                    database=database_path,
                ) as conn:
                    for schema_name in conn.catalog.get_schema_names():
                        for table_name in conn.catalog.get_table_names(schema_name):
                            table_def = conn.catalog.get_table_definition(table_name)
                            columns = []
                            for col in table_def.columns:
                                columns.append({
                                    "name": col.name.unescaped,
                                    "type": str(col.type),
                                })
                            tables_out.append({
                                "schema": str(schema_name),
                                "name": table_name.name.unescaped,
                                "columns": columns,
                            })
        except Exception as exc:
            raise RuntimeError(f"Unable to inspect Hyper schema: {exc}") from exc
        return {"tables": tables_out}
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


class ConnectionsMixin:
    """Mixin providing datasource connection methods for TWBEditor."""

    def _normalize_external_fields(
        self,
        fields: list[dict[str, Any]] | None,
        *,
        source_object: str,
    ) -> list[dict[str, Any]]:
        """Normalize field payloads to the shape required by Tableau metadata builders."""

        normalized_fields: list[dict[str, Any]] = []
        if not fields:
            return normalized_fields

        for ordinal, field in enumerate(fields):
            name = str(field.get("name", "")).strip()
            if not name:
                continue
            field_source_object = str(field.get("source_object", "")).strip() or source_object
            datatype = _normalize_external_datatype(
                str(field.get("datatype") or field.get("inferred_type") or "string")
            )
            role = str(field.get("role", "")).strip().lower() or _infer_external_role(name, datatype)
            field_type = (
                str(field.get("field_type") or field.get("type") or "").strip().lower()
                or _infer_external_field_type(role, datatype)
            )
            raw_semantic_role = str(field.get("semantic_role", "")).strip()
            if raw_semantic_role.casefold() == "geographic":
                semantic_role = infer_tableau_semantic_role(name)
            elif raw_semantic_role.startswith("["):
                semantic_role = raw_semantic_role
            elif raw_semantic_role:
                semantic_role = infer_tableau_semantic_role(raw_semantic_role) or infer_tableau_semantic_role(name)
            else:
                semantic_role = infer_tableau_semantic_role(name)

            normalized_fields.append(
                {
                    "name": name,
                    "ordinal": ordinal,
                    "datatype": datatype,
                    "role": role,
                    "field_type": field_type,
                    "semantic_role": semantic_role,
                    "source_object": field_source_object,
                }
            )

        return normalized_fields

    def _introspect_excel_fields(self, filepath: str, sheet_name: str) -> tuple[str, list[dict[str, Any]]]:
        """Inspect an Excel sheet locally when MCP callers do not provide schema fields."""

        actual_sheet_name, rows = _read_excel_sheet_rows(filepath, sheet_name=sheet_name)

        if not rows:
            return actual_sheet_name, []

        headers = _sanitize_headers(rows[0])
        value_rows = rows[1:]
        fields: list[dict[str, Any]] = []
        for index, header in enumerate(headers):
            values = _column_values_from_rows(value_rows, index)
            datatype = _infer_excel_datatype(header, values)
            role = _infer_external_role(header, datatype)
            fields.append(
                {
                    "name": header,
                    "datatype": datatype,
                    "role": role,
                    "field_type": _infer_external_field_type(role, datatype),
                    "semantic_role": infer_tableau_semantic_role(header),
                }
            )

        return actual_sheet_name, fields

    def _introspect_excel_tables(self, filepath: str, sheet_name: str = "") -> list[dict[str, Any]]:
        """Inspect every Excel sheet locally and return table-oriented schema specs."""

        sheet_names = _list_excel_sheet_names(filepath)
        if not sheet_names:
            return []
        if sheet_name and sheet_name in sheet_names:
            sheet_names = [sheet_name] + [name for name in sheet_names if name != sheet_name]

        tables: list[dict[str, Any]] = []
        for index, candidate_sheet in enumerate(sheet_names):
            actual_sheet_name, rows = _read_excel_sheet_rows(filepath, sheet_name=candidate_sheet)
            if not rows:
                continue

            headers = _sanitize_headers(rows[0])
            value_rows = rows[1:]
            fields: list[dict[str, Any]] = []
            for ordinal, header in enumerate(headers):
                values = _column_values_from_rows(value_rows, ordinal)
                datatype = _infer_excel_datatype(header, values)
                role = _infer_external_role(header, datatype)
                fields.append(
                    {
                        "name": header,
                        "ordinal": ordinal,
                        "datatype": datatype,
                        "role": role,
                        "field_type": _infer_external_field_type(role, datatype),
                        "semantic_role": infer_tableau_semantic_role(header),
                    }
                )

            tables.append(
                {
                    "name": actual_sheet_name,
                    "fields": fields,
                    "grid_origin": _excel_grid_origin(rows),
                    "outcome": "6" if len(rows) > 1 else "2",
                }
            )

        return tables

    def _introspect_csv_fields(
        self,
        filepath: str,
        delimiter: str = "",
        charset: str = "utf-8-sig",
    ) -> tuple[str, list[dict[str, Any]], str]:
        """Inspect a CSV file locally when MCP callers do not provide schema fields."""

        path = Path(filepath)
        actual_source_object = path.name
        actual_delimiter, rows = _read_csv_rows(
            filepath,
            delimiter=delimiter or None,
            encoding=charset,
        )

        if not rows:
            return actual_source_object, [], actual_delimiter

        headers = _sanitize_headers(rows[0])
        value_rows = rows[1:]
        fields: list[dict[str, Any]] = []
        for index, header in enumerate(headers):
            values = _column_values_from_rows(value_rows, index)
            datatype = _infer_excel_datatype(header, values)
            role = _infer_external_role(header, datatype)
            fields.append(
                {
                    "name": header,
                    "datatype": datatype,
                    "role": role,
                    "field_type": _infer_external_field_type(role, datatype),
                    "semantic_role": infer_tableau_semantic_role(header),
                }
            )

        return actual_source_object, fields, actual_delimiter

    def _rebuild_external_datasource_metadata(
        self,
        *,
        source_object: str,
        fields: list[dict[str, Any]],
        relation: etree._Element,
        prefer_existing_metadata: bool = True,
        local_name_source_object: str | None = None,
        relation_column_attrs_override: dict[str, str] | None = None,
        capability_attrs_override: list[tuple[str, str, str]] | None = None,
    ) -> None:
        """Rebuild a tabular datasource structure so Tableau sees a coherent schema."""

        relation_columns = relation.find("columns")
        relation_column_attrs = dict(relation_columns.attrib) if relation_columns is not None else {}
        has_override = relation_column_attrs_override is not None
        existing_cols = self._datasource.find(".//connection[@class='federated']/cols")
        existing_maps: dict[str, str] = {}
        if prefer_existing_metadata and existing_cols is not None:
            for map_el in existing_cols.findall("map"):
                key = map_el.get("key", "")
                remote_name = _extract_remote_name_from_map_value(map_el.get("value", ""))
                if key and remote_name:
                    existing_maps[remote_name] = key

        existing_metadata: dict[str, str] = {}
        if prefer_existing_metadata:
            for mr in self._datasource.findall(".//metadata-records/metadata-record[@class='column']"):
                remote_name = (mr.findtext("remote-name") or "").strip()
                local_name = (mr.findtext("local-name") or "").strip()
                if remote_name and local_name:
                    existing_metadata[remote_name] = local_name

        existing_top_level_columns = [
            deepcopy(col)
            for col in self._datasource.findall("column")
            if col.find("calculation") is None
        ]
        top_level_templates = {
            col.get("name", ""): deepcopy(col)
            for col in existing_top_level_columns
            if col.get("name")
        }

        object_el = self._datasource.find(".//object-graph//object")
        object_id = object_el.get("id", "") if object_el is not None else ""
        if object_el is not None:
            object_el.set("caption", source_object)

        normalized_fields: list[dict[str, Any]] = []
        local_name_source_object = source_object if local_name_source_object is None else local_name_source_object
        for field in fields:
            local_name = (
                (existing_metadata.get(field["name"]) if prefer_existing_metadata else "")
                or (existing_maps.get(field["name"]) if prefer_existing_metadata else "")
                or _default_local_name(field["name"], local_name_source_object, field["semantic_role"])
            )
            normalized_fields.append({**field, "local_name": local_name})

        # Keep calculated fields in place; rebuild external field scaffolding around them.
        calculated_columns = [
            deepcopy(col)
            for col in self._datasource.findall("column")
            if col.find("calculation") is not None
        ]
        for col in list(self._datasource.findall("column")):
            self._datasource.remove(col)

        alias_el = self._datasource.find("aliases")
        if alias_el is None:
            alias_el = etree.Element("aliases")
            alias_el.set("enabled", "yes")
            layout = self._datasource.find("layout")
            if layout is not None:
                layout.addprevious(alias_el)
            else:
                self._datasource.append(alias_el)

        top_level_columns: list[etree._Element] = []
        for field in normalized_fields:
            if field["role"] == "measure" or field["datatype"] == "date":
                continue
            template = top_level_templates.get(field["local_name"])
            if template is None:
                col = etree.Element("column")
                col.set("name", field["local_name"])
            else:
                col = template
                col.set("name", field["local_name"])
            col.set("datatype", field["datatype"])
            col.set("role", field["role"])
            col.set("type", field["field_type"])
            if field["semantic_role"]:
                col.set("semantic-role", field["semantic_role"])
            elif "semantic-role" in col.attrib:
                del col.attrib["semantic-role"]
            col.attrib.pop("caption", None)
            top_level_columns.append(col)

        internal_table_columns = [
            deepcopy(col)
            for col in existing_top_level_columns
            if (col.get("datatype") or "").strip().lower() == "table"
            or col.get("name", "").startswith("[__tableau_internal_object_id__].")
        ]
        for col in internal_table_columns:
            if source_object:
                col.set("caption", source_object)
            top_level_columns.append(col)

        insertion_anchor = None
        for tag in (
            "column-instance",
            "group",
            "layout",
            "semantic-values",
            "date-options",
            "object-graph",
        ):
            candidate = self._datasource.find(tag)
            if candidate is not None:
                insertion_anchor = candidate
                break
        for col in calculated_columns + top_level_columns:
            if insertion_anchor is not None:
                insertion_anchor.addprevious(col)
            else:
                self._datasource.append(col)

        # Rebuild relation columns for both the connection relation and object-graph relation.
        if relation_columns is None:
            relation_columns = etree.SubElement(relation, "columns")
        else:
            for child in list(relation_columns):
                relation_columns.remove(child)
        if relation_column_attrs:
            relation_columns.attrib.clear()
            relation_columns.attrib.update(relation_column_attrs)
        if relation_column_attrs_override:
            relation_columns.attrib.clear()
            relation_columns.attrib.update(relation_column_attrs_override)
        if "header" not in relation_columns.attrib:
            relation_columns.set("header", "yes")
        if not has_override and "outcome" not in relation_columns.attrib:
            relation_columns.set("outcome", "2")

        for field in normalized_fields:
            column_el = etree.SubElement(relation_columns, "column")
            column_el.set("datatype", field["datatype"])
            column_el.set("name", field["name"])
            column_el.set("ordinal", str(field["ordinal"]))

        for og_rel in self._datasource.findall(".//object-graph//relation"):
            og_columns = og_rel.find("columns")
            og_column_attrs = (
                dict(og_columns.attrib) if og_columns is not None and not has_override else relation_column_attrs_override or relation_columns.attrib
            )
            if og_columns is None:
                og_columns = etree.SubElement(og_rel, "columns")
            else:
                for child in list(og_columns):
                    og_columns.remove(child)
            og_columns.attrib.clear()
            og_columns.attrib.update(og_column_attrs)
            if "header" not in og_columns.attrib:
                og_columns.set("header", "yes")
            if not has_override and "outcome" not in og_columns.attrib:
                og_columns.set("outcome", "2")
            for field in normalized_fields:
                column_el = etree.SubElement(og_columns, "column")
                column_el.set("datatype", field["datatype"])
                column_el.set("name", field["name"])
                column_el.set("ordinal", str(field["ordinal"]))

        cols_el = self._datasource.find(".//connection[@class='federated']/cols")
        if cols_el is None:
            cols_el = etree.Element("cols")
            metadata_records = self._datasource.find(".//connection[@class='federated']/metadata-records")
            if metadata_records is not None:
                metadata_records.addprevious(cols_el)
            else:
                relation.addnext(cols_el)
        else:
            for child in list(cols_el):
                cols_el.remove(child)
        for field in normalized_fields:
            map_el = etree.SubElement(cols_el, "map")
            map_el.set("key", field["local_name"])
            map_el.set("value", f"[{source_object}].[{field['name']}]")

        metadata_records = self._datasource.find(".//connection[@class='federated']/metadata-records")
        if metadata_records is None:
            metadata_records = etree.Element("metadata-records")
            cols_el.addnext(metadata_records)
        else:
            for child in list(metadata_records):
                metadata_records.remove(child)

        capability_record = etree.SubElement(metadata_records, "metadata-record")
        capability_record.set("class", "capability")
        etree.SubElement(capability_record, "remote-name")
        etree.SubElement(capability_record, "remote-type").text = "0"
        etree.SubElement(capability_record, "parent-name").text = f"[{source_object}]"
        etree.SubElement(capability_record, "remote-alias")
        etree.SubElement(capability_record, "aggregation").text = "Count"
        etree.SubElement(capability_record, "contains-null").text = "true"
        capability_attrs = etree.SubElement(capability_record, "attributes")
        if capability_attrs_override:
            for datatype, name, value in capability_attrs_override:
                attr = etree.SubElement(capability_attrs, "attribute")
                attr.set("datatype", datatype)
                attr.set("name", name)
                attr.text = value
        else:
            attr_context = etree.SubElement(capability_attrs, "attribute")
            attr_context.set("datatype", "integer")
            attr_context.set("name", "context")
            attr_context.text = "0"
            if relation_columns.get("gridOrigin"):
                attr_grid = etree.SubElement(capability_attrs, "attribute")
                attr_grid.set("datatype", "string")
                attr_grid.set("name", "gridOrigin")
                attr_grid.text = f"\"{relation_columns.get('gridOrigin')}\""
            attr_header = etree.SubElement(capability_attrs, "attribute")
            attr_header.set("datatype", "boolean")
            attr_header.set("name", "header")
            attr_header.text = "true"
            attr_outcome = etree.SubElement(capability_attrs, "attribute")
            attr_outcome.set("datatype", "integer")
            attr_outcome.set("name", "outcome")
            attr_outcome.text = relation_columns.get("outcome", "2")

        for field in normalized_fields:
            metadata_record = etree.SubElement(metadata_records, "metadata-record")
            metadata_record.set("class", "column")
            etree.SubElement(metadata_record, "remote-name").text = field["name"]
            etree.SubElement(metadata_record, "remote-type").text = _REMOTE_TYPE_BY_DATATYPE[field["datatype"]]
            etree.SubElement(metadata_record, "local-name").text = field["local_name"]
            etree.SubElement(metadata_record, "parent-name").text = f"[{source_object}]"
            etree.SubElement(metadata_record, "remote-alias").text = field["name"]
            etree.SubElement(metadata_record, "ordinal").text = str(field["ordinal"])
            etree.SubElement(metadata_record, "local-type").text = field["datatype"]
            etree.SubElement(metadata_record, "aggregation").text = _default_aggregation(field["datatype"])
            if field["datatype"] == "real":
                etree.SubElement(metadata_record, "precision").text = "15"
            etree.SubElement(metadata_record, "contains-null").text = "true"
            if field["datatype"] == "string":
                collation = etree.SubElement(metadata_record, "collation")
                collation.set("flag", "1")
                collation.set("name", "LZH_RCN_S2")
            attributes = etree.SubElement(metadata_record, "attributes")
            debug_attr = etree.SubElement(attributes, "attribute")
            debug_attr.set("datatype", "string")
            debug_attr.set("name", "DebugRemoteType")
            debug_attr.text = f"\"{_DEBUG_REMOTE_TYPE_BY_DATATYPE[field['datatype']]}\""
            if object_id:
                etree.SubElement(metadata_record, "object-id").text = f"[{object_id}]"

        self._reinit_fields()
        self.field_registry.set_unknown_field_policy(allow_unknown_fields=True)

    def _rebuild_excel_multi_table_metadata(
        self,
        *,
        conn_name: str,
        tables: list[dict[str, Any]],
    ) -> None:
        """Rebuild datasource metadata for an Excel workbook with multiple sheets."""

        if not tables:
            self._reinit_fields()
            self.field_registry.set_unknown_field_policy(allow_unknown_fields=True)
            return

        name_counts = Counter(
            field["name"]
            for table in tables
            for field in table.get("fields", [])
        )

        # Capture calculated columns so we can restore them after rebuilding
        calculated_columns = [
            deepcopy(col)
            for col in self._datasource.findall("column")
            if col.find("calculation") is not None
        ]
        for col in list(self._datasource.findall("column")):
            self._datasource.remove(col)

        alias_el = self._datasource.find("aliases")
        if alias_el is None:
            alias_el = etree.Element("aliases")
            alias_el.set("enabled", "yes")
            layout = self._datasource.find("layout")
            if layout is not None:
                layout.addprevious(alias_el)
            else:
                self._datasource.append(alias_el)

        # Remove old federated payload we are about to regenerate.
        fed_conn = self._datasource.find("connection[@class='federated']")
        if fed_conn is None:
            raise RuntimeError("Expected federated connection to exist before rebuilding Excel metadata.")

        old_cols = fed_conn.find("cols")
        if old_cols is not None:
            fed_conn.remove(old_cols)
        old_metadata_records = fed_conn.find("metadata-records")
        if old_metadata_records is not None:
            fed_conn.remove(old_metadata_records)

        for old_rel in fed_conn.findall("relation"):
            fed_conn.remove(old_rel)

        collection = etree.SubElement(fed_conn, "relation")
        collection.set("type", "collection")

        normalized_tables: list[dict[str, Any]] = []
        for index, table in enumerate(tables):
            table_name = str(table.get("name", "")).strip()
            fields = list(table.get("fields", []))
            is_primary = index == 0
            object_id = f"{table_name}_{_generate_uuid().strip('{}').replace('-', '')}"
            table_fields: list[dict[str, Any]] = []
            for field in fields:
                field_name = str(field.get("name", "")).strip()
                if not field_name:
                    continue
                is_shared = name_counts[field_name] > 1
                local_name = _excel_local_name(
                    field_name,
                    table_name,
                    is_primary=is_primary,
                    is_shared=is_shared,
                )
                table_fields.append(
                    {
                        **field,
                        "local_name": local_name,
                        "is_primary": is_primary,
                        "is_shared": is_shared,
                    }
                )

            relation = etree.SubElement(collection, "relation")
            relation.set("connection", conn_name)
            relation.set("name", table_name)
            relation.set("table", f"[{table_name}$]")
            relation.set("type", "table")
            columns = etree.SubElement(relation, "columns")
            columns.attrib.clear()
            columns.set("gridOrigin", str(table.get("grid_origin", "A1:A1:no:A1:A1:0")))
            columns.set("header", "yes")
            columns.set("outcome", str(table.get("outcome", "2")))
            for ordinal, field in enumerate(table_fields):
                column_el = etree.SubElement(columns, "column")
                column_el.set("datatype", field["datatype"])
                column_el.set("name", field["name"])
                column_el.set("ordinal", str(ordinal))

            normalized_tables.append(
                {
                    "name": table_name,
                    "id": object_id,
                    "fields": table_fields,
                    "is_primary": is_primary,
                    "grid_origin": table.get("grid_origin", "A1:A1:no:A1:A1:0"),
                    "outcome": table.get("outcome", "6"),
                    "relation": relation,
                }
            )

        # Update object-graph to mirror the workbook's table graph.
        object_graph = self._datasource.find("object-graph")
        if object_graph is None:
            object_graph = etree.SubElement(self._datasource, "object-graph")
        objects_el = object_graph.find("objects")
        if objects_el is None:
            objects_el = etree.SubElement(object_graph, "objects")
        else:
            for child in list(objects_el):
                objects_el.remove(child)

        relationships_el = object_graph.find("relationships")
        if relationships_el is None:
            relationships_el = etree.SubElement(object_graph, "relationships")
        else:
            for child in list(relationships_el):
                relationships_el.remove(child)

        table_objects: list[dict[str, Any]] = []
        for table in normalized_tables:
            table_name = table["name"]
            table_object_id = table["id"]
            relation = table["relation"]
            obj = etree.SubElement(objects_el, "object")
            obj.set("caption", table_name)
            obj.set("id", table_object_id)
            props = etree.SubElement(obj, "properties")
            props.set("context", "")
            props.append(deepcopy(relation))

            table_objects.append(
                {
                    "name": table_name,
                    "id": table_object_id,
                    "relation": relation,
                    "fields": table["fields"],
                    "is_primary": table["is_primary"],
                }
            )

        # Infer relationships from shared field names, preferring the primary table.
        if table_objects:
            primary = table_objects[0]
            primary_field_map = {field["name"]: field["local_name"] for field in primary["fields"]}
            for secondary in table_objects[1:]:
                secondary_field_map = {field["name"]: field["local_name"] for field in secondary["fields"]}
                shared_field_name = None
                for field in primary["fields"]:
                    if field["name"] in secondary_field_map:
                        shared_field_name = field["name"]
                        break
                if not shared_field_name:
                    continue
                rel_el = etree.SubElement(relationships_el, "relationship")
                expr = etree.SubElement(rel_el, "expression")
                expr.set("op", "=")
                left_expr = etree.SubElement(expr, "expression")
                left_expr.set("op", primary_field_map[shared_field_name])
                right_expr = etree.SubElement(expr, "expression")
                right_expr.set("op", secondary_field_map[shared_field_name])
                first_end = etree.SubElement(rel_el, "first-end-point")
                first_end.set("object-id", primary["id"])
                second_end = etree.SubElement(rel_el, "second-end-point")
                second_end.set("object-id", secondary["id"])

        # Rebuild shared <cols> mappings and metadata records.
        cols_el = etree.SubElement(fed_conn, "cols")
        metadata_records = etree.SubElement(fed_conn, "metadata-records")
        top_level_columns: list[etree._Element] = []

        existing_top_level_columns = [
            deepcopy(col)
            for col in self._datasource.findall("column")
            if col.find("calculation") is None
        ]
        top_level_templates = {
            col.get("name", ""): deepcopy(col)
            for col in existing_top_level_columns
            if col.get("name")
        }

        insertion_anchor = None
        for tag in (
            "column-instance",
            "group",
            "layout",
            "semantic-values",
            "date-options",
            "object-graph",
        ):
            candidate = self._datasource.find(tag)
            if candidate is not None:
                insertion_anchor = candidate
                break

        for table in normalized_tables:
            table_name = table["name"]
            table_object_id = table["id"]
            capability_record = etree.SubElement(metadata_records, "metadata-record")
            capability_record.set("class", "capability")
            etree.SubElement(capability_record, "remote-name")
            etree.SubElement(capability_record, "remote-type").text = "0"
            etree.SubElement(capability_record, "parent-name").text = f"[{table_name}]"
            etree.SubElement(capability_record, "remote-alias")
            etree.SubElement(capability_record, "aggregation").text = "Count"
            etree.SubElement(capability_record, "contains-null").text = "true"
            capability_attributes = etree.SubElement(capability_record, "attributes")
            attr_context = etree.SubElement(capability_attributes, "attribute")
            attr_context.set("datatype", "integer")
            attr_context.set("name", "context")
            attr_context.text = "0"
            attr_grid = etree.SubElement(capability_attributes, "attribute")
            attr_grid.set("datatype", "string")
            attr_grid.set("name", "gridOrigin")
            attr_grid.text = f"\"{table.get('grid_origin', 'A1:A1:no:A1:A1:0')}\""
            attr_header = etree.SubElement(capability_attributes, "attribute")
            attr_header.set("datatype", "boolean")
            attr_header.set("name", "header")
            attr_header.text = "true"
            attr_outcome = etree.SubElement(capability_attributes, "attribute")
            attr_outcome.set("datatype", "integer")
            attr_outcome.set("name", "outcome")
            attr_outcome.text = str(table.get("outcome", "6"))

            for field in table["fields"]:
                map_el = etree.SubElement(cols_el, "map")
                map_el.set("key", field["local_name"])
                map_el.set("value", f"[{table_name}].[{field['name']}]")

                metadata_record = etree.SubElement(metadata_records, "metadata-record")
                metadata_record.set("class", "column")
                etree.SubElement(metadata_record, "remote-name").text = field["name"]
                etree.SubElement(metadata_record, "remote-type").text = _REMOTE_TYPE_BY_DATATYPE[field["datatype"]]
                etree.SubElement(metadata_record, "local-name").text = field["local_name"]
                etree.SubElement(metadata_record, "parent-name").text = f"[{table_name}]"
                etree.SubElement(metadata_record, "remote-alias").text = field["name"]
                etree.SubElement(metadata_record, "ordinal").text = str(field["ordinal"])
                etree.SubElement(metadata_record, "local-type").text = field["datatype"]
                etree.SubElement(metadata_record, "aggregation").text = _default_aggregation(field["datatype"])
                if field["datatype"] == "real":
                    etree.SubElement(metadata_record, "precision").text = "15"
                etree.SubElement(metadata_record, "contains-null").text = "true"
                if field["datatype"] == "string":
                    collation = etree.SubElement(metadata_record, "collation")
                    collation.set("flag", "1")
                    collation.set("name", "LZH_RCN_S2")
                attributes = etree.SubElement(metadata_record, "attributes")
                debug_attr = etree.SubElement(attributes, "attribute")
                debug_attr.set("datatype", "string")
                debug_attr.set("name", "DebugRemoteType")
                debug_attr.text = f"\"{_DEBUG_REMOTE_TYPE_BY_DATATYPE[field['datatype']]}\""
                etree.SubElement(metadata_record, "object-id").text = f"[{table_object_id}]"

                if field["is_primary"] or field["is_shared"]:
                    template = top_level_templates.get(field["local_name"])
                    if template is None:
                        col = etree.Element("column")
                        col.set("name", field["local_name"])
                    else:
                        col = template
                        col.set("name", field["local_name"])
                    col.set("datatype", field["datatype"])
                    col.set("role", field["role"])
                    col.set("type", field["field_type"])
                    if field["semantic_role"]:
                        col.set("semantic-role", field["semantic_role"])
                        if field["semantic_role"] == "[ZipCode].[Name]":
                            col.set("default-format", "*00000")
                    elif "semantic-role" in col.attrib:
                        del col.attrib["semantic-role"]
                    if not field["is_primary"]:
                        col.set("hidden", "true")
                    elif "hidden" in col.attrib:
                        del col.attrib["hidden"]
                    col.attrib.pop("caption", None)
                    top_level_columns.append(col)

        for table in normalized_tables:
            internal_col = etree.Element("column")
            internal_col.set("caption", table["name"])
            internal_col.set(
                "name",
                f"[__tableau_internal_object_id__].[{table['id']}]",
            )
            internal_col.set("datatype", "table")
            internal_col.set("role", "measure")
            internal_col.set("type", "quantitative")
            top_level_columns.append(internal_col)

        # Restore calculated fields and top-level columns.
        for col in calculated_columns + top_level_columns:
            if insertion_anchor is not None:
                insertion_anchor.addprevious(col)
            else:
                self._datasource.append(col)

        # Re-register fields so calculation rewriting keeps working.
        self._reinit_fields()
        for table in normalized_tables:
            for field in table["fields"]:
                display_name = field["local_name"].strip("[]")
                self.field_registry.register(
                    display_name=display_name,
                    local_name=field["local_name"],
                    datatype=field["datatype"],
                    role=field["role"],
                    field_type=field["field_type"],
                    is_calculated=False,
                )
        self.field_registry.set_unknown_field_policy(allow_unknown_fields=True)

    def _register_external_fields(self, fields: list[dict[str, Any]] | None) -> None:
        """Replace datasource field stubs from an inspected external schema."""

        if not fields:
            self.field_registry.set_unknown_field_policy(allow_unknown_fields=True)
            return

        for col in list(self._datasource.findall("column")):
            calc = col.find("calculation")
            if calc is None:
                self._datasource.remove(col)

        insert_before = self._datasource.find("layout")
        normalized_fields = self._normalize_external_fields(fields, source_object="")
        for field in normalized_fields:
            name = field["name"]

            col = etree.Element("column")
            col.set("name", f"[{name}]")
            col.set("caption", name)
            col.set("datatype", field["datatype"])
            col.set("role", field["role"])
            col.set("type", field["field_type"])
            if field["semantic_role"]:
                col.set("semantic-role", field["semantic_role"])

            if insert_before is not None:
                insert_before.addprevious(col)
            else:
                self._datasource.append(col)

        self._reinit_fields()
        self.field_registry.set_unknown_field_policy(allow_unknown_fields=True)

    def set_mysql_connection(
        self,
        server: str,
        dbname: str,
        username: str,
        table_name: str,
        port: str = "3306",
    ) -> str:
        """Configure the datasource to use a Local MySQL connection."""
        # 1. Update <connection class='federated'>
        fed_conn = self._datasource.find("connection[@class='federated']")
        if fed_conn is None:
            for old_conn in self._datasource.findall("connection"):
                self._datasource.remove(old_conn)
            fed_conn = etree.Element("connection")
            fed_conn.set("class", "federated")
            self._datasource.insert(0, fed_conn)

        # Update <named-connections>
        named_conns = fed_conn.find("named-connections")
        if named_conns is None:
            named_conns = etree.SubElement(fed_conn, "named-connections")
        else:
            for child in list(named_conns):
                named_conns.remove(child)

        conn_name = f"mysql.{_generate_uuid().strip('{}').lower()}"

        nc = etree.SubElement(named_conns, "named-connection")
        nc.set("caption", server)
        nc.set("name", conn_name)

        mysql_conn = etree.SubElement(nc, "connection")
        mysql_conn.set("class", "mysql")
        mysql_conn.set("dbname", dbname)
        mysql_conn.set("odbc-native-protocol", "")
        mysql_conn.set("one-time-sql", "")
        mysql_conn.set("port", str(port))
        mysql_conn.set("server", server)
        mysql_conn.set("source-charset", "")
        mysql_conn.set("username", username)

        # 2. Update <relation>
        relation = fed_conn.find("relation")
        if relation is None:
            relation = etree.SubElement(fed_conn, "relation")

        relation.set("connection", conn_name)
        relation.set("name", table_name)
        relation.set("table", f"[{table_name}]")
        relation.set("type", "table")
        for cols in relation.findall("columns"):
            relation.remove(cols)

        # 3. Update <object-graph> relation
        for og_rel in self._datasource.findall(".//object-graph//relation"):
            og_rel.set("connection", conn_name)
            og_rel.set("name", table_name)
            og_rel.set("table", f"[{table_name}]")
            og_rel.set("type", "table")
            for cols in og_rel.findall("columns"):
                og_rel.remove(cols)

        # 4. Cleanup old generic/excel connections and leftover fields
        excel_conn = self._datasource.find("connection[@class='excel-direct']")
        if excel_conn is not None:
            self._datasource.remove(excel_conn)
            
        old_cols = fed_conn.find("cols")
        if old_cols is not None:
            fed_conn.remove(old_cols)
            
        for c in self._datasource.findall("column"):
            self._datasource.remove(c)
            
        aliases = self._datasource.find("aliases")
        if aliases is not None:
            self._datasource.remove(aliases)

        # 6. Clean metadata-records
        for mr in self._datasource.findall(".//metadata-record"):
            mr.getparent().remove(mr)

        self._reinit_fields()
        self.field_registry.set_unknown_field_policy(allow_unknown_fields=True)
        return f"Configured MySQL connection to {server}/{dbname} (table: {table_name})"

    def set_tableauserver_connection(
        self,
        server: str,
        dbname: str,
        username: str,
        table_name: str,
        directory: str = "/dataserver",
        port: str = "82",
    ) -> str:
        """Configure the datasource to use a Tableau Server connection."""
        # 1. Remove all old connections
        for conn in self._datasource.findall("connection"):
            self._datasource.remove(conn)

        # 2. Add <repository-location>
        repo = self._datasource.find("repository-location")
        if repo is None:
            repo = etree.Element("repository-location")
            self._datasource.insert(0, repo)

        repo.set("derived-from", f"{directory}/{dbname}?rev=1.0")
        repo.set("id", dbname)
        repo.set("path", "/datasources")
        repo.set("revision", "1.0")

        # 3. Add <connection class='sqlproxy'>
        sqlproxy_conn = etree.Element("connection")
        channel = "https" if str(port) in ("443", "82") else "http"
        sqlproxy_conn.set("channel", channel)
        sqlproxy_conn.set("class", "sqlproxy")
        sqlproxy_conn.set("dbname", dbname)
        sqlproxy_conn.set("directory", directory)
        sqlproxy_conn.set("port", str(port))
        sqlproxy_conn.set("server", server)
        sqlproxy_conn.set("username", username)

        relation = etree.SubElement(sqlproxy_conn, "relation")
        relation.set("name", table_name)
        relation.set("table", f"[{table_name}]")
        relation.set("type", "table")

        # Insert after repository-location
        idx = list(self._datasource).index(repo)
        self._datasource.insert(idx + 1, sqlproxy_conn)

        # 4. Update <object-graph> relation
        for og_rel in self._datasource.findall(".//object-graph//relation"):
            if "connection" in og_rel.attrib:
                del og_rel.attrib["connection"]
            og_rel.set("name", table_name)
            og_rel.set("table", f"[{table_name}]")
            og_rel.set("type", "table")
            for cols in og_rel.findall("columns"):
                og_rel.remove(cols)
                
        # 5. Cleanup old fields and aliases
        for c in self._datasource.findall("column"):
            self._datasource.remove(c)
            
        aliases = self._datasource.find("aliases")
        if aliases is not None:
            self._datasource.remove(aliases)

        # 6. Clean metadata-records
        for mr in self._datasource.findall(".//metadata-record"):
            mr.getparent().remove(mr)

        self._reinit_fields()
        self.field_registry.set_unknown_field_policy(allow_unknown_fields=True)
        return f"Configured Tableau Server connection to {server}/{dbname} (table: {table_name})"

    def set_excel_connection(
        self,
        filepath: str,
        sheet_name: str = "",
        fields: list[dict[str, Any]] | None = None,
    ) -> str:
        """Configure the datasource to use a local Excel file."""

        fed_conn = self._datasource.find("connection[@class='federated']")
        if fed_conn is None:
            for old_conn in self._datasource.findall("connection"):
                self._datasource.remove(old_conn)
            fed_conn = etree.Element("connection")
            fed_conn.set("class", "federated")
            self._datasource.insert(0, fed_conn)

        named_conns = fed_conn.find("named-connections")
        if named_conns is None:
            named_conns = etree.SubElement(fed_conn, "named-connections")
        else:
            for child in list(named_conns):
                named_conns.remove(child)

        conn_name = f"excel-direct.{_generate_uuid().strip('{}').lower()}"
        nc = etree.SubElement(named_conns, "named-connection")
        nc.set("caption", os.path.basename(filepath))
        nc.set("name", conn_name)

        excel_conn = etree.SubElement(nc, "connection")
        excel_conn.set("class", "excel-direct")
        excel_conn.set("cleaning", "no")
        excel_conn.set("compat", "no")
        excel_conn.set("dataRefreshTime", "")
        excel_conn.set("filename", filepath.replace("\\", "/"))
        excel_conn.set("interpretationMode", "0")
        excel_conn.set("password", "")
        excel_conn.set("server", "")
        excel_conn.set("validate", "no")

        table_specs: list[dict[str, Any]] = []
        if not fields:
            table_specs = self._introspect_excel_tables(filepath, sheet_name)
        normalized_fields = self._normalize_external_fields(fields, source_object=sheet_name)
        if not normalized_fields and table_specs:
            sheet_name = table_specs[0]["name"]
        elif not sheet_name and normalized_fields:
            sheet_name = str(normalized_fields[0].get("source_object", "")).strip()

        if not sheet_name:
            sheet_name = "Sheet1"

        if table_specs and len(table_specs) > 1:
            self._rebuild_excel_multi_table_metadata(
                conn_name=conn_name,
                tables=table_specs,
            )
        else:
            if not normalized_fields:
                sheet_name, introspected_fields = self._introspect_excel_fields(filepath, sheet_name)
                normalized_fields = self._normalize_external_fields(
                    introspected_fields,
                    source_object=sheet_name,
                )

            relation = fed_conn.find("relation")
            if relation is None:
                relation = etree.SubElement(fed_conn, "relation")
            relation.set("connection", conn_name)
            relation.set("name", sheet_name)
            relation.set("table", f"[{sheet_name}$]")
            relation.set("type", "table")

            for og_rel in self._datasource.findall(".//object-graph//relation"):
                og_rel.set("connection", conn_name)
                og_rel.set("name", sheet_name)
                og_rel.set("table", f"[{sheet_name}$]")
                og_rel.set("type", "table")

            if normalized_fields:
                self._rebuild_external_datasource_metadata(
                    source_object=sheet_name,
                    fields=normalized_fields,
                    relation=relation,
                )
            else:
                self._reinit_fields()
                self.field_registry.set_unknown_field_policy(allow_unknown_fields=True)
        return f"Configured Excel connection to {filepath} (sheet: {sheet_name})"

    def set_csv_connection(
        self,
        filepath: str,
        delimiter: str = "",
        charset: str = "utf-8-sig",
        fields: list[dict[str, Any]] | None = None,
    ) -> str:
        """Configure the datasource to use a local CSV file."""

        fed_conn = self._datasource.find("connection[@class='federated']")
        if fed_conn is None:
            for old_conn in self._datasource.findall("connection"):
                self._datasource.remove(old_conn)
            fed_conn = etree.Element("connection")
            fed_conn.set("class", "federated")
            self._datasource.insert(0, fed_conn)

        named_conns = fed_conn.find("named-connections")
        if named_conns is None:
            named_conns = etree.SubElement(fed_conn, "named-connections")
        else:
            for child in list(named_conns):
                named_conns.remove(child)

        conn_name = f"textscan.{_generate_uuid().strip('{}').lower()}"
        csv_path = Path(filepath)
        source_filename = csv_path.name
        source_stem = csv_path.stem
        self._datasource.set("caption", source_stem)

        nc = etree.SubElement(named_conns, "named-connection")
        nc.set("caption", source_stem)
        nc.set("name", conn_name)

        csv_conn = etree.SubElement(nc, "connection")
        csv_conn.set("class", "textscan")
        csv_conn.set("directory", str(csv_path.parent.resolve()).replace("\\", "/"))
        csv_conn.set("filename", source_filename)
        csv_conn.set("password", "")
        csv_conn.set("server", "")
        actual_delimiter = delimiter or ","

        normalized_fields = self._normalize_external_fields(fields, source_object=source_filename)
        if not normalized_fields:
            source_filename, introspected_fields, actual_delimiter = self._introspect_csv_fields(
                filepath,
                delimiter=delimiter,
                charset=charset,
            )
            normalized_fields = self._normalize_external_fields(
                introspected_fields,
                source_object=source_filename,
            )
        relation = fed_conn.find("relation")
        if relation is None:
            relation = etree.SubElement(fed_conn, "relation")
        relation.set("connection", conn_name)
        relation.set("name", source_filename)
        relation.set("table", f"[{source_stem}#csv]")
        relation.set("type", "table")

        for og_rel in self._datasource.findall(".//object-graph//relation"):
            og_rel.set("connection", conn_name)
            og_rel.set("name", source_filename)
            og_rel.set("table", f"[{source_stem}#csv]")
            og_rel.set("type", "table")

        if normalized_fields:
            relation_columns_attrs = {
                "character-set": "UTF-8",
                "header": "yes",
                "locale": "zh_CN",
                "separator": actual_delimiter,
            }
            capability_attrs = [
                ("string", "character-set", "\"UTF-8\""),
                ("string", "collation", "\"zh_Hans_CN\""),
                ("string", "currency", "\"¥\""),
                ("string", "field-delimiter", f"\"{actual_delimiter}\""),
                ("string", "header-row", "\"true\""),
                ("string", "locale", "\"zh_CN\""),
                ("string", "single-char", "\"\""),
            ]
            self._rebuild_external_datasource_metadata(
                source_object=source_filename,
                fields=normalized_fields,
                relation=relation,
                prefer_existing_metadata=False,
                local_name_source_object="",
                relation_column_attrs_override=relation_columns_attrs,
                capability_attrs_override=capability_attrs,
            )
        else:
            self._reinit_fields()
            self.field_registry.set_unknown_field_policy(allow_unknown_fields=True)

        # When users only switch datasource and save immediately, the workbook
        # can be left without any worksheet/window nodes (depending on editor init mode),
        # which Tableau Desktop may fail to open.
        if self.root.find(".//worksheets/worksheet") is None:
            self.add_worksheet("Sheet 1")

        return f"Configured CSV connection to {filepath} (file: {source_filename})"

    def set_hyper_connection(
        self,
        filepath: str,
        table_name: str = "Extract",
        tables: Optional[List[dict]] = None,
    ) -> str:
        """Configure the datasource to use a local Hyper extract connection.

        Parameters
        ----------
        filepath : str
            Path to the ``.hyper`` file.
        table_name : str
            Table name for single-table mode (ignored when *tables* is given).
        tables : list[dict] | None
            For multi-table hyper files.  Each dict must have a ``"name"``
            key and may have an optional ``"columns"`` list of column-name
            strings.  The first entry is the *primary* table.
        """
        # 1. Update <connection class='federated'>
        fed_conn = self._datasource.find("connection[@class='federated']")
        if fed_conn is None:
            for old_conn in self._datasource.findall("connection"):
                self._datasource.remove(old_conn)
            fed_conn = etree.Element("connection")
            fed_conn.set("class", "federated")
            self._datasource.insert(0, fed_conn)

        # Update <named-connections>
        named_conns = fed_conn.find("named-connections")
        if named_conns is None:
            named_conns = etree.SubElement(fed_conn, "named-connections")
        else:
            for child in list(named_conns):
                named_conns.remove(child)

        conn_name = f"hyper.{_generate_uuid().strip('{}').lower()}"

        nc = etree.SubElement(named_conns, "named-connection")
        nc.set("caption", filepath.split("/")[-1].split("\\")[-1])
        nc.set("name", conn_name)

        hyper_conn = etree.SubElement(nc, "connection")
        hyper_conn.set("authentication", "auth-none")
        hyper_conn.set("author-locale", "en_US")
        hyper_conn.set("class", "hyper")
        hyper_conn.set("dbname", filepath)
        hyper_conn.set("default-settings", "yes")
        hyper_conn.set("schema", "Extract")
        hyper_conn.set("sslmode", "")
        hyper_conn.set("tablename", "Extract")
        hyper_conn.set("username", "")

        # Remove existing relation(s)
        for old_rel in fed_conn.findall("relation"):
            fed_conn.remove(old_rel)

        if tables and len(tables) > 1:
            # --- Multi-table mode ---
            self._set_hyper_multi_table(fed_conn, conn_name, tables)
        else:
            # --- Single-table mode (original behaviour) ---
            if tables and len(tables) == 1:
                table_name = tables[0]["name"]

            relation = etree.SubElement(fed_conn, "relation")
            relation.set("connection", conn_name)
            relation.set("name", table_name)
            relation.set("table", f"[Extract].[{table_name}]")
            relation.set("type", "table")

            # Update <object-graph> relation
            for og_rel in self._datasource.findall(".//object-graph//relation"):
                og_rel.set("connection", conn_name)
                og_rel.set("name", table_name)
                og_rel.set("table", f"[Extract].[{table_name}]")
                og_rel.set("type", "table")
                for cols in og_rel.findall("columns"):
                    og_rel.remove(cols)

        # Cleanup old generic/excel connections and leftover fields
        excel_conn = self._datasource.find("connection[@class='excel-direct']")
        if excel_conn is not None:
            self._datasource.remove(excel_conn)

        old_cols = fed_conn.find("cols")
        if old_cols is not None:
            fed_conn.remove(old_cols)

        for c in self._datasource.findall("column"):
            self._datasource.remove(c)

        aliases = self._datasource.find("aliases")
        if aliases is not None:
            self._datasource.remove(aliases)

        # Clean metadata-records
        for mr in self._datasource.findall(".//metadata-record"):
            mr.getparent().remove(mr)

        registered_fields: list[dict[str, Any]] = []
        if tables:
            for table in tables:
                for column in table.get("columns", []):
                    if isinstance(column, dict):
                        field_name = str(column.get("name", "")).strip()
                    else:
                        field_name = str(column).strip()
                    if not field_name:
                        continue
                    registered_fields.append(
                        {
                            "name": field_name,
                            "role": "dimension",
                            "field_type": "nominal",
                            "datatype": "string",
                        }
                    )

        self._register_external_fields(registered_fields)
        if tables and len(tables) > 1:
            names = ", ".join(t["name"] for t in tables)
            return f"Configured Hyper connection to {filepath} (tables: {names})"
        return f"Configured Hyper connection to {filepath} (table: {table_name})"

    # ------------------------------------------------------------------ #
    #  Multi-table helpers                                                #
    # ------------------------------------------------------------------ #

    def _set_hyper_multi_table(
        self,
        fed_conn: etree._Element,
        conn_name: str,
        tables: List[dict],
    ) -> None:
        """Build ``<relation type='collection'>`` for multi-table hyper files."""
        # -- Build the collection relation under fed_conn --
        collection = etree.SubElement(fed_conn, "relation")
        collection.set("type", "collection")
        for tbl in tables:
            child = etree.SubElement(collection, "relation")
            child.set("connection", conn_name)
            child.set("name", tbl["name"])
            child.set("table", f"[Extract].[{tbl['name']}]")
            child.set("type", "table")

        # -- Generate <cols> with <map> entries --
        primary = tables[0]
        primary_columns = set(primary.get("columns", []))

        cols_el = etree.SubElement(fed_conn, "cols")

        # Primary table maps: [Column] -> [PrimaryTable].[Column]
        for col_name in primary.get("columns", []):
            m = etree.SubElement(cols_el, "map")
            m.set("key", f"[{col_name}]")
            m.set("value", f"[{primary['name']}].[{col_name}]")

        # Non-primary tables
        for tbl in tables[1:]:
            for col_name in tbl.get("columns", []):
                m = etree.SubElement(cols_el, "map")
                if col_name in primary_columns:
                    # Overlapping column: suffix with (table_name)
                    m.set("key", f"[{col_name} ({tbl['name']})]")
                else:
                    m.set("key", f"[{col_name}]")
                m.set("value", f"[{tbl['name']}].[{col_name}]")

        # -- Update <object-graph> relations --
        for og_rel in self._datasource.findall(".//object-graph//relation"):
            old_name = og_rel.get("name", "")
            best_match = None
            # Find best match using exact or split-prefix
            for tbl in tables:
                base_name = tbl["name"].split("_")[0]
                if old_name == tbl["name"] or base_name in old_name:
                    best_match = tbl
                    break
            
            # Fallback
            if not best_match and tables:
                best_match = tables[0]
                
            if best_match:
                og_rel.set("connection", conn_name)
                og_rel.set("name", best_match["name"])
                og_rel.set("table", f"[Extract].[{best_match['name']}]")
                og_rel.set("type", "table")

